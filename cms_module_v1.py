"""
CMS Builder v1 — Portal TPQ HMarisa
====================================
Memungkinkan Admin membuat modul data baru melalui form tanpa menulis kode.
Setiap modul otomatis mendapat: halaman daftar, tambah, edit, hapus,
ekspor CSV/Excel, impor, download template, dan entri sidebar.

Dipasang via install_cms_module_v1(app, db, app_globals) dari app.py.
"""

import csv
import io
import json
import math
import os
import re
from datetime import datetime
from functools import wraps

from flask import (
    abort, flash, redirect, render_template, request,
    send_file, url_for
)
from flask_login import current_user, login_required
from sqlalchemy import inspect, text

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

CMS_PER_PAGE = 20
_GRADIENT_COLORS = [
    "linear-gradient(135deg,#075f46,#10b981)",
    "linear-gradient(135deg,#1d4ed8,#60a5fa)",
    "linear-gradient(135deg,#7c3aed,#a78bfa)",
    "linear-gradient(135deg,#b45309,#fbbf24)",
    "linear-gradient(135deg,#be185d,#f472b6)",
    "linear-gradient(135deg,#0e7490,#22d3ee)",
    "linear-gradient(135deg,#dc2626,#f87171)",
    "linear-gradient(135deg,#15803d,#4ade80)",
]

FIELD_TYPES = [
    ("text",     "Teks Pendek"),
    ("textarea", "Teks Panjang"),
    ("number",   "Angka"),
    ("date",     "Tanggal"),
    ("datetime", "Tanggal & Waktu"),
    ("select",   "Pilihan (Dropdown)"),
    ("boolean",  "Ya / Tidak"),
    ("email",    "Email"),
    ("phone",    "Telepon"),
    ("file",     "File / Lampiran"),
]

ALL_FEATURES = [
    ("add",       "Tambah",           "fa-plus"),
    ("edit",      "Edit",             "fa-pen"),
    ("delete",    "Hapus",            "fa-trash"),
    ("import",    "Import",           "fa-file-import"),
    ("export",    "Export",           "fa-file-export"),
    ("template",  "Download Template","fa-file-arrow-down"),
    ("upload",    "Upload File",      "fa-upload"),
    ("print",     "Cetak",            "fa-print"),
    ("search",    "Pencarian",        "fa-magnifying-glass"),
    ("filter",    "Filter",           "fa-filter"),
    ("pagination","Pagination",       "fa-table-list"),
]


def _slugify(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s[:60] or "modul"


def install_cms_module_v1(app, db, app_globals):
    """Daftarkan CMS Builder ke Flask app."""

    # ── Models ──────────────────────────────────────────────────────────────

    class CmsModule(db.Model):
        __tablename__ = "cms_module"
        __table_args__ = {"extend_existing": True}
        id           = db.Column(db.Integer, primary_key=True)
        name         = db.Column(db.String(100), nullable=False)
        slug         = db.Column(db.String(80), unique=True, nullable=False, index=True)
        icon         = db.Column(db.String(80), default="fa-table-list")
        color        = db.Column(db.Integer, default=0)   # index into _GRADIENT_COLORS
        description  = db.Column(db.Text, default="")
        purpose      = db.Column(db.Text, default="")
        role_access  = db.Column(db.String(200), default="admin_utama,admin")
        features     = db.Column(db.Text, default="{}")   # JSON
        parent_menu_id = db.Column(db.Integer, db.ForeignKey("sidebar_menu.id", ondelete="SET NULL"), nullable=True)
        sort_order   = db.Column(db.Integer, default=0)
        is_active    = db.Column(db.Boolean, default=True)
        created_by   = db.Column(db.String(80), default="")
        created_at   = db.Column(db.DateTime, default=datetime.utcnow)
        updated_at   = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

        fields_rel = db.relationship(
            "CmsField", backref="module", order_by="CmsField.sort_order",
            cascade="all, delete-orphan", lazy="select"
        )

        def roles_list(self):
            return [r.strip() for r in self.role_access.split(",") if r.strip()]

        def features_dict(self):
            try:
                return json.loads(self.features or "{}")
            except Exception:
                return {}

        def has_feature(self, key):
            return self.features_dict().get(key, False)

        def gradient(self):
            return _GRADIENT_COLORS[self.color % len(_GRADIENT_COLORS)]

        def active_fields(self):
            return [f for f in self.fields_rel if f.is_active]

    class CmsField(db.Model):
        __tablename__ = "cms_field"
        __table_args__ = {"extend_existing": True}
        id          = db.Column(db.Integer, primary_key=True)
        module_id   = db.Column(db.Integer, db.ForeignKey("cms_module.id", ondelete="CASCADE"), nullable=False, index=True)
        name        = db.Column(db.String(80), nullable=False)
        label       = db.Column(db.String(120), nullable=False)
        field_type  = db.Column(db.String(30), default="text")
        options_raw = db.Column(db.Text, default="")   # comma-separated for select
        placeholder = db.Column(db.String(200), default="")
        required    = db.Column(db.Boolean, default=False)
        searchable  = db.Column(db.Boolean, default=True)
        exportable  = db.Column(db.Boolean, default=True)
        sort_order  = db.Column(db.Integer, default=0)
        is_active   = db.Column(db.Boolean, default=True)

        def options_list(self):
            if not self.options_raw:
                return []
            return [o.strip() for o in self.options_raw.split(",") if o.strip()]

    class CmsRecord(db.Model):
        __tablename__ = "cms_record"
        __table_args__ = {"extend_existing": True}
        id         = db.Column(db.Integer, primary_key=True)
        module_id  = db.Column(db.Integer, db.ForeignKey("cms_module.id", ondelete="CASCADE"), nullable=False, index=True)
        data       = db.Column(db.Text, default="{}")   # JSON
        is_deleted = db.Column(db.Boolean, default=False)
        deleted_at = db.Column(db.DateTime, nullable=True)
        deleted_by = db.Column(db.String(80), default="")
        created_by = db.Column(db.String(80), default="")
        created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
        updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

        def data_dict(self):
            try:
                return json.loads(self.data or "{}")
            except Exception:
                return {}

        def get(self, key, default=""):
            return self.data_dict().get(key, default)

    # Expose to app
    app_globals["CmsModule"] = CmsModule
    app_globals["CmsField"] = CmsField
    app_globals["CmsRecord"] = CmsRecord

    # Create tables
    with app.app_context():
        db.create_all()
        # Add columns that may be missing (schema migration)
        try:
            existing = {c["name"] for c in inspect(db.engine).get_columns("cms_module")}
            with db.engine.connect() as conn:
                if "purpose" not in existing:
                    conn.execute(text("ALTER TABLE cms_module ADD COLUMN purpose TEXT DEFAULT ''"))
                if "color" not in existing:
                    conn.execute(text("ALTER TABLE cms_module ADD COLUMN color INTEGER DEFAULT 0"))
                conn.commit()
        except Exception:
            pass

    # ── Access helpers ───────────────────────────────────────────────────────

    def _admin_required(view):
        @wraps(view)
        @login_required
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("admin_login"))
            role = getattr(current_user, "role", "")
            if role not in ("admin_utama", "admin"):
                abort(403)
            return view(*args, **kwargs)
        return wrapped

    def _module_access_required(module):
        """Check if current user has access to this module."""
        if not current_user.is_authenticated:
            return False
        role = getattr(current_user, "role", "")
        return role in module.roles_list()

    def _get_sidebar_model():
        return app_globals.get("SidebarMenu")

    # ── Sidebar auto-registration ────────────────────────────────────────────

    def _register_sidebar(module):
        SidebarMenu = _get_sidebar_model()
        if not SidebarMenu:
            return
        try:
            existing = SidebarMenu.query.filter_by(
                endpoint="cms_module_list",
                url="/m/" + module.slug
            ).first()
            if existing:
                existing.label = module.name
                existing.icon = module.icon
                existing.roles = module.role_access
                existing.sort_order = module.sort_order or 200
                existing.parent_id = module.parent_menu_id
                db.session.commit()
                return
            entry = SidebarMenu(
                label=module.name,
                icon=module.icon,
                endpoint="cms_module_list",
                url="/m/" + module.slug,
                active_endpoints="cms_module_list,cms_record_add,cms_record_edit",
                parent_id=module.parent_menu_id,
                sort_order=module.sort_order or 200,
                roles=module.role_access,
                is_active=module.is_active,
                is_system=False,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )
            db.session.add(entry)
            db.session.commit()
        except Exception as e:
            db.session.rollback()

    def _unregister_sidebar(module_slug):
        SidebarMenu = _get_sidebar_model()
        if not SidebarMenu:
            return
        try:
            entries = SidebarMenu.query.filter_by(url="/m/" + module_slug).all()
            for e in entries:
                db.session.delete(e)
            db.session.commit()
        except Exception:
            db.session.rollback()

    # ── Export helpers ───────────────────────────────────────────────────────

    def _build_csv(module, records):
        fields = module.active_fields()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["No"] + [f.label for f in fields if f.exportable])
        for i, rec in enumerate(records, 1):
            d = rec.data_dict()
            writer.writerow([i] + [d.get(f.name, "") for f in fields if f.exportable])
        output.seek(0)
        return output

    def _build_xlsx(module, records):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = module.name[:31]
        fields = [f for f in module.active_fields() if f.exportable]
        headers = ["No"] + [f.label for f in fields]
        ws.append(headers)
        for i, rec in enumerate(records, 1):
            d = rec.data_dict()
            ws.append([i] + [d.get(f.name, "") for f in fields])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _build_template_xlsx(module):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = module.name[:31]
        fields = [f for f in module.active_fields() if f.exportable]
        ws.append([f.label for f in fields])
        # Add one example row
        ws.append(["Contoh " + f.label for f in fields])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Import helpers ───────────────────────────────────────────────────────

    def _parse_import_file(module, file_obj, filename):
        fields = [f for f in module.active_fields() if f.exportable]
        rows = []
        errors = []
        ext = os.path.splitext(filename)[1].lower()

        if ext == ".csv":
            try:
                content = file_obj.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(content))
                label_to_name = {f.label: f.name for f in fields}
                for idx, row in enumerate(reader, 2):
                    record = {}
                    for label, val in row.items():
                        name = label_to_name.get(label)
                        if name:
                            record[name] = val.strip() if val else ""
                    rows.append(record)
            except Exception as e:
                errors.append(f"Gagal membaca CSV: {e}")
        elif ext in (".xlsx", ".xls") and _HAS_OPENPYXL:
            try:
                wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
                ws = wb.active
                headers = []
                label_to_name = {f.label: f.name for f in fields}
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if idx == 0:
                        headers = [str(c) if c else "" for c in row]
                        continue
                    if all(c is None for c in row):
                        continue
                    record = {}
                    for col_i, val in enumerate(row):
                        if col_i < len(headers):
                            name = label_to_name.get(headers[col_i])
                            if name:
                                record[name] = str(val) if val is not None else ""
                    rows.append(record)
                wb.close()
            except Exception as e:
                errors.append(f"Gagal membaca Excel: {e}")
        else:
            errors.append("Format file tidak didukung. Gunakan .csv atau .xlsx")

        return rows, errors

    # ═══════════════════════════════════════════════════════════════════════
    # ROUTES — CMS Builder Admin
    # ═══════════════════════════════════════════════════════════════════════

    @app.route("/control-center/cms-builder")
    @_admin_required
    def cms_builder_index():
        modules = CmsModule.query.order_by(CmsModule.sort_order, CmsModule.name).all()
        return render_template(
            "cms_builder/builder.html",
            modules=modules,
            all_features=ALL_FEATURES,
            gradient_colors=_GRADIENT_COLORS,
        )

    @app.route("/control-center/cms-builder/new", methods=["GET", "POST"])
    @_admin_required
    def cms_builder_new():
        SidebarMenu = _get_sidebar_model()
        parent_menus = []
        if SidebarMenu:
            parent_menus = SidebarMenu.query.filter(
                SidebarMenu.parent_id.is_(None),
                SidebarMenu.is_active == True,
            ).order_by(SidebarMenu.sort_order).all()

        if request.method == "POST":
            name = request.form.get("name", "").strip()
            slug = _slugify(request.form.get("slug", name))
            icon = request.form.get("icon", "fa-table-list").strip()
            color = int(request.form.get("color", 0))
            description = request.form.get("description", "").strip()
            purpose = request.form.get("purpose", "").strip()
            role_access = ",".join(request.form.getlist("role_access")) or "admin_utama,admin"
            sort_order = int(request.form.get("sort_order", 200) or 200)
            parent_menu_id = request.form.get("parent_menu_id") or None
            if parent_menu_id:
                parent_menu_id = int(parent_menu_id)

            # Features
            feat = {}
            for key, _, _ in ALL_FEATURES:
                feat[key] = bool(request.form.get("feat_" + key))
            feat["search"] = True   # always enabled
            feat["pagination"] = True

            # Validate slug uniqueness
            if not slug:
                flash("Nama modul tidak boleh kosong.", "danger")
                return render_template("cms_builder/module_form.html",
                                       mode="new", form=request.form,
                                       all_features=ALL_FEATURES,
                                       parent_menus=parent_menus,
                                       gradient_colors=_GRADIENT_COLORS,
                                       field_types=FIELD_TYPES)
            if CmsModule.query.filter_by(slug=slug).first():
                flash(f"Slug '{slug}' sudah dipakai. Ganti nama modul.", "danger")
                return render_template("cms_builder/module_form.html",
                                       mode="new", form=request.form,
                                       all_features=ALL_FEATURES,
                                       parent_menus=parent_menus,
                                       gradient_colors=_GRADIENT_COLORS,
                                       field_types=FIELD_TYPES)

            module = CmsModule(
                name=name, slug=slug, icon=icon, color=color,
                description=description, purpose=purpose,
                role_access=role_access, features=json.dumps(feat),
                parent_menu_id=parent_menu_id, sort_order=sort_order,
                created_by=current_user.username,
            )
            db.session.add(module)
            db.session.flush()

            # Parse fields
            field_names = request.form.getlist("field_name[]")
            field_labels = request.form.getlist("field_label[]")
            field_types_ = request.form.getlist("field_type[]")
            field_required = request.form.getlist("field_required[]")
            field_options = request.form.getlist("field_options[]")

            for i, (fn, fl, ft) in enumerate(zip(field_names, field_labels, field_types_)):
                fn = fn.strip(); fl = fl.strip()
                if not fn or not fl:
                    continue
                f = CmsField(
                    module_id=module.id,
                    name=fn, label=fl, field_type=ft,
                    options_raw=field_options[i] if i < len(field_options) else "",
                    required=("1" in (field_required[i] if i < len(field_required) else "")),
                    searchable=True, exportable=True,
                    sort_order=i,
                )
                db.session.add(f)

            db.session.commit()
            _register_sidebar(module)
            flash(f"Modul '{name}' berhasil dibuat. Sidebar sudah diperbarui.", "success")
            return redirect(url_for("cms_builder_index"))

        return render_template(
            "cms_builder/module_form.html",
            mode="new", form={},
            all_features=ALL_FEATURES,
            parent_menus=parent_menus,
            gradient_colors=_GRADIENT_COLORS,
            field_types=FIELD_TYPES,
        )

    @app.route("/control-center/cms-builder/<int:module_id>/edit", methods=["GET", "POST"])
    @_admin_required
    def cms_builder_edit(module_id):
        module = CmsModule.query.get_or_404(module_id)
        SidebarMenu = _get_sidebar_model()
        parent_menus = []
        if SidebarMenu:
            parent_menus = SidebarMenu.query.filter(
                SidebarMenu.parent_id.is_(None),
                SidebarMenu.is_active == True,
            ).order_by(SidebarMenu.sort_order).all()

        if request.method == "POST":
            module.name = request.form.get("name", module.name).strip()
            module.icon = request.form.get("icon", module.icon).strip()
            module.color = int(request.form.get("color", module.color))
            module.description = request.form.get("description", "").strip()
            module.purpose = request.form.get("purpose", "").strip()
            module.role_access = ",".join(request.form.getlist("role_access")) or "admin_utama,admin"
            module.sort_order = int(request.form.get("sort_order", 200) or 200)
            parent_id = request.form.get("parent_menu_id") or None
            module.parent_menu_id = int(parent_id) if parent_id else None

            feat = {}
            for key, _, _ in ALL_FEATURES:
                feat[key] = bool(request.form.get("feat_" + key))
            feat["search"] = True
            feat["pagination"] = True
            module.features = json.dumps(feat)
            module.updated_at = datetime.utcnow()

            db.session.commit()
            _register_sidebar(module)
            flash(f"Modul '{module.name}' berhasil diperbarui.", "success")
            return redirect(url_for("cms_builder_index"))

        return render_template(
            "cms_builder/module_form.html",
            mode="edit", module=module,
            form={},
            all_features=ALL_FEATURES,
            parent_menus=parent_menus,
            gradient_colors=_GRADIENT_COLORS,
            field_types=FIELD_TYPES,
        )

    @app.route("/control-center/cms-builder/<int:module_id>/delete", methods=["POST"])
    @_admin_required
    def cms_builder_delete(module_id):
        module = CmsModule.query.get_or_404(module_id)
        slug = module.slug
        name = module.name
        record_count = CmsRecord.query.filter_by(module_id=module_id, is_deleted=False).count()
        if record_count > 0 and request.form.get("force") != "1":
            flash(f"Modul '{name}' masih memiliki {record_count} data. Centang 'Paksa Hapus' untuk melanjutkan.", "warning")
            return redirect(url_for("cms_builder_index"))
        _unregister_sidebar(slug)
        db.session.delete(module)
        db.session.commit()
        flash(f"Modul '{name}' berhasil dihapus.", "success")
        return redirect(url_for("cms_builder_index"))

    @app.route("/control-center/cms-builder/<int:module_id>/toggle", methods=["POST"])
    @_admin_required
    def cms_builder_toggle(module_id):
        module = CmsModule.query.get_or_404(module_id)
        module.is_active = not module.is_active
        db.session.commit()
        # Sync sidebar
        SidebarMenu = _get_sidebar_model()
        if SidebarMenu:
            entry = SidebarMenu.query.filter_by(url="/m/" + module.slug).first()
            if entry:
                entry.is_active = module.is_active
                db.session.commit()
        status = "diaktifkan" if module.is_active else "dinonaktifkan"
        flash(f"Modul '{module.name}' berhasil {status}.", "success")
        return redirect(url_for("cms_builder_index"))

    # ═══════════════════════════════════════════════════════════════════════
    # ROUTES — Generic Module Pages
    # ═══════════════════════════════════════════════════════════════════════

    def _get_module_by_slug(slug):
        m = CmsModule.query.filter_by(slug=slug, is_active=True).first_or_404()
        if not _module_access_required(m):
            abort(403)
        return m

    @app.route("/m/<slug>")
    @login_required
    def cms_module_list(slug):
        module = _get_module_by_slug(slug)
        q = request.args.get("q", "").strip()
        page = int(request.args.get("page", 1))
        status_filter = request.args.get("status", "active")

        query = CmsRecord.query.filter_by(module_id=module.id)
        if status_filter == "active":
            query = query.filter_by(is_deleted=False)
        elif status_filter == "deleted":
            query = query.filter_by(is_deleted=True)

        # Search across searchable fields
        if q:
            matching_ids = []
            all_recs = query.all()
            for rec in all_recs:
                d = rec.data_dict()
                for field in module.active_fields():
                    if field.searchable and q.lower() in str(d.get(field.name, "")).lower():
                        matching_ids.append(rec.id)
                        break
            query = CmsRecord.query.filter(CmsRecord.id.in_(matching_ids))

        total = query.count()
        total_pages = max(1, math.ceil(total / CMS_PER_PAGE))
        page = min(max(1, page), total_pages)
        records = query.order_by(CmsRecord.created_at.desc()) \
                       .offset((page - 1) * CMS_PER_PAGE).limit(CMS_PER_PAGE).all()

        # Sibling submenus for page-level tabs
        SidebarMenu = _get_sidebar_model()
        sibling_tabs = []
        if SidebarMenu and module.parent_menu_id:
            siblings = SidebarMenu.query.filter_by(
                parent_id=module.parent_menu_id, is_active=True
            ).order_by(SidebarMenu.sort_order).all()
            for s in siblings:
                sibling_tabs.append({
                    "label": s.label,
                    "url": s.url or (url_for(s.endpoint) if s.endpoint else "#"),
                    "icon": s.icon,
                })

        return render_template(
            "cms_builder/record_list.html",
            module=module,
            records=records,
            fields=module.active_fields(),
            q=q, page=page, total_pages=total_pages, total=total,
            status_filter=status_filter,
            sibling_tabs=sibling_tabs,
        )

    @app.route("/m/<slug>/new", methods=["GET", "POST"])
    @login_required
    def cms_record_add(slug):
        module = _get_module_by_slug(slug)
        if not module.has_feature("add"):
            abort(403)

        errors = {}
        form_data = {}
        if request.method == "POST":
            form_data = {f.name: request.form.get(f.name, "").strip() for f in module.active_fields()}
            # Validate required
            for field in module.active_fields():
                if field.required and not form_data.get(field.name):
                    errors[field.name] = f"{field.label} wajib diisi."

            if not errors:
                record = CmsRecord(
                    module_id=module.id,
                    data=json.dumps(form_data),
                    created_by=current_user.username,
                )
                db.session.add(record)
                db.session.commit()
                flash(f"Data berhasil ditambahkan ke {module.name}.", "success")
                return redirect(url_for("cms_module_list", slug=slug))

        return render_template(
            "cms_builder/record_form.html",
            module=module, mode="add",
            fields=module.active_fields(),
            form_data=form_data, errors=errors,
            field_types=dict(FIELD_TYPES),
        )

    @app.route("/m/<slug>/<int:record_id>/edit", methods=["GET", "POST"])
    @login_required
    def cms_record_edit(slug, record_id):
        module = _get_module_by_slug(slug)
        if not module.has_feature("edit"):
            abort(403)
        record = CmsRecord.query.filter_by(id=record_id, module_id=module.id).first_or_404()

        errors = {}
        form_data = record.data_dict()
        if request.method == "POST":
            form_data = {f.name: request.form.get(f.name, "").strip() for f in module.active_fields()}
            for field in module.active_fields():
                if field.required and not form_data.get(field.name):
                    errors[field.name] = f"{field.label} wajib diisi."
            if not errors:
                record.data = json.dumps(form_data)
                record.updated_at = datetime.utcnow()
                db.session.commit()
                flash("Data berhasil diperbarui.", "success")
                return redirect(url_for("cms_module_list", slug=slug))

        return render_template(
            "cms_builder/record_form.html",
            module=module, mode="edit",
            record=record, fields=module.active_fields(),
            form_data=form_data, errors=errors,
            field_types=dict(FIELD_TYPES),
        )

    @app.route("/m/<slug>/<int:record_id>/delete", methods=["POST"])
    @login_required
    def cms_record_delete(slug, record_id):
        module = _get_module_by_slug(slug)
        if not module.has_feature("delete"):
            abort(403)
        record = CmsRecord.query.filter_by(id=record_id, module_id=module.id).first_or_404()
        delete_type = request.form.get("delete_type", "soft")
        if delete_type == "permanent":
            db.session.delete(record)
            flash("Data berhasil dihapus permanen.", "success")
        else:
            record.is_deleted = True
            record.deleted_at = datetime.utcnow()
            record.deleted_by = current_user.username
            flash("Data berhasil dinonaktifkan.", "success")
        db.session.commit()
        return redirect(url_for("cms_module_list", slug=slug))

    @app.route("/m/<slug>/export.csv")
    @login_required
    def cms_module_export_csv(slug):
        module = _get_module_by_slug(slug)
        if not module.has_feature("export"):
            abort(403)
        records = CmsRecord.query.filter_by(module_id=module.id, is_deleted=False) \
                                  .order_by(CmsRecord.created_at).all()
        output = _build_csv(module, records)
        filename = f"{slug}_{datetime.now().strftime('%Y%m%d')}.csv"
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename,
        )

    @app.route("/m/<slug>/export.xlsx")
    @login_required
    def cms_module_export_xlsx(slug):
        module = _get_module_by_slug(slug)
        if not module.has_feature("export") or not _HAS_OPENPYXL:
            abort(403)
        records = CmsRecord.query.filter_by(module_id=module.id, is_deleted=False) \
                                  .order_by(CmsRecord.created_at).all()
        buf = _build_xlsx(module, records)
        filename = f"{slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(
            buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename,
        )

    @app.route("/m/<slug>/template.xlsx")
    @login_required
    def cms_module_template(slug):
        module = _get_module_by_slug(slug)
        if not module.has_feature("template") or not _HAS_OPENPYXL:
            abort(403)
        buf = _build_template_xlsx(module)
        filename = f"template_{slug}.xlsx"
        return send_file(
            buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename,
        )

    @app.route("/m/<slug>/import", methods=["GET", "POST"])
    @login_required
    def cms_module_import(slug):
        module = _get_module_by_slug(slug)
        if not module.has_feature("import"):
            abort(403)

        if request.method == "POST":
            f = request.files.get("import_file")
            if not f or not f.filename:
                flash("Pilih file terlebih dahulu.", "warning")
                return redirect(url_for("cms_module_list", slug=slug))

            rows, errors = _parse_import_file(module, f, f.filename)
            if errors:
                for e in errors:
                    flash(e, "danger")
                return redirect(url_for("cms_module_list", slug=slug))

            ok = 0
            fail = 0
            for row in rows:
                # Validate required fields
                row_ok = True
                for field in module.active_fields():
                    if field.required and not row.get(field.name):
                        fail += 1
                        row_ok = False
                        break
                if row_ok:
                    record = CmsRecord(
                        module_id=module.id,
                        data=json.dumps(row),
                        created_by=current_user.username + " (import)",
                    )
                    db.session.add(record)
                    ok += 1

            db.session.commit()
            flash(f"Import selesai: {ok} berhasil, {fail} gagal (field wajib kosong).", "success" if fail == 0 else "warning")
            return redirect(url_for("cms_module_list", slug=slug))

        return redirect(url_for("cms_module_list", slug=slug))

    # ── Module Hub Page (parent menu with children as cards) ─────────────────

    @app.route("/hub/<int:menu_id>")
    @login_required
    def cms_module_hub(menu_id):
        """Halaman hub untuk parent menu — menampilkan submenu sebagai kartu."""
        SidebarMenu = _get_sidebar_model()
        if not SidebarMenu:
            abort(404)
        parent = SidebarMenu.query.get_or_404(menu_id)
        children = SidebarMenu.query.filter_by(
            parent_id=menu_id, is_active=True
        ).order_by(SidebarMenu.sort_order).all()
        return render_template(
            "cms_builder/module_hub.html",
            parent=parent,
            children=children,
        )

    app_globals["install_cms_module_v1"] = True
