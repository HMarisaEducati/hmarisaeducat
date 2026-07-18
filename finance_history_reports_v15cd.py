"""Keuangan V15-C/D — Riwayat Pembayaran dan Laporan administrasi iuran.

Modul ini dipasang setelah V15-A, V15-B, dan V15-B2. Tidak membuat sistem
akuntansi; seluruh keluaran hanya berasal dari tagihan dan pembayaran santri.
"""
from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO
from pathlib import Path
from typing import Any

from flask import abort, flash, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required
from sqlalchemy import func


PAGE_SIZES = {25, 50, 100}
PAYMENT_METHODS = ["Tunai", "Transfer", "Lainnya", "Migrasi Data Lama"]
REPORT_TYPES = {
    "all_bills": "Rekap Seluruh Tagihan",
    "paid": "Daftar Santri Lunas",
    "partial": "Daftar Santri Sebagian",
    "unpaid": "Daftar Santri Belum Lunas",
    "waived": "Daftar Santri Dibebaskan",
    "payments": "Riwayat Pembayaran",
    "class_summary": "Rekap Pembayaran per Kelas",
    "monthly_summary": "Rekap Pembayaran per Bulan",
}
STATUS_BY_REPORT = {
    "paid": "Lunas",
    "partial": "Sebagian",
    "unpaid": "Belum Lunas",
    "waived": "Dibebaskan",
}
MONTH_ORDER = {
    "Januari": 1, "Februari": 2, "Maret": 3, "April": 4,
    "Mei": 5, "Juni": 6, "Juli": 7, "Agustus": 8,
    "September": 9, "Oktober": 10, "November": 11, "Desember": 12,
}


def install_finance_history_reports_v15cd(app, db, namespace: dict[str, Any]):
    """Aktifkan Riwayat Pembayaran dan Laporan tanpa mengubah tabel lama."""
    if app.extensions.get("finance_history_reports_v15cd"):
        return app.extensions["finance_history_reports_v15cd"]

    v15a = app.extensions.get("finance_v15a")
    v15b = app.extensions.get("finance_v15b")
    v15b2 = app.extensions.get("finance_administration_v15b2")
    if not v15a or not v15b or not v15b2:
        raise RuntimeError("Keuangan V15-A, V15-B, dan V15-B2 harus terpasang sebelum V15-C/D.")

    User = namespace["User"]
    Santri = namespace["Santri"]
    CLASSES = namespace.get("CLASSES", [])
    MONTHS = namespace.get("MONTHS", list(MONTH_ORDER))
    upload_root = Path(namespace.get("UPLOAD_DIR") or (Path(app.root_path) / "uploads"))

    models_a = v15a["models"]
    GeneralSetting = models_a["general"]
    ChargeType = models_a["charge_type"]
    PaymentChannel = models_a["payment_channel"]
    AuditLog = models_a["audit"]

    models_b = v15b["models"]
    FinanceBill = models_b["bill"]

    models_b2 = v15b2["models"]
    FinancePayment = models_b2["payment"]

    def _finance_admin_required(view):
        @wraps(view)
        @login_required
        def wrapped(*args, **kwargs):
            if not current_user.is_admin or current_user.is_teacher or not getattr(current_user, "is_active", True):
                abort(403)
            return view(*args, **kwargs)
        return wrapped

    def _clean(value: Any, limit: int = 255) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip())[:limit]

    def _integer(value: Any, default: int, minimum: int = 1, maximum: int = 100000) -> int:
        try:
            result = int(str(value or default))
        except (TypeError, ValueError):
            result = default
        return min(max(result, minimum), maximum)

    def _parse_date(value: Any) -> date | None:
        raw = _clean(value, 20)
        if not raw:
            return None
        try:
            return date.fromisoformat(raw)
        except ValueError:
            return None

    def _format_date(value: date | None) -> str:
        return value.strftime("%d/%m/%Y") if value else "Belum diisi"

    def _format_wib(value: datetime | None) -> str:
        if not value:
            return "Belum diisi"
        return (value + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M WIB")

    def _rupiah(value: Any) -> str:
        try:
            return "Rp{:,.0f}".format(int(value or 0)).replace(",", ".")
        except (TypeError, ValueError):
            return "Rp0"

    def _active_general():
        general, _, _, _ = v15a["ensure_defaults"]()
        return general

    def _academic_year_options(active: str) -> list[str]:
        options = [f"{year}/{year + 1}" for year in range(2020, 2050)]
        if active and active not in options:
            options.append(active)
        return sorted(set(options), key=lambda item: item[:4])

    def _pager(total: int, page: int, per_page: int) -> dict[str, Any]:
        pages = max(1, (total + per_page - 1) // per_page)
        page = min(max(page, 1), pages)
        return {
            "page": page,
            "per_page": per_page,
            "total": total,
            "pages": pages,
            "has_prev": page > 1,
            "has_next": page < pages,
            "prev_num": page - 1,
            "next_num": page + 1,
        }

    def _payment_filters() -> dict[str, Any]:
        general = _active_general()
        status = _clean(request.args.get("transaction_status") or "active", 20)
        if status not in {"all", "active", "cancelled"}:
            status = "active"
        return {
            "q": _clean(request.args.get("q"), 100),
            "academic_year": _clean(request.args.get("academic_year") if "academic_year" in request.args else general.academic_year_active, 20),
            "semester": _clean(request.args.get("semester") if "semester" in request.args else general.semester_active, 20),
            "class_name": _clean(request.args.get("class_name"), 60),
            "student_name": _clean(request.args.get("student_name"), 140),
            "charge_type_id": _clean(request.args.get("charge_type_id"), 20),
            "period": _clean(request.args.get("period"), 80),
            "method": _clean(request.args.get("method"), 40),
            "date_from": _parse_date(request.args.get("date_from")),
            "date_to": _parse_date(request.args.get("date_to")),
            "transaction_status": status,
            "sort": _clean(request.args.get("sort") or "newest", 20),
        }

    def _payment_query(filters: dict[str, Any]):
        query = db.session.query(FinancePayment, FinanceBill, Santri, ChargeType).join(
            FinanceBill, FinanceBill.id == FinancePayment.bill_id
        ).join(Santri, Santri.id == FinanceBill.santri_id).join(
            ChargeType, ChargeType.id == FinanceBill.charge_type_id
        ).filter(FinanceBill.is_archived.is_(False))

        if filters["q"]:
            like = f"%{filters['q']}%"
            query = query.filter(FinancePayment.transaction_number.ilike(like))
        if filters["academic_year"]:
            query = query.filter(FinanceBill.academic_year == filters["academic_year"])
        if filters["semester"]:
            query = query.filter(FinanceBill.semester == filters["semester"])
        if filters["class_name"]:
            query = query.filter(Santri.class_name == filters["class_name"])
        if filters["student_name"]:
            query = query.filter(Santri.name.ilike(f"%{filters['student_name']}%"))
        if filters["charge_type_id"].isdigit():
            query = query.filter(FinanceBill.charge_type_id == int(filters["charge_type_id"]))
        if filters["period"]:
            query = query.filter(FinanceBill.period_label == filters["period"])
        if filters["method"]:
            query = query.filter(FinancePayment.method == filters["method"])
        if filters["date_from"]:
            query = query.filter(FinancePayment.payment_date >= filters["date_from"])
        if filters["date_to"]:
            query = query.filter(FinancePayment.payment_date <= filters["date_to"])
        if filters["transaction_status"] == "active":
            query = query.filter(FinancePayment.is_cancelled.is_(False))
        elif filters["transaction_status"] == "cancelled":
            query = query.filter(FinancePayment.is_cancelled.is_(True))

        if filters["sort"] == "oldest":
            query = query.order_by(FinancePayment.payment_date.asc(), FinancePayment.id.asc())
        elif filters["sort"] == "amount_high":
            query = query.order_by(FinancePayment.amount.desc(), FinancePayment.id.desc())
        elif filters["sort"] == "name":
            query = query.order_by(Santri.name.asc(), FinancePayment.payment_date.desc())
        else:
            query = query.order_by(FinancePayment.payment_date.desc(), FinancePayment.id.desc())
        return query

    @_finance_admin_required
    def finance_payments():
        filters = _payment_filters()
        page = _integer(request.args.get("page"), 1)
        per_page = _integer(request.args.get("per_page"), 25, 1, 100)
        if per_page not in PAGE_SIZES:
            per_page = 25
        query = _payment_query(filters)
        total = query.count()
        pager = _pager(total, page, per_page)
        rows = query.offset((pager["page"] - 1) * per_page).limit(per_page).all()
        active_total = int(db.session.query(func.coalesce(func.sum(FinancePayment.amount), 0)).filter(
            FinancePayment.is_cancelled.is_(False)
        ).scalar() or 0)
        cancelled_count = FinancePayment.query.filter_by(is_cancelled=True).count()
        query_args = request.args.to_dict(flat=True)
        query_args.pop("page", None)
        prev_url = url_for("finance_payments", **query_args, page=pager["prev_num"]) if pager["has_prev"] else ""
        next_url = url_for("finance_payments", **query_args, page=pager["next_num"]) if pager["has_next"] else ""
        return render_template(
            "finance_v15cd/payment_history.html",
            active_tab="payments",
            rows=rows,
            filters=filters,
            pager=pager,
            prev_url=prev_url,
            next_url=next_url,
            classes=CLASSES,
            months=MONTHS,
            charge_types=ChargeType.query.order_by(ChargeType.sort_order, ChargeType.name).all(),
            academic_year_options=_academic_year_options(_active_general().academic_year_active),
            methods=PAYMENT_METHODS,
            active_total=active_total,
            cancelled_count=cancelled_count,
        )

    @app.route("/finance/payments/<int:payment_id>")
    @_finance_admin_required
    def finance_payment_detail_v15cd(payment_id: int):
        payment = db.get_or_404(FinancePayment, payment_id)
        bill = db.session.get(FinanceBill, payment.bill_id)
        if bill is None:
            abort(404)
        student = db.session.get(Santri, bill.santri_id)
        charge = db.session.get(ChargeType, bill.charge_type_id)
        channel = db.session.get(PaymentChannel, payment.payment_channel_id) if payment.payment_channel_id else None
        creator = db.session.get(User, payment.created_by) if payment.created_by else None
        audit_logs = AuditLog.query.filter_by(entity_type="finance_bill", entity_id=str(bill.id)).filter(
            AuditLog.action.in_(["Mencatat pembayaran", "Membatalkan pembayaran"])
        ).order_by(AuditLog.created_at.desc(), AuditLog.id.desc()).limit(100).all()
        return render_template(
            "finance_v15cd/payment_detail.html",
            active_tab="payments",
            payment=payment,
            bill=bill,
            student=student,
            charge=charge,
            channel=channel,
            creator=creator,
            audit_logs=audit_logs,
            format_wib=_format_wib,
        )

    def _report_filters() -> dict[str, Any]:
        general = _active_general()
        report_type = _clean(request.args.get("report_type") or "all_bills", 30)
        if report_type not in REPORT_TYPES:
            report_type = "all_bills"
        status = _clean(request.args.get("status"), 30)
        return {
            "report_type": report_type,
            "academic_year": _clean(request.args.get("academic_year") if "academic_year" in request.args else general.academic_year_active, 20),
            "semester": _clean(request.args.get("semester") if "semester" in request.args else general.semester_active, 20),
            "class_name": _clean(request.args.get("class_name"), 60),
            "charge_type_id": _clean(request.args.get("charge_type_id"), 20),
            "period": _clean(request.args.get("period"), 80),
            "status": status,
            "method": _clean(request.args.get("method"), 40),
            "date_from": _parse_date(request.args.get("date_from")),
            "date_to": _parse_date(request.args.get("date_to")),
        }

    def _base_bill_query(filters: dict[str, Any]):
        query = db.session.query(FinanceBill, Santri, ChargeType).join(
            Santri, Santri.id == FinanceBill.santri_id
        ).join(ChargeType, ChargeType.id == FinanceBill.charge_type_id).filter(
            FinanceBill.is_archived.is_(False)
        )
        if filters["academic_year"]:
            query = query.filter(FinanceBill.academic_year == filters["academic_year"])
        if filters["semester"]:
            query = query.filter(FinanceBill.semester == filters["semester"])
        if filters["class_name"]:
            query = query.filter(Santri.class_name == filters["class_name"])
        if filters["charge_type_id"].isdigit():
            query = query.filter(FinanceBill.charge_type_id == int(filters["charge_type_id"]))
        if filters["period"]:
            query = query.filter(FinanceBill.period_label == filters["period"])
        forced = STATUS_BY_REPORT.get(filters["report_type"])
        if forced:
            query = query.filter(FinanceBill.status == forced)
        elif filters["status"]:
            query = query.filter(FinanceBill.status == filters["status"])
        if filters["date_from"]:
            query = query.filter(func.date(FinanceBill.created_at) >= filters["date_from"].isoformat())
        if filters["date_to"]:
            query = query.filter(func.date(FinanceBill.created_at) <= filters["date_to"].isoformat())
        return query.order_by(Santri.class_name, Santri.name, FinanceBill.period_year, FinanceBill.period_month)

    def _report_result(filters: dict[str, Any]) -> dict[str, Any]:
        report_type = filters["report_type"]
        title = REPORT_TYPES[report_type]
        columns: list[tuple[str, str]] = []
        rows: list[dict[str, Any]] = []
        totals: dict[str, Any] = {}

        if report_type == "payments":
            pfilters = {
                "q": "",
                "academic_year": filters["academic_year"],
                "semester": filters["semester"],
                "class_name": filters["class_name"],
                "student_name": "",
                "charge_type_id": filters["charge_type_id"],
                "period": filters["period"],
                "method": filters["method"],
                "date_from": filters["date_from"],
                "date_to": filters["date_to"],
                "transaction_status": "all",
                "sort": "oldest",
            }
            data = _payment_query(pfilters).all()
            columns = [
                ("date", "Tanggal"), ("transaction", "Nomor Transaksi"),
                ("student", "Nama Santri"), ("class", "Kelas"),
                ("charge", "Jenis Tagihan"), ("period", "Periode"),
                ("amount", "Nominal Dibayar"), ("method", "Metode"),
                ("receiver", "Admin/Penerima"), ("state", "Status Transaksi"),
            ]
            for payment, bill, student, charge in data:
                rows.append({
                    "date": _format_date(payment.payment_date),
                    "transaction": payment.transaction_number,
                    "student": student.name,
                    "class": student.class_name,
                    "charge": charge.name,
                    "period": f"{bill.period_label} {bill.period_year}",
                    "amount": int(payment.amount or 0),
                    "method": payment.method,
                    "receiver": payment.receiver_name or "Belum diisi",
                    "state": "Dibatalkan" if payment.is_cancelled else "Aktif",
                })
            totals = {
                "total_transactions": len(rows),
                "total_amount": sum(row["amount"] for row in rows if row["state"] == "Aktif"),
                "cancelled": sum(1 for row in rows if row["state"] == "Dibatalkan"),
            }
        else:
            data = _base_bill_query(filters).all()
            if report_type in {"class_summary", "monthly_summary"}:
                grouped: dict[str, dict[str, Any]] = {}
                for bill, student, _charge in data:
                    if report_type == "class_summary":
                        key = student.class_name
                        label = student.class_name
                    else:
                        key = f"{bill.period_year:04d}-{int(bill.period_month or 99):02d}-{bill.period_label}"
                        label = f"{bill.period_label} {bill.period_year}"
                    item = grouped.setdefault(key, {
                        "label": label, "students": set(), "bills": 0,
                        "nominal": 0, "paid": 0, "remaining": 0,
                        "lunas": 0, "sebagian": 0, "belum": 0, "dibebaskan": 0,
                    })
                    item["students"].add(bill.santri_id)
                    item["bills"] += 1
                    item["nominal"] += int(bill.amount or 0)
                    item["paid"] += int(bill.paid_amount or 0)
                    item["remaining"] += int(bill.remaining_amount or 0)
                    if bill.status == "Lunas": item["lunas"] += 1
                    elif bill.status == "Sebagian": item["sebagian"] += 1
                    elif bill.status == "Belum Lunas": item["belum"] += 1
                    elif bill.status == "Dibebaskan": item["dibebaskan"] += 1
                columns = [
                    ("label", "Kelas" if report_type == "class_summary" else "Periode"),
                    ("students", "Jumlah Santri"), ("bills", "Jumlah Tagihan"),
                    ("nominal", "Total Tagihan"), ("paid", "Total Dibayar"),
                    ("remaining", "Total Sisa"), ("lunas", "Lunas"),
                    ("sebagian", "Sebagian"), ("belum", "Belum Lunas"),
                    ("dibebaskan", "Dibebaskan"),
                ]
                for key in sorted(grouped):
                    item = grouped[key]
                    item["students"] = len(item["students"])
                    rows.append(item)
            else:
                columns = [
                    ("student", "Nama Santri"), ("class", "Kelas"),
                    ("charge", "Jenis Tagihan"), ("period", "Periode"),
                    ("nominal", "Nominal"), ("paid", "Dibayar"),
                    ("remaining", "Sisa"), ("status", "Status"),
                    ("due_date", "Jatuh Tempo"),
                ]
                for bill, student, charge in data:
                    rows.append({
                        "student": student.name,
                        "class": student.class_name,
                        "charge": charge.name,
                        "period": f"{bill.period_label} {bill.period_year}",
                        "nominal": int(bill.amount or 0),
                        "paid": int(bill.paid_amount or 0),
                        "remaining": int(bill.remaining_amount or 0),
                        "status": bill.status,
                        "due_date": _format_date(bill.due_date),
                    })
            totals = {
                "total_bills": len(data),
                "total_nominal": sum(int(bill.amount or 0) for bill, *_ in data),
                "total_paid": sum(int(bill.paid_amount or 0) for bill, *_ in data),
                "total_remaining": sum(int(bill.remaining_amount or 0) for bill, *_ in data),
                "lunas": sum(1 for bill, *_ in data if bill.status == "Lunas"),
                "sebagian": sum(1 for bill, *_ in data if bill.status == "Sebagian"),
                "belum": sum(1 for bill, *_ in data if bill.status == "Belum Lunas"),
                "dibebaskan": sum(1 for bill, *_ in data if bill.status == "Dibebaskan"),
            }
        return {"title": title, "columns": columns, "rows": rows, "totals": totals}

    def _filter_payload(filters: dict[str, Any]) -> dict[str, str]:
        payload = {}
        for key, value in filters.items():
            if isinstance(value, date):
                payload[key] = value.isoformat()
            else:
                payload[key] = str(value or "")
        return payload

    def _preview_hash(filters: dict[str, Any]) -> str:
        raw = json.dumps(_filter_payload(filters), ensure_ascii=False, sort_keys=True)
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    def _preview_valid(filters: dict[str, Any]) -> bool:
        return session.get("finance_report_preview_hash") == _preview_hash(filters)

    @_finance_admin_required
    def finance_reports():
        filters = _report_filters()
        preview = request.args.get("preview") == "1"
        result = _report_result(filters) if preview else None
        if preview:
            session["finance_report_preview_hash"] = _preview_hash(filters)
        download_pdf_url = ""
        download_xlsx_url = ""
        if preview:
            payload = _filter_payload(filters)
            download_pdf_url = url_for("finance_report_pdf_v15cd", **payload)
            download_xlsx_url = url_for("finance_report_xlsx_v15cd", **payload)
        return render_template(
            "finance_v15cd/report.html",
            active_tab="reports",
            filters=filters,
            preview=preview,
            result=result,
            download_pdf_url=download_pdf_url,
            download_xlsx_url=download_xlsx_url,
            report_types=REPORT_TYPES,
            classes=CLASSES,
            months=MONTHS,
            charge_types=ChargeType.query.order_by(ChargeType.sort_order, ChargeType.name).all(),
            academic_year_options=_academic_year_options(_active_general().academic_year_active),
            methods=PAYMENT_METHODS,
        )

    def _report_filename(filters: dict[str, Any], extension: str) -> str:
        slug = re.sub(r"[^a-z0-9]+", "_", REPORT_TYPES[filters["report_type"]].lower()).strip("_")
        return f"Laporan_{slug}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{extension}"

    def _ensure_preview(filters: dict[str, Any]):
        if not _preview_valid(filters):
            flash("Tampilkan Preview terlebih dahulu sebelum mengunduh laporan.", "warning")
            return redirect(url_for("finance_reports", **_filter_payload(filters)))
        return None

    def _filter_description(filters: dict[str, Any]) -> str:
        values = [filters["academic_year"], filters["semester"]]
        if filters["class_name"]: values.append(filters["class_name"])
        if filters["period"]: values.append(filters["period"])
        if filters["status"]: values.append(filters["status"])
        if filters["date_from"] or filters["date_to"]:
            values.append(f"Tanggal {_format_date(filters['date_from'])} s.d. {_format_date(filters['date_to'])}")
        return " · ".join(value for value in values if value)

    @app.route("/finance/reports/download.xlsx")
    @_finance_admin_required
    def finance_report_xlsx_v15cd():
        filters = _report_filters()
        guard = _ensure_preview(filters)
        if guard:
            return guard
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        general = _active_general()
        result = _report_result(filters)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Laporan"
        sheet.append([general.tpq_name])
        sheet.append([general.tpq_address or "Belum diisi"])
        sheet.append([result["title"]])
        sheet.append([_filter_description(filters)])
        sheet.append([f"Dicetak: {(datetime.utcnow() + timedelta(hours=7)).strftime('%d/%m/%Y %H:%M WIB')}"])
        sheet.append([])
        sheet.append([label for _key, label in result["columns"]])
        header_row = sheet.max_row
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B8F52")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        money_keys = {"amount", "nominal", "paid", "remaining"}
        for row in result["rows"]:
            values = [row.get(key, "") for key, _label in result["columns"]]
            sheet.append(values)
            for index, (key, _label) in enumerate(result["columns"], start=1):
                if key in money_keys:
                    sheet.cell(sheet.max_row, index).number_format = '"Rp" #,##0'
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(result['columns']))}{sheet.max_row}"
        for column_index, (_key, label) in enumerate(result["columns"], start=1):
            lengths = [len(str(label))]
            lengths.extend(len(str(sheet.cell(row, column_index).value or "")) for row in range(header_row + 1, sheet.max_row + 1))
            max_length = max(lengths)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 34)
        sheet.append([])
        sheet.append(["Ringkasan"])
        for key, value in result["totals"].items():
            sheet.append([key.replace("_", " ").title(), value])
            if key.startswith("total_") and key not in {"total_bills", "total_transactions"}:
                sheet.cell(sheet.max_row, 2).number_format = '"Rp" #,##0'
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=_report_filename(filters, "xlsx"), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/finance/reports/download.pdf")
    @_finance_admin_required
    def finance_report_pdf_v15cd():
        filters = _report_filters()
        guard = _ensure_preview(filters)
        if guard:
            return guard
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        general = _active_general()
        result = _report_result(filters)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("FinanceTitle", parent=styles["Title"], fontSize=14, leading=17, alignment=TA_CENTER, textColor=colors.HexColor("#0B6B45"))
        small = ParagraphStyle("FinanceSmall", parent=styles["BodyText"], fontSize=7.2, leading=9)
        story = [
            Paragraph(general.tpq_name, title_style),
            Paragraph(general.tpq_address or "Belum diisi", ParagraphStyle("addr", parent=small, alignment=TA_CENTER)),
            Spacer(1, 3 * mm),
            Paragraph(result["title"], ParagraphStyle("reporttitle", parent=styles["Heading2"], alignment=TA_CENTER, fontSize=12)),
            Paragraph(_filter_description(filters), ParagraphStyle("filters", parent=small, alignment=TA_CENTER)),
            Spacer(1, 4 * mm),
        ]
        money_keys = {"amount", "nominal", "paid", "remaining"}
        table_data = [[Paragraph(label, small) for _key, label in result["columns"]]]
        for row in result["rows"]:
            values = []
            for key, _label in result["columns"]:
                value = row.get(key, "")
                if key in money_keys:
                    value = _rupiah(value)
                values.append(Paragraph(str(value or "Belum diisi"), small))
            table_data.append(values)
        if len(table_data) == 1:
            table_data.append([Paragraph("Belum ada data sesuai filter.", small)] + [""] * (len(result["columns"]) - 1))
        available_width = landscape(A4)[0] - 20 * mm
        col_width = available_width / max(1, len(result["columns"]))
        table = Table(table_data, colWidths=[col_width] * len(result["columns"]), repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B8F52")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C7D8D0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6FAF8")]),
        ]))
        story.extend([table, Spacer(1, 4 * mm)])
        total_lines = []
        for key, value in result["totals"].items():
            label = key.replace("_", " ").title()
            display = _rupiah(value) if key.startswith("total_") and key not in {"total_bills", "total_transactions"} else str(value)
            total_lines.append(f"<b>{label}:</b> {display}")
        story.append(Paragraph(" &nbsp; | &nbsp; ".join(total_lines), small))
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(f"Dicetak {(datetime.utcnow() + timedelta(hours=7)).strftime('%d/%m/%Y %H:%M WIB')}", small))
        doc.build(story)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=_report_filename(filters, "pdf"), mimetype="application/pdf")

    @app.route("/finance/payments/receipts.pdf", methods=["POST"])
    @_finance_admin_required
    def finance_payment_bulk_receipts_v15cd():
        ids = []
        for value in request.form.getlist("payment_ids"):
            try:
                ids.append(int(value))
            except (TypeError, ValueError):
                continue
        ids = list(dict.fromkeys(ids))[:100]
        if not ids:
            flash("Pilih minimal satu transaksi untuk dicetak.", "warning")
            return redirect(url_for("finance_payments"))
        payments = FinancePayment.query.filter(FinancePayment.id.in_(ids)).order_by(FinancePayment.payment_date, FinancePayment.id).all()
        if not payments:
            abort(404)
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm)
        styles = getSampleStyleSheet()
        story = []
        for index, payment in enumerate(payments):
            bill = db.session.get(FinanceBill, payment.bill_id)
            student = db.session.get(Santri, bill.santri_id) if bill else None
            charge = db.session.get(ChargeType, bill.charge_type_id) if bill else None
            snapshot = payment.snapshot()
            story.append(Paragraph(snapshot.get("tpq_name") or _active_general().tpq_name, styles["Title"]))
            story.append(Paragraph("BUKTI PEMBAYARAN", styles["Heading2"]))
            story.append(Paragraph(f"Nomor Transaksi: {payment.transaction_number}", styles["BodyText"]))
            story.append(Spacer(1, 4 * mm))
            data = [
                ["Nama Santri", snapshot.get("student_name") or (student.name if student else "Belum diisi")],
                ["Kelas", snapshot.get("student_class") or (student.class_name if student else "Belum diisi")],
                ["Jenis Tagihan", snapshot.get("charge_type") or (charge.name if charge else "Belum diisi")],
                ["Periode", snapshot.get("period") or (f"{bill.period_label} {bill.period_year}" if bill else "Belum diisi")],
                ["Nominal Dibayar", _rupiah(payment.amount)],
                ["Tanggal", _format_date(payment.payment_date)],
                ["Metode", payment.method],
                ["Admin/Penerima", payment.receiver_name or "Belum diisi"],
                ["Status", "Dibatalkan" if payment.is_cancelled else "Aktif"],
            ]
            table = Table(data, colWidths=[48 * mm, 105 * mm])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF7F0")),
                ("GRID", (0, 0), (-1, -1), .4, colors.HexColor("#BDD4C8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("PADDING", (0, 0), (-1, -1), 7),
            ]))
            story.append(table)
            story.append(Spacer(1, 5 * mm))
            story.append(Paragraph(snapshot.get("footer_note") or "Terima kasih atas pembayaran administrasi santri.", styles["BodyText"]))
            if index < len(payments) - 1:
                story.append(PageBreak())
        doc.build(story)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"Bukti_Pembayaran_Gabungan_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf", mimetype="application/pdf")

    # Ganti hanya view dari endpoint lama agar URL dan tab tetap konsisten.
    app.view_functions["finance_payments"] = finance_payments
    app.view_functions["finance_reports"] = finance_reports

    app.extensions["finance_history_reports_v15cd"] = {
        "version": "V15-C/D",
        "report_types": REPORT_TYPES,
    }
    return app.extensions["finance_history_reports_v15cd"]
