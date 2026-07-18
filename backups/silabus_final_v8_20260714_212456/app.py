import calendar
import csv
import io
import json
import os
import re
import shutil
import sqlite3
import uuid
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path
from urllib.parse import quote
from zoneinfo import ZoneInfo

from flask import (
    Flask, Response, abort, flash, has_request_context, redirect, render_template, request, session,
    send_file, send_from_directory, url_for
)
from flask_login import (
    LoginManager, UserMixin, current_user, login_required, login_user, logout_user
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import UniqueConstraint, func, inspect, text, or_
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from visual_outputs import render_monthly_poster, render_report_image, render_weekly_poster
from curriculum_documents import (
    build_database_excel, build_database_pdf, build_import_template,
    build_semester_excel, build_semester_pdf, build_single_curriculum_pdf,
    display_value as curriculum_display_value, safe_filename_part,
)


BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
PROOF_DIR = os.path.join(UPLOAD_DIR, "proofs")
BOOK_DIR = os.path.join(UPLOAD_DIR, "books")
PREVIEW_DIR = os.path.join(UPLOAD_DIR, "previews")
os.makedirs(PROOF_DIR, exist_ok=True)
os.makedirs(BOOK_DIR, exist_ok=True)
os.makedirs(PREVIEW_DIR, exist_ok=True)

app = Flask(__name__, instance_relative_config=True)
os.makedirs(app.instance_path, exist_ok=True)
app.config.update(
    SECRET_KEY=os.environ.get("SECRET_KEY", "change-this-secret-key-in-production"),
    SQLALCHEMY_DATABASE_URI=os.environ.get("DATABASE_URL", "sqlite:///tpq_hmarisa.db"),
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    MAX_CONTENT_LENGTH=25 * 1024 * 1024,
    REMEMBER_COOKIE_DURATION=timedelta(days=30),
    REMEMBER_COOKIE_HTTPONLY=True,
    REMEMBER_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    BSI_BANK_NAME="Bank Syariah Indonesia (BSI)",
    BSI_ACCOUNT_NUMBER=os.environ.get("BSI_ACCOUNT_NUMBER", "[ISI NOMOR REKENING BSI]"),
    BSI_ACCOUNT_NAME=os.environ.get("BSI_ACCOUNT_NAME", "TPQ HMarisa"),
)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "home"
login_manager.login_message = "Silakan masuk terlebih dahulu."

CLASSES = ["Ar Rahman", "Ar Rahim", "Al-Bayyan"]
MONTHS = ["Juli", "Agustus", "September", "Oktober", "November", "Desember",
          "Januari", "Februari", "Maret", "April", "Mei", "Juni"]
SURAH_JUZ30 = [
    "An-Naba'", "An-Nazi'at", "'Abasa", "At-Takwir", "Al-Infitar", "Al-Mutaffifin",
    "Al-Insyiqaq", "Al-Buruj", "At-Tariq", "Al-A'la", "Al-Ghasyiyah", "Al-Fajr",
    "Al-Balad", "Asy-Syams", "Al-Lail", "Ad-Duha", "Asy-Syarh", "At-Tin", "Al-'Alaq",
    "Al-Qadr", "Al-Bayyinah", "Az-Zalzalah", "Al-'Adiyat", "Al-Qari'ah", "At-Takasur",
    "Al-'Asr", "Al-Humazah", "Al-Fil", "Quraisy", "Al-Ma'un", "Al-Kausar", "Al-Kafirun",
    "An-Nasr", "Al-Lahab", "Al-Ikhlas", "Al-Falaq", "An-Nas"
]
SCORE_FIELDS = {
    "Ar Rahman": ["BTQ", "Hafalan Surat Juz 30 dan Doa Harian", "Materi dan Praktik Wudhu dan Shalat", "Hadits dan Adab"],
    "Ar Rahim": ["BTQ", "Hafalan Surat Juz 30 dan Doa Harian", "Materi dan Praktik Wudhu dan Shalat", "Fiqih Dasar"],
    "Al-Bayyan": ["Nadzhom Safinatunnajah", "Nadzhom Tauhid", "Nadzhom Akhlak", "Nadzhom Tajwid", "Tahfidz"],
}
TEACHERS = {
    "Ar Rahman": "Yeni Susilawati",
    "Ar Rahim": "Faisal Kazim Muyassar, S.H",
    "Al-Bayyan": "Hj. Maryamah, S.Ag",
}
PRINCIPAL = "Bunda Hj. Maryamah, S.Ag"
BOOK_CATEGORIES = [
    "Kitab Fiqih", "Kitab Akhlaq", "Kitab Nizhomi", "Kitab Tilawati",
    "Buku Edukasi Anak"
]
PREVIEW_PAGE_LIMIT = 3
SEARCH_LIMIT = 30


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="guardian")
    assigned_class = db.Column(db.String(60), default="")
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    last_login_at = db.Column(db.DateTime)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    students = db.relationship("Santri", backref="guardian", lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    @property
    def is_admin(self):
        # Kompatibilitas portal: seluruh akun staf masuk melalui panel admin.
        return self.role in {"admin_utama", "admin", "guru"}

    @property
    def is_superadmin(self):
        return self.role == "admin_utama"

    @property
    def is_teacher(self):
        return self.role == "guru"

    @property
    def role_label(self):
        return {"admin_utama": "Admin Utama", "admin": "Admin", "guru": "Guru", "guardian": "Wali Santri"}.get(self.role, self.role)


class Santri(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nis = db.Column(db.String(12), unique=True, nullable=False, index=True)
    name = db.Column(db.String(140), nullable=False, index=True)
    nickname = db.Column(db.String(80), default="")
    public_name = db.Column(db.String(100), default="")
    class_name = db.Column(db.String(60), nullable=False, index=True)
    guardian_name = db.Column(db.String(140), default="")
    guardian_phone = db.Column(db.String(40), default="")
    joined_date = db.Column(db.Date)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    guardian_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    raport = db.relationship("Raport", backref="santri", uselist=False, cascade="all, delete-orphan")
    bills = db.relationship("Iuran", backref="santri", cascade="all, delete-orphan", lazy=True)
    book_accesses = db.relationship("AksesKitab", backref="santri", cascade="all, delete-orphan", lazy=True)


class Raport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    santri_id = db.Column(db.Integer, db.ForeignKey("santri.id"), unique=True, nullable=False)
    semester = db.Column(db.String(20), default="Semester 1")
    academic_year = db.Column(db.String(20), default="2026/2027")
    scores_json = db.Column(db.Text, default="{}", nullable=False)
    mutabaah_json = db.Column(db.Text, default="[]", nullable=False)
    hafalan_json = db.Column(db.Text, default="{}", nullable=False)
    attitude_json = db.Column(db.Text, default="{}", nullable=False)
    absence_json = db.Column(db.Text, default="{}", nullable=False)
    development_notes = db.Column(db.Text, default="")
    status = db.Column(db.String(30), default="Draf", nullable=False)
    completeness = db.Column(db.Integer, default=0)
    publish_date = db.Column(db.Date)
    published_at = db.Column(db.DateTime)
    version = db.Column(db.Integer, default=1)
    snapshot_json = db.Column(db.Text, default="{}", nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def scores(self):
        try:
            return json.loads(self.scores_json or "{}")
        except json.JSONDecodeError:
            return {}

    def mutabaah(self):
        try:
            return json.loads(self.mutabaah_json or "[]")
        except json.JSONDecodeError:
            return []

    def hafalan(self):
        try:
            return json.loads(self.hafalan_json or "{}")
        except json.JSONDecodeError:
            return {}

    def attitude(self):
        try:
            return json.loads(self.attitude_json or "{}")
        except json.JSONDecodeError:
            return {}

    def absence(self):
        try:
            return json.loads(self.absence_json or "{}")
        except json.JSONDecodeError:
            return {}

    def snapshot(self):
        try:
            return json.loads(self.snapshot_json or "{}")
        except json.JSONDecodeError:
            return {}


class Iuran(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    santri_id = db.Column(db.Integer, db.ForeignKey("santri.id"), nullable=False, index=True)
    month = db.Column(db.String(20), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(30), default="Belum Lunas", nullable=False)
    nominal = db.Column(db.Integer, default=50000, nullable=False)
    proof_path = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    verified_at = db.Column(db.DateTime)
    __table_args__ = (UniqueConstraint("santri_id", "month", "year", name="uq_bill_period"),)


class Kitab(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(180), nullable=False)
    description = db.Column(db.Text, default="")
    category = db.Column(db.String(80), default="Kitab Fiqih", nullable=False, index=True)
    price = db.Column(db.Integer, default=0, nullable=False)
    filename = db.Column(db.String(255))
    original_filename = db.Column(db.String(255))
    original_size = db.Column(db.Integer, default=0)
    optimized_size = db.Column(db.Integer, default=0)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    accesses = db.relationship("AksesKitab", backref="kitab", cascade="all, delete-orphan", lazy=True)


class AksesKitab(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    kitab_id = db.Column(db.Integer, db.ForeignKey("kitab.id"), nullable=False)
    santri_id = db.Column(db.Integer, db.ForeignKey("santri.id"), nullable=False)
    status = db.Column(db.String(30), default="Terkunci", nullable=False)
    proof_path = db.Column(db.String(255))
    requested_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    confirmed_at = db.Column(db.DateTime)
    __table_args__ = (UniqueConstraint("kitab_id", "santri_id", name="uq_book_student"),)


class WeeklyCurriculum(db.Model):
    """Tabel lama dipertahankan, tetapi sekarang mewakili Silabus Bulanan."""
    id = db.Column(db.Integer, primary_key=True)
    class_name = db.Column(db.String(30), nullable=False, index=True)
    week_number = db.Column(db.Integer, nullable=False, default=1)  # kompatibilitas data lama
    month = db.Column(db.String(20), default="Juli", nullable=False, index=True)
    year = db.Column(db.Integer, default=2026, nullable=False, index=True)
    academic_year = db.Column(db.String(20), default="2026/2027", nullable=False)
    subject = db.Column(db.String(120), nullable=False)
    topic = db.Column(db.String(220), nullable=False)
    learning_target = db.Column(db.Text, default="")
    activities = db.Column(db.Text, default="")  # kompatibilitas
    week1 = db.Column(db.Text, default="")
    week2 = db.Column(db.Text, default="")
    week3 = db.Column(db.Text, default="")
    week4 = db.Column(db.Text, default="")
    week5 = db.Column(db.Text, default="")
    notes = db.Column(db.Text, default="")
    source_type = db.Column(db.String(40), default="Input Manual")
    status = db.Column(db.String(20), default="Aktif")
    version = db.Column(db.Integer, default=1)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class MasterClass(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(60), unique=True, nullable=False, index=True)
    code = db.Column(db.String(30), unique=True, nullable=False)
    teacher_name = db.Column(db.String(140), default="")
    sort_order = db.Column(db.Integer, default=0)
    academic_year_start = db.Column(db.String(20), default="2026/2027")
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    notes = db.Column(db.Text, default="")


class Teacher(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(140), nullable=False)
    title = db.Column(db.String(80), default="")
    phone = db.Column(db.String(40), default="")
    class_name = db.Column(db.String(60), default="")
    position = db.Column(db.String(80), default="Guru/Wali Kelas")
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    notes = db.Column(db.Text, default="")


class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(160), nullable=False)
    code = db.Column(db.String(40), nullable=False)
    class_name = db.Column(db.String(60), nullable=False, index=True)
    sort_order = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    notes = db.Column(db.Text, default="")
    __table_args__ = (UniqueConstraint("name", "class_name", name="uq_subject_class"),)


class AcademicYear(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(20), nullable=False)
    semester = db.Column(db.String(20), nullable=False, default="Semester 1")
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    is_active = db.Column(db.Boolean, default=True)
    is_primary = db.Column(db.Boolean, default=False)
    __table_args__ = (UniqueConstraint("name", "semester", name="uq_academic_year_semester"),)


class DailyHadith(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    arabic = db.Column(db.Text, nullable=False)
    translation = db.Column(db.Text, nullable=False)
    source = db.Column(db.String(160), default="")
    number = db.Column(db.String(60), default="")
    theme = db.Column(db.String(80), default="Adab")
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    priority = db.Column(db.Integer, default=0)
    special_date = db.Column(db.Date)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class CurriculumBank(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    material_code = db.Column(db.String(60), unique=True, nullable=False)
    class_name = db.Column(db.String(60), nullable=False, index=True)
    teacher_name = db.Column(db.String(140), default="")
    meeting_number = db.Column(db.Integer, default=1)
    meeting_date = db.Column(db.Date)
    month = db.Column(db.String(20), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    week_in_month = db.Column(db.Integer, default=1)
    day_name = db.Column(db.String(20), default="")
    subject = db.Column(db.String(160), nullable=False, index=True)
    topic = db.Column(db.String(240), nullable=False)
    learning_target = db.Column(db.Text, default="")
    notes = db.Column(db.Text, default="")
    calendar_agenda = db.Column(db.Text, default="")
    is_active = db.Column(db.Boolean, default=True)
    source_file = db.Column(db.String(255), default="")


class HafalanRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    santri_id = db.Column(db.Integer, db.ForeignKey("santri.id"), nullable=False, index=True)
    surah = db.Column(db.String(80), nullable=False, index=True)
    activity_type = db.Column(db.String(30), default="Setoran Baru")
    entry_date = db.Column(db.Date, nullable=False, default=date.today)
    status = db.Column(db.String(30), default="Sedang Proses")
    fluency = db.Column(db.Integer)
    tajwid = db.Column(db.Integer)
    makhraj = db.Column(db.Integer)
    notes = db.Column(db.Text, default="")
    is_deleted = db.Column(db.Boolean, default=False)
    created_by = db.Column(db.String(140), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    student = db.relationship("Santri", backref=db.backref("hafalan_records", lazy=True))


class WeeklyWinner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    period_start = db.Column(db.Date, nullable=False)
    period_end = db.Column(db.Date, nullable=False)
    class_name = db.Column(db.String(60), nullable=False)
    santri_id = db.Column(db.Integer, db.ForeignKey("santri.id"), nullable=False)
    score = db.Column(db.Float, default=0)
    reason = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    student = db.relationship("Santri")
    __table_args__ = (UniqueConstraint("period_start", "class_name", name="uq_weekly_winner"),)


class PosterRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    period_start = db.Column(db.Date, nullable=False)
    period_end = db.Column(db.Date, nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    status = db.Column(db.String(30), default="Siap Dibagikan")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


TEACHER_ALLOWED_ENDPOINTS = {
    "dashboard", "document_preview", "document_preview_page", "curriculum", "curriculum_database", "curriculum_detail", "edit_curriculum", "duplicate_curriculum", "curriculum_pdf", "curriculum_export", "curriculum_import", "curriculum_import_template", "curriculum_semester", "curriculum_semester_export",
    "daily_progress", "academics", "mutabaah_history", "mutabaah_detail", "edit_mutabaah", "delete_mutabaah",
    "hafalan_tracker", "save_hafalan_tracker", "edit_hafalan", "delete_hafalan", "reset_hafalan",
    "eraport", "report_edit", "report_preview", "report_pdf",
    "library", "book_preview", "download_book", "logout",
}


def admin_required(view):
    @wraps(view)
    @login_required
    def wrapped(*args, **kwargs):
        if not current_user.is_admin or not getattr(current_user, "is_active", True):
            abort(403)
        if current_user.is_teacher:
            if request.endpoint not in TEACHER_ALLOWED_ENDPOINTS:
                abort(403)
            assigned = normalize_class_name(current_user.assigned_class)
            requested_class = request.values.get("class_name") or request.values.get("syllabus_class")
            if assigned and requested_class and normalize_class_name(requested_class) != assigned:
                abort(403)
            view_args = request.view_args or {}
            student_id = view_args.get("student_id") or view_args.get("santri_id")
            if student_id:
                student = db.session.get(Santri, int(student_id))
                if student and normalize_class_name(student.class_name) != assigned:
                    abort(403)
            curriculum_id = view_args.get("row_id")
            if curriculum_id:
                curriculum_row = db.session.get(WeeklyCurriculum, int(curriculum_id))
                if curriculum_row and normalize_class_name(curriculum_row.class_name) != assigned:
                    abort(403)
        return view(*args, **kwargs)
    return wrapped


def superadmin_required(view):
    @wraps(view)
    @login_required
    def wrapped(*args, **kwargs):
        if not current_user.is_superadmin or not getattr(current_user, "is_active", True):
            abort(403)
        return view(*args, **kwargs)
    return wrapped


def download_requested(default=True):
    """Return whether a generated file should be downloaded as an attachment.

    All preview pages call the same final file route with ``download=0`` so the
    browser displays the exact PDF/image that will later be downloaded.
    """
    raw = request.args.get("download")
    if raw is None:
        return default
    return str(raw).strip().lower() not in {"0", "false", "no", "preview", "inline"}



def _pdf_stream_bytes(pdf_stream):
    """Return PDF bytes without leaving a BytesIO/file object at the wrong position."""
    if isinstance(pdf_stream, (bytes, bytearray)):
        return bytes(pdf_stream)
    if not hasattr(pdf_stream, "read"):
        raise TypeError("Sumber preview PDF tidak dapat dibaca.")
    original_position = None
    try:
        original_position = pdf_stream.tell()
    except Exception:
        pass
    try:
        pdf_stream.seek(0)
    except Exception:
        pass
    data = pdf_stream.read()
    if original_position is not None:
        try:
            pdf_stream.seek(original_position)
        except Exception:
            pass
    return data


def _cleanup_document_preview_cache(cache_root, max_age_seconds=24 * 60 * 60):
    """Remove stale rendered-page folders so instance storage does not grow forever."""
    import time

    if not os.path.isdir(cache_root):
        return
    cutoff = time.time() - max_age_seconds
    for name in os.listdir(cache_root):
        path = os.path.join(cache_root, name)
        try:
            if os.path.isdir(path) and os.path.getmtime(path) < cutoff:
                shutil.rmtree(path, ignore_errors=True)
        except OSError:
            continue


def render_pdf_preview_pages(pdf_stream):
    """Render the exact final PDF into cached PNG pages for browser-safe preview.

    Android Chrome often replaces an embedded PDF iframe with a generic
    ``PDF / Buka`` card. Rendering the generated PDF itself to images keeps the
    preview visible on mobile while guaranteeing that the preview comes from
    the same PDF bytes that are offered for download.
    """
    import hashlib
    import fitz

    pdf_bytes = _pdf_stream_bytes(pdf_stream)
    if not pdf_bytes:
        raise ValueError("PDF kosong dan tidak dapat dipratinjau.")

    owner_id = str(getattr(current_user, "id", "anonymous"))
    token = hashlib.sha256(owner_id.encode("utf-8") + b":" + pdf_bytes).hexdigest()[:32]
    cache_root = os.path.join(app.instance_path, "document_preview_pages")
    page_dir = os.path.join(cache_root, token)
    manifest_path = os.path.join(page_dir, "manifest.json")
    os.makedirs(cache_root, exist_ok=True)
    _cleanup_document_preview_cache(cache_root)

    page_count = 0
    if os.path.exists(manifest_path):
        try:
            manifest = json.loads(Path(manifest_path).read_text(encoding="utf-8"))
            if str(manifest.get("owner_id", "")) != owner_id:
                page_count = 0
            else:
                page_count = int(manifest.get("page_count", 0))
            if page_count < 1 or any(
                not os.path.exists(os.path.join(page_dir, f"page-{number}.png"))
                for number in range(1, page_count + 1)
            ):
                page_count = 0
        except (OSError, ValueError, TypeError, json.JSONDecodeError):
            page_count = 0

    if page_count < 1:
        shutil.rmtree(page_dir, ignore_errors=True)
        os.makedirs(page_dir, exist_ok=True)
        document = fitz.open(stream=pdf_bytes, filetype="pdf")
        try:
            page_count = document.page_count
            if page_count < 1:
                raise ValueError("PDF tidak memiliki halaman.")
            matrix = fitz.Matrix(1.45, 1.45)
            for index in range(page_count):
                page = document.load_page(index)
                pixmap = page.get_pixmap(matrix=matrix, alpha=False)
                pixmap.save(os.path.join(page_dir, f"page-{index + 1}.png"))
        finally:
            document.close()
        Path(manifest_path).write_text(
            json.dumps({"page_count": page_count, "owner_id": owner_id}, ensure_ascii=False),
            encoding="utf-8",
        )

    return token, page_count


def attach_pdf_page_preview(context, pdf_stream):
    """Attach direct page-image URLs to a document-preview template context."""
    try:
        token, page_count = render_pdf_preview_pages(pdf_stream)
        context.update(
            preview_kind="pdf-pages",
            preview_page_count=page_count,
            preview_pages=[
                url_for("document_preview_page", token=token, page_number=number)
                for number in range(1, page_count + 1)
            ],
        )
    except Exception:
        app.logger.exception("Gagal merender halaman preview PDF")
        context.update(
            preview_kind="pdf-fallback",
            preview_page_count=0,
            preview_pages=[],
            notice=(
                "Pratinjau halaman belum dapat dirender. Gunakan tombol Buka/Cetak "
                "untuk memeriksa PDF pada tab baru."
            ),
        )


def selected_guardian_student():
    """Return the one student selected from the public guardian entry form."""
    if not current_user.is_authenticated or current_user.is_admin:
        return None
    student_id = session.get("guardian_student_id")
    if not student_id:
        return None
    student = db.session.get(Santri, int(student_id))
    if not student or student.guardian_id != current_user.id:
        return None
    return student


def can_access_student(student):
    if not current_user.is_authenticated:
        return False
    if current_user.is_admin:
        return True
    selected = selected_guardian_student()
    return bool(selected and selected.id == student.id)


def get_or_create_raport(student):
    if student.raport:
        return student.raport
    raport = Raport(santri_id=student.id)
    db.session.add(raport)
    db.session.flush()
    return raport


def jakarta_now():
    return datetime.now(ZoneInfo("Asia/Jakarta"))


def normalize_class_name(value):
    mapping = {
        "Ar-Rahman": "Ar Rahman",
        "Ar-Rahim": "Ar Rahim",
        "Ar Rahman": "Ar Rahman",
        "Ar Rahim": "Ar Rahim",
        "Al Bayyan": "Al-Bayyan",
        "Al-Bayyan": "Al-Bayyan",
    }
    return mapping.get((value or "").strip(), (value or "").strip())


def active_class_names(include_inactive=False):
    try:
        if has_request_context() and current_user.is_authenticated and getattr(current_user, "is_teacher", False) and current_user.assigned_class:
            return [normalize_class_name(current_user.assigned_class)]
        query = MasterClass.query
        if not include_inactive:
            query = query.filter_by(is_active=True)
        names = [row.name for row in query.order_by(MasterClass.sort_order, MasterClass.name).all()]
        return names or list(CLASSES)
    except Exception:
        return list(CLASSES)


def class_teacher(class_name):
    class_name = normalize_class_name(class_name)
    row = MasterClass.query.filter_by(name=class_name).first()
    if row and row.teacher_name:
        return row.teacher_name
    return TEACHERS.get(class_name, "")


def class_subjects(class_name, include_inactive=False):
    class_name = normalize_class_name(class_name)
    try:
        query = Subject.query.filter_by(class_name=class_name)
        if not include_inactive:
            query = query.filter_by(is_active=True)
        rows = query.order_by(Subject.sort_order, Subject.name).all()
        if rows:
            return [row.name for row in rows]
    except Exception:
        pass
    return SCORE_FIELDS.get(class_name, [])


def current_academic_year():
    try:
        row = AcademicYear.query.filter_by(is_primary=True).first()
        if row:
            return row
    except Exception:
        pass
    return None


def current_daily_hadith(on_date=None):
    on_date = on_date or jakarta_now().date()
    special = (DailyHadith.query
               .filter_by(is_active=True, special_date=on_date)
               .order_by(DailyHadith.priority.desc(), DailyHadith.id)
               .first())
    if special:
        return special
    rows = (DailyHadith.query.filter_by(is_active=True)
            .filter(DailyHadith.special_date.is_(None))
            .order_by(DailyHadith.priority.desc(), DailyHadith.id)
            .all())
    if rows:
        index = on_date.toordinal() % len(rows)
        return rows[index]
    return None


def parse_date(value, fallback=None):
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return fallback


def format_date_id(value):
    if not value:
        return "-"
    if isinstance(value, str):
        value = parse_date(value)
    if not value:
        return "-"
    month_names = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    return f"{value.day} {month_names[value.month - 1]} {value.year}"


app.jinja_env.filters["tanggal_id"] = format_date_id


def score_predicate(value):
    if value is None or value == "":
        return "", "Belum diisi", "Belum diisi"
    try:
        value = int(value)
    except (TypeError, ValueError):
        return "", "Belum diisi", "Belum diisi"
    if value < 60 or value > 100:
        return "", "Nilai tidak valid", "Belum diisi"
    if value >= 91:
        pred, note = "A", "Sangat Baik"
    elif value >= 81:
        pred, note = "B", "Baik"
    elif value >= 71:
        pred, note = "C", "Cukup"
    else:
        pred, note = "D", "Perlu Bimbingan"
    return pred, note, "Tuntas" if value >= 70 else "Belum Tuntas"


def current_hafalan_status(student_id):
    records = (HafalanRecord.query
               .filter_by(santri_id=student_id, is_deleted=False)
               .order_by(HafalanRecord.entry_date.desc(), HafalanRecord.id.desc())
               .all())
    latest = {}
    for row in records:
        latest.setdefault(row.surah, row)
    return latest


def report_completeness(student, raport):
    subjects = class_subjects(student.class_name)
    scores = raport.scores()
    checks = []
    checks.extend(bool(scores.get(subject) not in (None, "", 0, "0")) for subject in subjects)
    attitude = raport.attitude()
    checks.extend(bool(attitude.get(key)) for key in ["Kehadiran", "Kedisiplinan", "Keterlibatan", "Pergaulan/Perilaku"])
    checks.append(bool((raport.development_notes or "").strip()))
    checks.append(bool(raport.publish_date))
    if not checks:
        return 0
    return int(round(sum(1 for item in checks if item) / len(checks) * 100))


def mutabaah_entries(student):
    raport = get_or_create_raport(student)
    entries = raport.mutabaah()
    changed = False
    for idx, entry in enumerate(entries):
        entry["_index"] = idx
        if "id" not in entry:
            entry["id"] = f"legacy-{student.id}-{idx}-{uuid.uuid4().hex[:8]}"
            changed = True
        if "subject" not in entry:
            entry["subject"] = entry.get("fiqih", "")
            changed = True
        if "material" not in entry:
            entry["material"] = entry.get("fiqih", "")
            changed = True
        entry.setdefault("attendance", "Hadir")
        entry.setdefault("on_time", True)
        entry.setdefault("adab_score", 80)
        entry.setdefault("tahsin_level", "Tilawati/Iqro" if entry.get("tilawati") else ("Al-Qur'an" if entry.get("alquran") else ""))
        entry.setdefault("tilawati", "")
        entry.setdefault("alquran", "")
        entry.setdefault("keterangan", "")
        entry.setdefault("notes", "")
        entry.setdefault("deleted", False)
    if changed:
        raport.mutabaah_json = json.dumps(entries, ensure_ascii=False)
    return entries


def monthly_period(reference=None):
    """Return the first and last day of the selected Gregorian month."""
    reference = reference or jakarta_now().date()
    first_day = reference.replace(day=1)
    last_day = reference.replace(day=calendar.monthrange(reference.year, reference.month)[1])
    return first_day, last_day


def monthly_label(reference=None):
    reference = reference or jakarta_now().date()
    month_names = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
                   "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    return f"{month_names[reference.month - 1]} {reference.year}"


def _is_monthly_period(row):
    if not row or not row.period_start or not row.period_end:
        return False
    expected_start, expected_end = monthly_period(row.period_start)
    return row.period_start == expected_start and row.period_end == expected_end


def latest_monthly_winners():
    """Keep showing the latest completed/generated month until a newer month exists."""
    rows = WeeklyWinner.query.order_by(WeeklyWinner.period_start.desc(), WeeklyWinner.id.desc()).all()
    target_start = None
    for row in rows:
        if _is_monthly_period(row):
            target_start = row.period_start
            break
    if not target_start:
        return []
    return [row for row in rows if row.period_start == target_start and _is_monthly_period(row)]


def latest_monthly_poster_record():
    for row in PosterRecord.query.order_by(PosterRecord.created_at.desc(), PosterRecord.id.desc()).all():
        if _is_monthly_period(row):
            return row
    return None


def calculate_monthly_winners(reference=None, force=False):
    period_start, period_end = monthly_period(reference)
    winners = []
    for class_name in active_class_names():
        existing = WeeklyWinner.query.filter_by(period_start=period_start, class_name=class_name).first()
        if existing and _is_monthly_period(existing) and not force:
            winners.append(existing)
            continue
        candidates = []
        for student in Santri.query.filter_by(class_name=class_name, is_active=True).all():
            entries = []
            for entry in mutabaah_entries(student):
                entry_date = parse_date(entry.get("date"))
                if entry_date and period_start <= entry_date <= period_end and not entry.get("deleted"):
                    entries.append(entry)
            attendance = sum(1 for e in entries if e.get("attendance", "Hadir") == "Hadir")
            on_time = sum(1 for e in entries if e.get("on_time", True))
            adab_values = [int(e.get("adab_score", 80) or 80) for e in entries]
            adab_avg = sum(adab_values) / len(adab_values) if adab_values else 0
            hafalan_rows = (HafalanRecord.query
                            .filter_by(santri_id=student.id, is_deleted=False)
                            .filter(HafalanRecord.entry_date >= period_start, HafalanRecord.entry_date <= period_end)
                            .all())
            new_hafalan = sum(1 for h in hafalan_rows
                              if h.activity_type in {"Setoran Baru", "Ujian Hafalan"}
                              and h.status == "Sudah Hafal")
            quality = [v for h in hafalan_rows for v in [h.fluency, h.tajwid, h.makhraj] if v is not None]
            quality_avg = sum(quality) / len(quality) if quality else 0
            journal_score = min(len(entries), 20) * 3
            score = (journal_score + attendance * 5 + on_time * 3 + new_hafalan * 12
                     + quality_avg * 0.08 + adab_avg * 0.08)
            candidates.append((score, new_hafalan, attendance, on_time, quality_avg, adab_avg, student))
        if not candidates:
            continue
        # Bila skor sama: hafalan terbanyak, kehadiran terbaik, ketepatan waktu,
        # kualitas hafalan, lalu adab. Tidak ada pemilihan acak.
        candidates.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4], x[5], x[6].name.lower()), reverse=True)
        best = candidates[0]
        reason = (f"Hafalan baru {best[1]}, kehadiran {best[2]} kali, tepat waktu {best[3]} kali, "
                  f"kualitas hafalan {best[4]:.0f}, adab {best[5]:.0f}.")
        if existing:
            existing.santri_id = best[6].id
            existing.period_end = period_end
            existing.score = round(best[0], 2)
            existing.reason = reason
            winner = existing
        else:
            winner = WeeklyWinner(period_start=period_start, period_end=period_end, class_name=class_name,
                                  santri_id=best[6].id, score=round(best[0], 2), reason=reason)
            db.session.add(winner)
        winners.append(winner)
    db.session.commit()
    return winners


def generate_monthly_poster(reference=None, force=False):
    """Create the approved monthly award poster for download and admin review."""
    reference = reference or jakarta_now().date()
    period_start, period_end = monthly_period(reference)
    winners = calculate_monthly_winners(reference, force=force)
    poster_dir = os.path.join(UPLOAD_DIR, "posters")
    os.makedirs(poster_dir, exist_ok=True)
    month_slug = monthly_label(reference).lower().replace(" ", "_")
    filename = f"santri_terbaik_{month_slug}.png"
    path = os.path.join(poster_dir, filename)
    if os.path.exists(path) and not force:
        return filename
    try:
        template = os.path.join(BASE_DIR, "static", "img", "weekly_award_template.png")
        ordered = []
        winner_map = {normalize_class_name(row.class_name): row for row in winners}
        for class_name in ["Ar Rahman", "Ar Rahim", "Al-Bayyan"]:
            row = winner_map.get(class_name)
            if row:
                public_name = (row.student.public_name or row.student.nickname or row.student.name or "").strip()
                if public_name and public_name == public_name.lower():
                    public_name = public_name.title()
            else:
                public_name = "Belum ditetapkan"
            ordered.append({"class_name": class_name, "name": public_name})
        render_monthly_poster(template, ordered, monthly_label(reference), path)
        record = PosterRecord.query.filter_by(period_start=period_start, period_end=period_end).first()
        if not record:
            record = PosterRecord(period_start=period_start, period_end=period_end,
                                  filename=filename, status="Siap Diperiksa")
            db.session.add(record)
        else:
            record.filename = filename
            record.status = "Siap Diperiksa"
            record.created_at = datetime.utcnow()
        db.session.commit()
    except Exception as exc:
        app.logger.exception("Poster bulanan gagal dibuat: %s", exc)
        return None
    return filename


# Alias kompatibilitas agar task/tautan lama tidak langsung rusak.
def weekly_period(reference=None):
    start, end = monthly_period(reference)
    return start, end, end


def calculate_weekly_winners(reference=None, force=False):
    return calculate_monthly_winners(reference=reference, force=force)


def generate_weekly_poster(reference=None, force=False):
    return generate_monthly_poster(reference=reference, force=force)


def generate_nis(year=None):
    year = year or datetime.now().year
    prefix = str(year)[-2:]
    existing = Santri.query.filter(Santri.nis.like(f"{prefix}%")).all()
    numbers = []
    for student in existing:
        suffix = student.nis[2:]
        if suffix.isdigit():
            numbers.append(int(suffix))
    seq = max(numbers, default=0) + 1
    return f"{prefix}{seq:03d}"


def save_upload(file_storage, folder, allowed_exts):
    if not file_storage or not file_storage.filename:
        return None
    original = secure_filename(file_storage.filename)
    ext = original.rsplit(".", 1)[-1].lower() if "." in original else ""
    if ext not in allowed_exts:
        raise ValueError(f"Format file tidak didukung. Gunakan: {', '.join(sorted(allowed_exts))}")
    filename = f"{uuid.uuid4().hex}.{ext}"
    path = os.path.join(folder, filename)
    file_storage.save(path)
    return filename


def rupiah(value):
    try:
        return "Rp {:,.0f}".format(int(value)).replace(",", ".")
    except (TypeError, ValueError):
        return "Rp 0"


app.jinja_env.filters["rupiah"] = rupiah


@app.context_processor
def inject_globals():
    classes = active_class_names() if "MasterClass" in globals() else list(CLASSES)
    teachers = {name: class_teacher(name) for name in classes} if "MasterClass" in globals() else dict(TEACHERS)
    hadith = current_daily_hadith() if "DailyHadith" in globals() else None
    return {
        "CLASSES": classes,
        "MONTHS": MONTHS,
        "SURAH_JUZ30": SURAH_JUZ30,
        "SCORE_FIELDS": {name: class_subjects(name) for name in classes} if "Subject" in globals() else SCORE_FIELDS,
        "TEACHERS": teachers,
        "PRINCIPAL": PRINCIPAL,
        "BOOK_CATEGORIES": BOOK_CATEGORIES,
        "PREVIEW_PAGE_LIMIT": PREVIEW_PAGE_LIMIT,
        "current_date": jakarta_now().date().isoformat() if "jakarta_now" in globals() else date.today().isoformat(),
        "current_year": jakarta_now().year if "jakarta_now" in globals() else datetime.now().year,
        "daily_hadith": hadith,
        "format_date_id": format_date_id if "format_date_id" in globals() else str,
        "score_predicate": score_predicate if "score_predicate" in globals() else None,
        "class_subjects": class_subjects if "class_subjects" in globals() else None,
        "class_teacher": class_teacher if "class_teacher" in globals() else None,
        "curriculum_display_value": curriculum_display_value if "curriculum_display_value" in globals() else str,
    }


@app.route("/", methods=["GET", "POST"])
def home():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for("dashboard"))
        if selected_guardian_student():
            return redirect(url_for("dashboard"))
        logout_user()
        session.pop("guardian_student_id", None)

    classes = active_class_names()
    students_by_class = {class_name: [] for class_name in classes}
    students = (Santri.query.filter_by(is_active=True)
                .order_by(Santri.class_name, func.lower(Santri.name), Santri.id).all())
    for student in students:
        normalized = normalize_class_name(student.class_name)
        students_by_class.setdefault(normalized, []).append({
            "id": student.id, "name": student.name, "nis": student.nis,
        })

    selected_class = normalize_class_name(request.form.get("class_name", "")) if request.method == "POST" else ""
    selected_student_id = request.form.get("student_id", type=int) if request.method == "POST" else None
    if request.method == "POST":
        if selected_class not in classes or not selected_student_id:
            flash("Pilih kelas dan nama santri terlebih dahulu.", "danger")
            return render_template("guardian_entry.html", students_by_class=students_by_class,
                                   class_name=selected_class, student_id=selected_student_id), 400
        student = db.session.get(Santri, selected_student_id)
        if not student or not student.is_active or normalize_class_name(student.class_name) != selected_class:
            flash("Data santri tidak ditemukan atau tidak sesuai dengan kelas yang dipilih.", "danger")
            return render_template("guardian_entry.html", students_by_class=students_by_class,
                                   class_name=selected_class, student_id=selected_student_id), 404
        login_user(student.guardian, remember=False)
        session["guardian_student_id"] = student.id
        session["guardian_entry"] = True
        flash(f"Data ananda {student.name} berhasil dibuka.", "success")
        return redirect(url_for("dashboard"))

    winners = latest_monthly_winners()
    winner_map = {row.class_name: row for row in winners}
    winner_period_label = monthly_label(winners[0].period_start) if winners else ""
    return render_template("guardian_entry.html", students_by_class=students_by_class,
                           monthly_winners=winner_map, monthly_period_label=winner_period_label,
                           classes=classes)

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for("dashboard"))
        logout_user()
        session.pop("guardian_student_id", None)
        session.pop("guardian_entry", None)

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = (User.query.filter(User.username == username,
                                  User.role.in_(["admin_utama", "admin", "guru"]))
                .first())
        if not user or not user.is_active or not user.check_password(password):
            flash("Username atau kata sandi tidak sesuai, atau akun sedang nonaktif.", "danger")
            return render_template("login.html", username=username), 401
        user.last_login_at = datetime.utcnow()
        db.session.commit()
        login_user(user, remember=bool(request.form.get("remember")))
        session.pop("guardian_student_id", None)
        session.pop("guardian_entry", None)
        flash(f"Selamat datang, {user.full_name}.", "success")
        return redirect(request.args.get("next") or url_for("dashboard"))
    return render_template("login.html")


@app.route("/login")
def login():
    return redirect(url_for("admin_login"))


@app.route("/logout")
def logout():
    was_admin = current_user.is_authenticated and current_user.is_admin
    if current_user.is_authenticated:
        logout_user()
    session.pop("guardian_student_id", None)
    session.pop("guardian_entry", None)
    flash("Anda telah keluar." if was_admin else "Akses data ananda telah ditutup.", "info")
    return redirect(url_for("home"))


@app.route("/dashboard")
@login_required
def dashboard():
    if current_user.is_admin:
        force = request.args.get("refresh_monthly") == "1" or request.args.get("refresh_weekly") == "1"
        now = jakarta_now()
        is_month_end_time = now.date() == monthly_period(now.date())[1] and now.hour >= 17
        poster_filename = None
        if force or is_month_end_time:
            winners = calculate_monthly_winners(reference=now.date(), force=force)
            poster_filename = generate_monthly_poster(reference=now.date(), force=force)
        else:
            winners = latest_monthly_winners()
        latest_poster = latest_monthly_poster_record()
        monthly_period_label = monthly_label(winners[0].period_start) if winners else ""
        poster_notice = None
        if latest_poster and latest_poster.status != "Sudah Dipublikasikan":
            poster_notice = (f"Poster Santri Terbaik Bulan {monthly_label(latest_poster.period_start)} "
                             "sudah tersedia. Silakan diperiksa dan dibagikan malam ini atau besok.")
        stats = {
            "students": Santri.query.filter_by(is_active=True).count(),
            "unpaid": Iuran.query.filter(Iuran.status != "Lunas").count(),
            "pending_bills": Iuran.query.filter_by(status="Menunggu Verifikasi").count(),
            "pending_books": AksesKitab.query.filter_by(status="Menunggu Verifikasi").count(),
        }
        recent_bills = Iuran.query.order_by(Iuran.created_at.desc()).limit(6).all()
        curriculum_class = normalize_class_name(request.args.get("syllabus_class", ""))
        curriculum_month = request.args.get("syllabus_month", MONTHS[0])
        if current_user.is_teacher and current_user.assigned_class:
            curriculum_class = normalize_class_name(current_user.assigned_class)
        curriculum_rows = []
        syllabus_filtered = bool(curriculum_class and curriculum_month)
        primary_year = current_academic_year()
        if curriculum_class in active_class_names() and curriculum_month in MONTHS:
            query = WeeklyCurriculum.query.filter_by(class_name=curriculum_class, month=curriculum_month)
            if primary_year:
                query = query.filter_by(academic_year=primary_year.name)
            curriculum_rows = query.order_by(WeeklyCurriculum.subject, WeeklyCurriculum.version.desc()).all()
        return render_template("dashboard_admin.html", stats=stats, recent_bills=recent_bills,
                               curriculum_class=curriculum_class, curriculum_month=curriculum_month,
                               curriculum_rows=curriculum_rows, syllabus_filtered=syllabus_filtered,
                               primary_year=primary_year, monthly_winners=winners,
                               monthly_period_label=monthly_period_label, poster_notice=poster_notice,
                               latest_poster=latest_poster, poster_filename=poster_filename)

    student = selected_guardian_student()
    if not student:
        logout_user(); session.pop("guardian_student_id", None)
        flash("Silakan masukkan kembali nama santri dan kelas.", "info")
        return redirect(url_for("home"))
    raport = get_or_create_raport(student)
    entries = [e for e in mutabaah_entries(student) if not e.get("deleted")]
    db.session.commit()
    bills = Iuran.query.filter_by(santri_id=student.id).order_by(Iuran.year.desc(), Iuran.id.desc()).all()
    latest_hafalan = current_hafalan_status(student.id)
    legacy = raport.hafalan()
    hafalan_done = sum(1 for surah in SURAH_JUZ30 if
                       (latest_hafalan.get(surah) and latest_hafalan[surah].status == "Sudah Hafal") or
                       (not latest_hafalan.get(surah) and legacy.get(surah)))
    unpaid_count = sum(1 for bill in bills if bill.status != "Lunas")
    open_access_ids = {access.kitab_id for access in student.book_accesses if access.status == "Terbuka"}
    available_books = Kitab.query.filter((Kitab.price == 0) | (Kitab.id.in_(open_access_ids or {-1}))).count()
    return render_template("dashboard_guardian.html", student=student, selected_student=student,
                           raport=raport, bills=bills, mutabaah=entries, hafalan=legacy,
                           hafalan_done=hafalan_done, unpaid_count=unpaid_count,
                           available_books=available_books)


def student_export_query(class_filter="", status_filter="Semua Status", q=""):
    """Build a consistent student query for on-screen lists and exports."""
    classes = active_class_names(include_inactive=True)
    class_filter = normalize_class_name(class_filter or "")
    query = Santri.query
    if class_filter in classes:
        query = query.filter_by(class_name=class_filter)
    if status_filter == "Aktif":
        query = query.filter_by(is_active=True)
    elif status_filter == "Nonaktif":
        query = query.filter_by(is_active=False)
    if q:
        query = query.filter(
            or_(
                Santri.name.ilike(f"%{q}%"),
                Santri.nis.ilike(f"%{q}%"),
                Santri.guardian_name.ilike(f"%{q}%"),
            )
        )
    return query


@app.route("/students", methods=["GET", "POST"])
@admin_required
def students():
    classes = active_class_names(include_inactive=True)
    active_tab = request.args.get("tab", "database")
    if active_tab not in {"database", "add"}:
        active_tab = "database"

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        class_name = normalize_class_name(request.form.get("class_name", ""))
        guardian_name = request.form.get("guardian_name", "").strip() or f"Wali {name}"
        guardian_phone = request.form.get("guardian_phone", "").strip()
        public_name = request.form.get("public_name", "").strip() or name
        if not name or class_name not in classes:
            flash("Nama santri dan kelas wajib diisi.", "danger")
            return redirect(url_for("students", tab="add"))
        nis = request.form.get("nis", "").strip() or generate_nis()
        if Santri.query.filter_by(nis=nis).first():
            flash("NIS tersebut sudah digunakan oleh santri lain.", "danger")
            return redirect(url_for("students", tab="add"))
        duplicate = Santri.query.filter(func.lower(Santri.name) == name.lower(), Santri.class_name == class_name).first()
        if duplicate and request.form.get("confirm_duplicate") != "1":
            flash("Terdapat nama yang sama di kelas ini. Periksa kembali sebelum menyimpan.", "warning")
            return redirect(url_for("students", tab="add", q=name, class_name=class_name))
        guardian_username = request.form.get("guardian_username", "").strip() or f"wali{nis}"
        base_username = guardian_username
        counter = 1
        while User.query.filter_by(username=guardian_username).first():
            guardian_username = f"{base_username}{counter}"
            counter += 1
        guardian = User(username=guardian_username, full_name=guardian_name, role="guardian")
        guardian.set_password(uuid.uuid4().hex)
        db.session.add(guardian)
        db.session.flush()
        student = Santri(
            nis=nis,
            name=name,
            public_name=public_name,
            class_name=class_name,
            guardian_name=guardian_name,
            guardian_phone=guardian_phone,
            joined_date=parse_date(request.form.get("joined_date"), date.today()),
            is_active=True,
            guardian_id=guardian.id,
        )
        db.session.add(student)
        db.session.flush()
        db.session.add(Raport(santri_id=student.id))
        db.session.commit()
        flash(f"Santri {name} berhasil ditambahkan dengan NIS {student.nis}.", "success")
        return redirect(url_for("students", tab="database", student_id=student.id, class_name=class_name))

    q = request.args.get("q", "").strip()
    class_filter = normalize_class_name(request.args.get("class_name", ""))
    status_filter = request.args.get("status", "Aktif")
    selected_id = request.args.get("student_id", type=int)
    query = student_export_query(class_filter, status_filter, q)
    if selected_id:
        query = query.filter(Santri.id == selected_id)
    total_matches = query.count()
    student_rows = query.order_by(func.lower(Santri.name)).limit(SEARCH_LIMIT).all()
    counts = {name: Santri.query.filter_by(class_name=name, is_active=True).count() for name in classes}
    students_by_class = {
        name: [
            {"id": st.id, "name": st.name, "nis": st.nis}
            for st in Santri.query.filter_by(class_name=name).order_by(func.lower(Santri.name)).all()
        ]
        for name in classes
    }
    return render_template(
        "students.html",
        students=student_rows,
        class_filter=class_filter,
        q=q,
        status_filter=status_filter,
        selected_id=selected_id,
        class_counts=counts,
        students_by_class=students_by_class,
        total_matches=total_matches,
        search_limit=SEARCH_LIMIT,
        active_tab=active_tab,
    )


@app.route("/students/import", methods=["GET", "POST"])
@app.route("/students/bulk", methods=["GET", "POST"])
@admin_required
def bulk_students():
    """Two-step Excel import with validation preview before saving."""
    from openpyxl import load_workbook

    preview = []
    token = request.form.get("token", "")
    preview_dir = os.path.join(app.instance_path, "student_import_previews")
    os.makedirs(preview_dir, exist_ok=True)

    if request.method == "POST" and request.form.get("action") == "commit":
        token = secure_filename(token)
        preview_path = os.path.join(preview_dir, f"{token}.json")
        if not token or not os.path.exists(preview_path):
            flash("Data pratinjau tidak ditemukan. Unggah ulang file Excel.", "danger")
            return redirect(url_for("bulk_students"))
        with open(preview_path, "r", encoding="utf-8") as handle:
            preview = json.load(handle)
        created = 0
        skipped = 0
        for row in preview:
            if row.get("status") != "Valid":
                skipped += 1
                continue
            try:
                nis = str(row.get("nis") or "").strip() or generate_nis()
                if Santri.query.filter_by(nis=nis).first():
                    skipped += 1
                    continue
                name = row["name"].strip()
                guardian_name = (row.get("guardian_name") or f"Wali {name}").strip()
                username_base = f"wali{nis}"
                username = username_base
                counter = 1
                while User.query.filter_by(username=username).first():
                    username = f"{username_base}{counter}"
                    counter += 1
                guardian = User(username=username, full_name=guardian_name, role="guardian")
                guardian.set_password(uuid.uuid4().hex)
                db.session.add(guardian)
                db.session.flush()
                student = Santri(
                    nis=nis,
                    name=name,
                    public_name=(row.get("public_name") or name).strip(),
                    class_name=row["class_name"],
                    guardian_name=guardian_name,
                    guardian_phone=str(row.get("guardian_phone") or "").strip(),
                    joined_date=parse_date(row.get("joined_date"), date.today()),
                    is_active=True,
                    guardian_id=guardian.id,
                )
                db.session.add(student)
                db.session.flush()
                db.session.add(Raport(santri_id=student.id))
                db.session.commit()
                created += 1
            except Exception:
                db.session.rollback()
                skipped += 1
        try:
            os.remove(preview_path)
        except OSError:
            pass
        flash(f"Impor selesai. {created} santri berhasil ditambahkan dan {skipped} baris dilewati.", "success" if created else "warning")
        return redirect(url_for("students", tab="database"))

    if request.method == "POST":
        upload = request.files.get("excel_file")
        if not upload or not upload.filename:
            flash("Pilih file Excel terlebih dahulu.", "danger")
            return redirect(url_for("bulk_students"))
        if not upload.filename.lower().endswith(".xlsx"):
            flash("Format import yang didukung adalah Excel .xlsx.", "danger")
            return redirect(url_for("bulk_students"))
        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError("File Excel kosong")
            headers = [str(value or "").strip().lower() for value in rows[0]]
            aliases = {
                "name": {"nama lengkap", "nama santri", "nama"},
                "public_name": {"nama tampilan publik", "nama publik", "nama panggilan"},
                "nis": {"nis"},
                "class_name": {"kelas"},
                "guardian_name": {"nama wali", "wali"},
                "guardian_phone": {"nomor kontak wali", "nomor kontak", "no hp wali", "telepon wali"},
                "joined_date": {"tanggal masuk", "tgl masuk"},
            }
            index_map = {}
            for key, names in aliases.items():
                index_map[key] = next((idx for idx, header in enumerate(headers) if header in names), None)
            if index_map["name"] is None or index_map["class_name"] is None:
                raise ValueError("Kolom Nama Lengkap dan Kelas wajib tersedia")

            classes = active_class_names(include_inactive=True)
            seen_nis = set()
            for excel_row, values in enumerate(rows[1:], start=2):
                if not any(value not in (None, "") for value in values):
                    continue
                def value_for(key):
                    idx = index_map.get(key)
                    return values[idx] if idx is not None and idx < len(values) else ""
                name = str(value_for("name") or "").strip()
                class_name = normalize_class_name(str(value_for("class_name") or "").strip())
                nis_raw = value_for("nis")
                if isinstance(nis_raw, float) and nis_raw.is_integer():
                    nis_raw = int(nis_raw)
                nis = str(nis_raw or "").strip()
                issues = []
                if not name:
                    issues.append("Nama kosong")
                if class_name not in classes:
                    issues.append("Kelas tidak dikenali")
                if nis and (nis in seen_nis or Santri.query.filter_by(nis=nis).first()):
                    issues.append("NIS duplikat")
                if nis:
                    seen_nis.add(nis)
                joined_value = value_for("joined_date")
                if isinstance(joined_value, datetime):
                    joined_value = joined_value.date().isoformat()
                elif isinstance(joined_value, date):
                    joined_value = joined_value.isoformat()
                else:
                    joined_value = str(joined_value or "").strip()
                preview.append({
                    "excel_row": excel_row,
                    "name": name,
                    "public_name": str(value_for("public_name") or name).strip(),
                    "nis": nis,
                    "class_name": class_name,
                    "guardian_name": str(value_for("guardian_name") or "").strip(),
                    "guardian_phone": str(value_for("guardian_phone") or "").strip(),
                    "joined_date": joined_value,
                    "status": "Valid" if not issues else "; ".join(issues),
                })
            token = uuid.uuid4().hex
            preview_path = os.path.join(preview_dir, f"{token}.json")
            with open(preview_path, "w", encoding="utf-8") as handle:
                json.dump(preview, handle, ensure_ascii=False)
        except Exception as exc:
            flash(f"File Excel tidak dapat dibaca: {exc}", "danger")
            return redirect(url_for("bulk_students"))

    valid_count = sum(1 for row in preview if row.get("status") == "Valid")
    return render_template("bulk_students.html", preview=preview, token=token, valid_count=valid_count)


def curriculum_academic_year_options():
    """Tahun ajaran untuk form Silabus, selalu tersedia hingga 2049/2050."""
    built_in = [f"{year}/{year + 1}" for year in range(2026, 2050)]
    try:
        stored = [row[0] for row in db.session.query(AcademicYear.name).distinct().all() if row[0]]
    except Exception:
        stored = []
    return sorted(set(built_in + stored), key=lambda value: int(value.split('/')[0]))


def curriculum_semester_for_month(month):
    return "Semester 1" if month in ["Juli", "Agustus", "September", "Oktober", "November", "Desember"] else "Semester 2"


def curriculum_months_for_semester(semester):
    return MONTHS[:6] if semester == "Semester 1" else MONTHS[6:]


def curriculum_year_for_month(academic_year, month):
    parts = str(academic_year or "").split("/")
    if len(parts) != 2:
        raise ValueError("Format tahun ajaran harus YYYY/YYYY")
    return int(parts[0] if curriculum_semester_for_month(month) == "Semester 1" else parts[1])


def curriculum_filter_description(filters):
    labels = []
    if filters.get("class_filter"):
        labels.append(f"Kelas: {filters['class_filter']}")
    if filters.get("month_filter"):
        labels.append(f"Bulan: {filters['month_filter']}")
    if filters.get("academic_year_filter"):
        labels.append(f"Tahun Ajaran: {filters['academic_year_filter']}")
    if filters.get("semester_filter"):
        labels.append(filters["semester_filter"])
    if filters.get("subject_filter"):
        labels.append(f"Bidang Pelajaran: {filters['subject_filter']}")
    if filters.get("status_filter"):
        labels.append(f"Status: {filters['status_filter']}")
    if filters.get("keyword"):
        labels.append(f"Kata Kunci: {filters['keyword']}")
    return " | ".join(labels) if labels else "Seluruh data silabus"


def curriculum_semester_rows(class_name, academic_year, semester):
    months = curriculum_months_for_semester(semester)
    query = (WeeklyCurriculum.query
             .filter_by(class_name=class_name, academic_year=academic_year, status="Aktif")
             .filter(WeeklyCurriculum.month.in_(months)))
    candidates = query.order_by(WeeklyCurriculum.version.desc(), WeeklyCurriculum.updated_at.desc()).all()
    latest = {}
    for row in candidates:
        key = (row.month, row.subject)
        if key not in latest:
            latest[key] = row
    return sorted(latest.values(), key=lambda row: (months.index(row.month), row.subject.lower()))


def curriculum_semester_overview(rows, class_name, semester):
    months = curriculum_months_for_semester(semester)
    subjects = class_subjects(class_name)
    by_subject = []
    for subject in subjects:
        subject_rows = [row for row in rows if row.subject == subject]
        filled_months = [month for month in months if any(row.month == month for row in subject_rows)]
        by_subject.append({
            "subject": subject,
            "count": len(subject_rows),
            "months": ", ".join(filled_months) if filled_months else "Belum diisi",
        })
    completeness = []
    for month in months:
        existing = {row.subject for row in rows if row.month == month}
        missing = [subject for subject in subjects if subject not in existing]
        completeness.append({
            "month": month,
            "status": "Lengkap" if not missing else ("Belum ada data" if not existing else "Belum lengkap"),
            "missing": missing,
            "missing_text": ", ".join(missing) if missing else "Tidak ada",
        })
    return by_subject, completeness


def curriculum_query_from_request():
    classes = active_class_names()
    class_filter = normalize_class_name(request.args.get("class_name", ""))
    if current_user.is_authenticated and getattr(current_user, "is_teacher", False) and current_user.assigned_class:
        class_filter = normalize_class_name(current_user.assigned_class)
    month_filter = request.args.get("month", "")
    academic_year_filter = request.args.get("academic_year", "")
    semester_filter = request.args.get("semester", "")
    subject_filter = request.args.get("subject", "").strip()
    status_filter = request.args.get("status", "Aktif").strip()
    keyword = request.args.get("keyword", "").strip()
    query = WeeklyCurriculum.query
    if class_filter in classes:
        query = query.filter_by(class_name=class_filter)
    if month_filter in MONTHS:
        query = query.filter_by(month=month_filter)
    if academic_year_filter:
        query = query.filter_by(academic_year=academic_year_filter)
    if semester_filter == "Semester 1":
        query = query.filter(WeeklyCurriculum.month.in_(["Juli", "Agustus", "September", "Oktober", "November", "Desember"]))
    elif semester_filter == "Semester 2":
        query = query.filter(WeeklyCurriculum.month.in_(["Januari", "Februari", "Maret", "April", "Mei", "Juni"]))
    if subject_filter:
        query = query.filter_by(subject=subject_filter)
    if status_filter in {"Aktif", "Nonaktif"}:
        query = query.filter_by(status=status_filter)
    if keyword:
        like = f"%{keyword}%"
        query = query.filter(or_(WeeklyCurriculum.topic.ilike(like), WeeklyCurriculum.learning_target.ilike(like),
                                 WeeklyCurriculum.notes.ilike(like), WeeklyCurriculum.subject.ilike(like)))
    filters = dict(class_filter=class_filter, month_filter=month_filter,
                   academic_year_filter=academic_year_filter, semester_filter=semester_filter,
                   subject_filter=subject_filter, status_filter=status_filter, keyword=keyword)
    return query, filters


@app.route("/curriculum", methods=["GET", "POST"])
@admin_required
def curriculum():
    """Halaman khusus input Silabus Bulanan; database berada pada URL terpisah."""
    classes = active_class_names()
    academic_year_options = curriculum_academic_year_options()
    primary = current_academic_year()
    primary_name = primary.name if primary else "2026/2027"
    if request.method == "POST":
        class_name = normalize_class_name(request.form.get("class_name", ""))
        month = request.form.get("month", "")
        academic_year = request.form.get("academic_year", primary_name).strip()
        try:
            year = int(academic_year.split('/')[0]) if month in ["Juli", "Agustus", "September", "Oktober", "November", "Desember"] else int(academic_year.split('/')[1])
        except Exception:
            year = request.form.get("year", type=int) or jakarta_now().year
        subject = request.form.get("subject", "").strip()
        topic = request.form.get("topic", "").strip()
        learning_target = request.form.get("learning_target", "").strip()
        weeks = [request.form.get(f"week{i}", "").strip() for i in range(1, 6)]
        notes = request.form.get("notes", "").strip()
        source_type = request.form.get("source_type", "Input Manual")
        if class_name not in classes or month not in MONTHS or academic_year not in academic_year_options or not subject or not topic or not any(weeks):
            flash("Kelas, bulan, tahun ajaran, bidang, materi pokok, dan minimal satu rencana pekan wajib diisi.", "danger")
            return redirect(url_for("curriculum"))
        existing = WeeklyCurriculum.query.filter_by(class_name=class_name, month=month, year=year,
                                                     academic_year=academic_year, subject=subject, status="Aktif").first()
        if existing and request.form.get("duplicate_action") != "revision":
            flash("Silabus untuk kelas, bulan, tahun ajaran, dan bidang pelajaran ini sudah tersedia. Edit data lama atau simpan sebagai revisi.", "warning")
            return redirect(url_for("edit_curriculum", row_id=existing.id))
        version = existing.version + 1 if existing else 1
        row = WeeklyCurriculum(class_name=class_name, month=month, year=year, academic_year=academic_year,
                               week_number=1, subject=subject, topic=topic, learning_target=learning_target,
                               activities="\n".join(filter(None, weeks)), week1=weeks[0], week2=weeks[1],
                               week3=weeks[2], week4=weeks[3], week5=weeks[4], notes=notes,
                               source_type=source_type, status="Aktif", version=version)
        db.session.add(row); db.session.commit()
        flash("Silabus Bulanan berhasil disimpan.", "success")
        return redirect(url_for("curriculum_database", class_name=class_name, month=month, academic_year=academic_year))

    bank_rows = CurriculumBank.query.filter_by(is_active=True).order_by(CurriculumBank.class_name,
                                                                         CurriculumBank.subject,
                                                                         CurriculumBank.meeting_number).all()
    bank_data = [{"class_name": b.class_name, "subject": b.subject, "month": b.month, "year": b.year,
                  "week": b.week_in_month, "topic": b.topic, "target": b.learning_target,
                  "agenda": b.calendar_agenda, "code": b.material_code} for b in bank_rows]
    return render_template("curriculum.html", bank_data=bank_data, bank_count=len(bank_rows),
                           primary_year=primary, academic_year_options=academic_year_options)


@app.route("/curriculum/database")
@admin_required
def curriculum_database():
    query, filters = curriculum_query_from_request()
    rows = query.order_by(WeeklyCurriculum.year.desc(), WeeklyCurriculum.month, WeeklyCurriculum.class_name,
                          WeeklyCurriculum.subject, WeeklyCurriculum.version.desc()).all()
    return render_template("curriculum_database.html", rows=rows,
                           academic_year_options=curriculum_academic_year_options(), **filters)


@app.route("/curriculum/<int:row_id>/detail")
@admin_required
def curriculum_detail(row_id):
    row = db.session.get(WeeklyCurriculum, row_id)
    if row is None:
        abort(404)
    teacher_name = class_teacher(row.class_name) or "Belum diisi"
    return render_template(
        "curriculum_detail.html",
        row=row,
        semester=curriculum_semester_for_month(row.month),
        teacher_name=teacher_name,
        academic_year_options=curriculum_academic_year_options(),
    )


@app.route("/curriculum/<int:row_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_curriculum(row_id):
    row = db.get_or_404(WeeklyCurriculum, row_id)
    if request.method == "POST":
        row.class_name = normalize_class_name(request.form.get("class_name", row.class_name))
        row.month = request.form.get("month", row.month)
        row.academic_year = request.form.get("academic_year", row.academic_year).strip()
        try:
            row.year = int(row.academic_year.split('/')[0]) if curriculum_semester_for_month(row.month) == "Semester 1" else int(row.academic_year.split('/')[1])
        except Exception:
            pass
        row.subject = request.form.get("subject", row.subject).strip()
        row.topic = request.form.get("topic", row.topic).strip()
        row.learning_target = request.form.get("learning_target", "").strip()
        for i in range(1, 6):
            setattr(row, f"week{i}", request.form.get(f"week{i}", "").strip())
        row.activities = "\n".join(filter(None, [row.week1, row.week2, row.week3, row.week4, row.week5]))
        row.notes = request.form.get("notes", "").strip(); row.updated_at = datetime.utcnow()
        db.session.commit(); flash("Silabus Bulanan berhasil diperbarui.", "success")
        return redirect(url_for("curriculum_detail", row_id=row.id))
    return render_template("curriculum_edit.html", row=row,
                           academic_year_options=curriculum_academic_year_options())


@app.route("/curriculum/<int:row_id>/duplicate", methods=["POST"])
@admin_required
def duplicate_curriculum(row_id):
    source = db.get_or_404(WeeklyCurriculum, row_id)
    month = request.form.get("month", "")
    academic_year = request.form.get("academic_year", source.academic_year)
    if month not in MONTHS:
        flash("Pilih bulan tujuan yang valid.", "danger")
        return redirect(url_for("curriculum_detail", row_id=row_id))
    try:
        year = int(academic_year.split('/')[0]) if curriculum_semester_for_month(month) == "Semester 1" else int(academic_year.split('/')[1])
    except Exception:
        year = source.year
    existing = WeeklyCurriculum.query.filter_by(class_name=source.class_name, month=month,
                                                 academic_year=academic_year, subject=source.subject,
                                                 status="Aktif").first()
    if existing:
        flash("Silabus tujuan sudah tersedia.", "warning")
        return redirect(url_for("curriculum_detail", row_id=existing.id))
    copy = WeeklyCurriculum(class_name=source.class_name, month=month, year=year,
                            academic_year=academic_year, week_number=1, subject=source.subject,
                            topic=source.topic, learning_target=source.learning_target,
                            activities=source.activities, week1=source.week1, week2=source.week2,
                            week3=source.week3, week4=source.week4, week5=source.week5,
                            notes=source.notes, source_type=f"Duplikat dari {source.month} {source.year}",
                            status="Aktif", version=1)
    db.session.add(copy); db.session.commit()
    flash("Silabus berhasil diduplikat.", "success")
    return redirect(url_for("curriculum_detail", row_id=copy.id))


@app.route("/curriculum/<int:row_id>/delete", methods=["POST"])
@admin_required
def delete_curriculum(row_id):
    row = db.get_or_404(WeeklyCurriculum, row_id)
    row.status = "Nonaktif" if row.status == "Aktif" else "Aktif"
    db.session.commit()
    flash(f"Silabus Bulanan sekarang berstatus {row.status}.", "info")
    return redirect(request.referrer or url_for("curriculum_database"))


@app.route("/curriculum/<int:row_id>/destroy", methods=["POST"])
@admin_required
def destroy_curriculum(row_id):
    row = db.get_or_404(WeeklyCurriculum, row_id)
    if row.status == "Aktif":
        flash("Nonaktifkan silabus terlebih dahulu sebelum menghapus permanen.", "warning")
        return redirect(request.referrer or url_for("curriculum_database"))
    if request.form.get("confirm_text", "").strip().upper() != "HAPUS":
        flash("Konfirmasi penghapusan tidak sesuai.", "danger")
        return redirect(request.referrer or url_for("curriculum_database"))
    db.session.delete(row); db.session.commit()
    flash("Silabus Bulanan berhasil dihapus permanen.", "success")
    return redirect(url_for("curriculum_database"))


@app.route("/curriculum/<int:row_id>/pdf")
@admin_required
def curriculum_pdf(row_id):
    row = db.session.get(WeeklyCurriculum, row_id)
    if row is None:
        abort(404)
    teacher_name = class_teacher(row.class_name) or "Belum diisi"
    letterhead_path = os.path.join(BASE_DIR, "static", "img", "kop_surat_tpq_hmarisa.png")
    buffer = build_single_curriculum_pdf(
        row,
        teacher_name=teacher_name,
        principal=PRINCIPAL,
        date_text=format_date_id(jakarta_now().date()),
        letterhead_path=letterhead_path,
    )
    filename = "silabus_{}_{}_{}.pdf".format(
        safe_filename_part(row.class_name),
        safe_filename_part(row.month),
        row.year,
    )
    return send_file(buffer, mimetype="application/pdf", as_attachment=download_requested(), download_name=filename)


@app.route("/curriculum/export/<file_type>")
@admin_required
def curriculum_export(file_type):
    query, filters = curriculum_query_from_request()
    rows = query.all()
    rows = sorted(
        rows,
        key=lambda row: (row.year, MONTHS.index(row.month) if row.month in MONTHS else 99,
                         row.class_name.lower(), row.subject.lower(), -int(row.version or 1)),
    )
    distinct_classes = sorted({row.class_name for row in rows})
    if len(distinct_classes) == 1:
        second_role = "Guru/Wali Kelas"
        second_name = class_teacher(distinct_classes[0]) or "Belum diisi"
    else:
        second_role = "Admin/Petugas Kurikulum"
        second_name = getattr(current_user, "full_name", "Administrator TPQ HMarisa")
    filter_text = curriculum_filter_description(filters)
    letterhead_path = os.path.join(BASE_DIR, "static", "img", "kop_surat_tpq_hmarisa.png")
    date_text = format_date_id(jakarta_now().date())
    if file_type == "excel":
        output = build_database_excel(
            rows,
            filter_text=filter_text,
            principal=PRINCIPAL,
            second_role=second_role,
            second_name=second_name,
            date_text=date_text,
            letterhead_path=letterhead_path,
        )
        return send_file(
            output,
            as_attachment=True,
            download_name=f"database_silabus_bulanan_{jakarta_now().date().isoformat()}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    if file_type == "pdf":
        output = build_database_pdf(
            rows,
            filter_text=filter_text,
            principal=PRINCIPAL,
            second_role=second_role,
            second_name=second_name,
            date_text=date_text,
            letterhead_path=letterhead_path,
        )
        return send_file(
            output,
            as_attachment=download_requested(),
            download_name=f"database_silabus_bulanan_{jakarta_now().date().isoformat()}.pdf",
            mimetype="application/pdf",
        )
    abort(404)


@app.route("/curriculum/import/template")
@admin_required
def curriculum_import_template():
    classes = active_class_names()
    output = build_import_template(
        classes=classes,
        subjects_map={class_name: class_subjects(class_name) for class_name in classes},
        academic_years=curriculum_academic_year_options(),
        months=MONTHS,
    )
    return send_file(
        output,
        as_attachment=True,
        download_name="Template_Import_Silabus_Bulanan_TPQ_HMarisa.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/curriculum/import", methods=["GET", "POST"])
@admin_required
def curriculum_import():
    preview = []
    token = request.form.get("token", "")
    preview_dir = os.path.join(app.instance_path, "curriculum_import_previews")
    os.makedirs(preview_dir, exist_ok=True)

    if request.method == "POST" and request.form.get("action") == "confirm":
        path = os.path.join(preview_dir, secure_filename(token) + ".json")
        if not token or not os.path.exists(path):
            flash("Pratinjau import sudah tidak tersedia. Unggah ulang file Excel.", "danger")
            return redirect(url_for("curriculum_import"))
        preview = json.loads(Path(path).read_text(encoding="utf-8"))
        imported = 0
        skipped = 0
        for item in preview:
            if item.get("validation_status") != "Valid":
                skipped += 1
                continue
            existing = WeeklyCurriculum.query.filter_by(
                class_name=item["class_name"],
                month=item["month"],
                academic_year=item["academic_year"],
                subject=item["subject"],
                status=item["record_status"],
            ).first()
            if existing:
                skipped += 1
                continue
            row = WeeklyCurriculum(
                class_name=item["class_name"],
                month=item["month"],
                year=item["year"],
                academic_year=item["academic_year"],
                week_number=1,
                subject=item["subject"],
                topic=item["topic"],
                learning_target=item["learning_target"],
                week1=item["week1"],
                week2=item["week2"],
                week3=item["week3"],
                week4=item["week4"],
                week5=item["week5"],
                activities="\n".join(filter(None, [item["week1"], item["week2"], item["week3"], item["week4"], item["week5"]])),
                notes=item["notes"],
                source_type="Import Excel",
                status=item["record_status"],
                version=item["version"],
            )
            db.session.add(row)
            imported += 1
        db.session.commit()
        try:
            os.remove(path)
        except OSError:
            pass
        flash(f"Import selesai: {imported} data disimpan, {skipped} data dilewati.", "success")
        return redirect(url_for("curriculum_database"))

    if request.method == "POST" and request.files.get("file"):
        upload = request.files["file"]
        if not upload.filename.lower().endswith(".xlsx"):
            flash("Gunakan file Excel berformat .xlsx.", "danger")
            return redirect(url_for("curriculum_import"))
        from openpyxl import load_workbook
        try:
            wb = load_workbook(upload, data_only=False, read_only=True)
            ws = wb["Data Silabus"] if "Data Silabus" in wb.sheetnames else wb.active
            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
            aliases = {
                "kelas": "class_name",
                "bulan": "month",
                "tahun ajaran": "academic_year",
                "semester": "semester",
                "bidang pelajaran": "subject",
                "materi pokok": "topic",
                "target pembelajaran": "learning_target",
                "pekan 1": "week1",
                "pekan 2": "week2",
                "pekan 3": "week3",
                "pekan 4": "week4",
                "pekan 5": "week5",
                "rencana materi pekan 1": "week1",
                "rencana materi pekan 2": "week2",
                "rencana materi pekan 3": "week3",
                "rencana materi pekan 4": "week4",
                "rencana materi pekan 5": "week5",
                "catatan guru": "notes",
                "status": "record_status",
                "versi": "version",
            }
            indexes = {aliases[header]: index for index, header in enumerate(headers) if header in aliases}
            required = {"class_name", "month", "academic_year", "subject", "topic"}
            if not required.issubset(indexes):
                missing = sorted(required.difference(indexes))
                raise ValueError(f"Kolom wajib tidak lengkap ({', '.join(missing)}). Gunakan Template Import terbaru.")

            classes = active_class_names()
            years = curriculum_academic_year_options()
            keys = sorted(set(aliases.values()))
            for number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                def value_for(key):
                    index = indexes.get(key)
                    if index is None or index >= len(values):
                        return ""
                    raw = values[index]
                    if raw is None:
                        return ""
                    return str(raw).strip()

                item = {key: value_for(key) for key in keys}
                if not any(item.get(key) for key in ["class_name", "month", "academic_year", "subject", "topic"]):
                    continue
                item["class_name"] = normalize_class_name(item["class_name"])
                issues = []
                if item["class_name"] not in classes:
                    issues.append("Kelas tidak dikenal")
                if item["month"] not in MONTHS:
                    issues.append("Bulan tidak valid")
                if item["academic_year"] not in years:
                    issues.append("Tahun ajaran tidak valid")
                computed_semester = curriculum_semester_for_month(item["month"]) if item["month"] in MONTHS else ""
                supplied_semester = item.get("semester", "")
                if supplied_semester and not supplied_semester.startswith("=") and supplied_semester not in {computed_semester, "Semester 1", "Semester 2"}:
                    issues.append("Semester tidak valid")
                if supplied_semester in {"Semester 1", "Semester 2"} and computed_semester and supplied_semester != computed_semester:
                    issues.append("Semester tidak sesuai bulan")
                item["semester"] = computed_semester
                if item["subject"] and item["class_name"] in classes and item["subject"] not in class_subjects(item["class_name"]):
                    issues.append("Bidang pelajaran tidak sesuai kelas")
                if not item["topic"]:
                    issues.append("Materi pokok kosong")
                if not any(item.get(f"week{index}") for index in range(1, 6)):
                    issues.append("Minimal satu rencana materi pekan wajib diisi")
                item["record_status"] = item.get("record_status") or "Aktif"
                if item["record_status"] not in {"Aktif", "Nonaktif"}:
                    issues.append("Status harus Aktif atau Nonaktif")
                try:
                    item["version"] = max(1, int(float(item.get("version") or 1)))
                except (TypeError, ValueError):
                    item["version"] = 1
                    issues.append("Versi harus berupa angka")
                try:
                    item["year"] = curriculum_year_for_month(item["academic_year"], item["month"])
                except Exception:
                    item["year"] = jakarta_now().year
                duplicate = WeeklyCurriculum.query.filter_by(
                    class_name=item["class_name"],
                    month=item["month"],
                    academic_year=item["academic_year"],
                    subject=item["subject"],
                    status=item["record_status"],
                ).first()
                if duplicate:
                    issues.append("Data sudah tersedia")
                item["excel_row"] = number
                item["validation_status"] = "Valid" if not issues else "; ".join(issues)
                preview.append(item)
            wb.close()
            token = uuid.uuid4().hex
            Path(os.path.join(preview_dir, token + ".json")).write_text(
                json.dumps(preview, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception as exc:
            flash(f"File Excel tidak dapat dibaca: {exc}", "danger")
            return redirect(url_for("curriculum_import"))

    valid_count = sum(1 for item in preview if item.get("validation_status") == "Valid")
    invalid_count = len(preview) - valid_count
    return render_template(
        "curriculum_import.html",
        preview=preview,
        token=token,
        valid_count=valid_count,
        invalid_count=invalid_count,
    )


@app.route("/curriculum/semester")
@admin_required
def curriculum_semester():
    classes = active_class_names()
    primary = current_academic_year()
    default_year = primary.name if primary else "2026/2027"
    class_name = normalize_class_name(request.args.get("class_name", ""))
    academic_year = request.args.get("academic_year", default_year).strip()
    semester = request.args.get("semester", "Semester 1").strip()
    rows = []
    summary = []
    completeness = []
    filtered = class_name in classes and academic_year in curriculum_academic_year_options() and semester in {"Semester 1", "Semester 2"}
    if filtered:
        rows = curriculum_semester_rows(class_name, academic_year, semester)
        summary, completeness = curriculum_semester_overview(rows, class_name, semester)
    return render_template(
        "curriculum_semester.html",
        classes=classes,
        class_name=class_name,
        academic_year=academic_year,
        semester=semester,
        academic_year_options=curriculum_academic_year_options(),
        rows=rows,
        summary=summary,
        completeness=completeness,
        filtered=filtered,
    )


@app.route("/curriculum/semester/export/<file_type>")
@admin_required
def curriculum_semester_export(file_type):
    classes = active_class_names()
    class_name = normalize_class_name(request.args.get("class_name", ""))
    academic_year = request.args.get("academic_year", "").strip()
    semester = request.args.get("semester", "").strip()
    if class_name not in classes or academic_year not in curriculum_academic_year_options() or semester not in {"Semester 1", "Semester 2"}:
        flash("Pilih kelas, tahun ajaran, dan semester sebelum mengunduh rekap.", "danger")
        return redirect(url_for("curriculum_semester"))
    rows = curriculum_semester_rows(class_name, academic_year, semester)
    summary, completeness = curriculum_semester_overview(rows, class_name, semester)
    teacher_name = class_teacher(class_name) or "Belum diisi"
    letterhead_path = os.path.join(BASE_DIR, "static", "img", "kop_surat_tpq_hmarisa.png")
    date_text = format_date_id(jakarta_now().date())
    base_name = "rekap_silabus_{}_{}_{}".format(
        safe_filename_part(semester),
        safe_filename_part(class_name),
        safe_filename_part(academic_year),
    )
    if file_type == "pdf":
        output = build_semester_pdf(
            rows,
            class_name=class_name,
            academic_year=academic_year,
            semester=semester,
            teacher_name=teacher_name,
            principal=PRINCIPAL,
            date_text=date_text,
            letterhead_path=letterhead_path,
        )
        return send_file(output, as_attachment=download_requested(), download_name=base_name + ".pdf", mimetype="application/pdf")
    if file_type == "excel":
        output = build_semester_excel(
            rows,
            class_name=class_name,
            academic_year=academic_year,
            semester=semester,
            teacher_name=teacher_name,
            principal=PRINCIPAL,
            date_text=date_text,
            letterhead_path=letterhead_path,
            summary=summary,
            completeness=completeness,
        )
        return send_file(
            output,
            as_attachment=True,
            download_name=base_name + ".xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    abort(404)


@app.route("/curriculum/bank/download")
@admin_required
def curriculum_bank_download():
    """Kompatibilitas tautan lama; arahkan ke preview export Silabus."""
    return redirect(url_for("document_preview", document_type="curriculum-database"))


@app.route("/preview/<document_type>")
@admin_required
def document_preview(document_type):
    """Global preview gate used before every administrative download.

    PDF previews render the exact final PDF bytes into PNG pages so they are
    visible directly on Android, tablet, and desktop without an iframe viewer.
    Excel/template previews render a web table that mirrors the workbook fields.
    Existing modules that already are previews (E-Raport, receipt, library) keep
    their own preview screen and link to their final download from there.
    """
    if current_user.is_teacher and document_type not in {
        "curriculum-single", "curriculum-database", "curriculum-semester",
        "curriculum-template", "poster",
    }:
        abort(403)

    context = {
        "document_type": document_type,
        "preview_kind": "summary",
        "title": "Pratinjau Dokumen",
        "subtitle": "Periksa isi sebelum mengunduh.",
        "preview_url": "",
        "preview_pages": [],
        "preview_page_count": 0,
        "download_buttons": [],
        "back_url": url_for("dashboard"),
        "edit_url": "",
        "print_url": "",
        "orientation": "",
        "preview_headers": [],
        "preview_rows": [],
        "summary_items": [],
        "notice": "Preview dan file akhir menggunakan sumber data yang sama.",
    }

    args = request.args.to_dict(flat=True)
    args.pop("download", None)

    if document_type == "curriculum-single":
        row_id = request.args.get("row_id", type=int)
        row = db.session.get(WeeklyCurriculum, row_id) if row_id else None
        if row is None:
            abort(404)
        if current_user.is_teacher:
            assigned = normalize_class_name(current_user.assigned_class)
            if assigned and normalize_class_name(row.class_name) != assigned:
                abort(403)
        context.update(
            title="Preview Silabus Bulanan",
            subtitle=f"{row.class_name} · {row.month} {row.year} · {row.subject}",
            preview_url=url_for("curriculum_pdf", row_id=row.id, download=0),
            print_url=url_for("curriculum_pdf", row_id=row.id, download=0),
            orientation="A4 landscape",
            back_url=url_for("curriculum_detail", row_id=row.id),
            edit_url=url_for("edit_curriculum", row_id=row.id),
            download_buttons=[{
                "label": "Unduh PDF", "url": url_for("curriculum_pdf", row_id=row.id, download=1),
                "icon": "fa-file-pdf", "class": "btn-primary",
            }],
        )
        preview_pdf = build_single_curriculum_pdf(
            row,
            teacher_name=class_teacher(row.class_name) or "Belum diisi",
            principal=PRINCIPAL,
            date_text=format_date_id(jakarta_now().date()),
            letterhead_path=os.path.join(BASE_DIR, "static", "img", "kop_surat_tpq_hmarisa.png"),
        )
        attach_pdf_page_preview(context, preview_pdf)

    elif document_type == "curriculum-database":
        query, filters = curriculum_query_from_request()
        count = query.count()
        context.update(
            title="Preview Database Silabus Bulanan",
            subtitle=f"{curriculum_filter_description(filters)} · {count} data",
            preview_url=url_for("curriculum_export", file_type="pdf", download=0, **args),
            print_url=url_for("curriculum_export", file_type="pdf", download=0, **args),
            orientation="A4 landscape",
            back_url=url_for("curriculum_database", **args),
            edit_url=url_for("curriculum"),
            download_buttons=[
                {"label": "Unduh PDF", "url": url_for("curriculum_export", file_type="pdf", download=1, **args), "icon": "fa-file-pdf", "class": "btn-primary"},
                {"label": "Unduh Excel", "url": url_for("curriculum_export", file_type="excel", download=1, **args), "icon": "fa-file-excel", "class": "btn-outline"},
            ],
        )
        preview_rows = sorted(
            query.all(),
            key=lambda item: (
                item.year,
                MONTHS.index(item.month) if item.month in MONTHS else 99,
                item.class_name.lower(),
                item.subject.lower(),
                -int(item.version or 1),
            ),
        )
        preview_classes = sorted({item.class_name for item in preview_rows})
        if len(preview_classes) == 1:
            preview_second_role = "Guru/Wali Kelas"
            preview_second_name = class_teacher(preview_classes[0]) or "Belum diisi"
        else:
            preview_second_role = "Admin/Petugas Kurikulum"
            preview_second_name = getattr(current_user, "full_name", "Administrator TPQ HMarisa")
        preview_pdf = build_database_pdf(
            preview_rows,
            filter_text=curriculum_filter_description(filters),
            principal=PRINCIPAL,
            second_role=preview_second_role,
            second_name=preview_second_name,
            date_text=format_date_id(jakarta_now().date()),
            letterhead_path=os.path.join(BASE_DIR, "static", "img", "kop_surat_tpq_hmarisa.png"),
        )
        attach_pdf_page_preview(context, preview_pdf)

    elif document_type == "curriculum-semester":
        class_name = normalize_class_name(request.args.get("class_name", ""))
        academic_year = request.args.get("academic_year", "").strip()
        semester = request.args.get("semester", "").strip()
        if class_name not in active_class_names() or academic_year not in curriculum_academic_year_options() or semester not in {"Semester 1", "Semester 2"}:
            flash("Pilih kelas, tahun ajaran, dan semester terlebih dahulu.", "danger")
            return redirect(url_for("curriculum_semester"))
        if current_user.is_teacher:
            assigned = normalize_class_name(current_user.assigned_class)
            if assigned and class_name != assigned:
                abort(403)
        export_args = {"class_name": class_name, "academic_year": academic_year, "semester": semester}
        context.update(
            title="Preview Rekap Silabus Semester",
            subtitle=f"{class_name} · {academic_year} · {semester}",
            preview_url=url_for("curriculum_semester_export", file_type="pdf", download=0, **export_args),
            print_url=url_for("curriculum_semester_export", file_type="pdf", download=0, **export_args),
            orientation="A4 landscape",
            back_url=url_for("curriculum_semester", **export_args),
            edit_url=url_for("curriculum_database", class_name=class_name, academic_year=academic_year, semester=semester, status="Aktif"),
            download_buttons=[
                {"label": "Unduh PDF", "url": url_for("curriculum_semester_export", file_type="pdf", download=1, **export_args), "icon": "fa-file-pdf", "class": "btn-primary"},
                {"label": "Unduh Excel", "url": url_for("curriculum_semester_export", file_type="excel", download=1, **export_args), "icon": "fa-file-excel", "class": "btn-outline"},
            ],
        )
        preview_semester_rows = curriculum_semester_rows(class_name, academic_year, semester)
        preview_pdf = build_semester_pdf(
            preview_semester_rows,
            class_name=class_name,
            academic_year=academic_year,
            semester=semester,
            teacher_name=class_teacher(class_name) or "Belum diisi",
            principal=PRINCIPAL,
            date_text=format_date_id(jakarta_now().date()),
            letterhead_path=os.path.join(BASE_DIR, "static", "img", "kop_surat_tpq_hmarisa.png"),
        )
        attach_pdf_page_preview(context, preview_pdf)

    elif document_type == "curriculum-template":
        context.update(
            title="Preview Template Import Silabus",
            subtitle="Kolom template sudah dipetakan satu-ke-satu dengan Tambah Rencana Bulanan.",
            preview_kind="table",
            back_url=url_for("curriculum_import"),
            orientation="Excel landscape",
            preview_headers=[
                "Kelas", "Bulan", "Tahun Ajaran", "Semester", "Bidang Pelajaran",
                "Materi Pokok", "Target Pembelajaran", "Rencana Materi Pekan 1",
                "Rencana Materi Pekan 2", "Rencana Materi Pekan 3",
                "Rencana Materi Pekan 4", "Rencana Materi Pekan 5",
                "Catatan Guru", "Status", "Versi",
            ],
            preview_rows=[["Ar Rahim", "Juli", "2026/2027", "Semester 1", "BTQ", "Huruf berharakat fathah", "Santri mampu mengenal dan membaca huruf berharakat fathah", "Pengenalan", "Latihan", "Penguatan", "Evaluasi", "Pengulangan", "Belum diisi", "Aktif", "1"]],
            download_buttons=[{
                "label": "Unduh Template Excel", "url": url_for("curriculum_import_template", download=1),
                "icon": "fa-file-excel", "class": "btn-primary",
            }],
            notice="Baris contoh hanya panduan dan dapat dihapus sebelum import.",
        )

    elif document_type == "students-database":
        if current_user.is_teacher:
            abort(403)
        class_filter = normalize_class_name(request.args.get("class_name", ""))
        status_filter = request.args.get("status", "Semua Status")
        q = request.args.get("q", "").strip()
        export_args = {"class_name": class_filter, "status": status_filter, "q": q}
        count = student_export_query(class_filter, status_filter, q).count()
        context.update(
            title="Preview Database Santri",
            subtitle=f"{class_filter or 'Semua Kelas'} · {status_filter} · {count} data",
            preview_url=url_for("export_students_pdf", download=0, **export_args),
            print_url=url_for("export_students_pdf", download=0, **export_args),
            orientation="A4 landscape",
            back_url=url_for("students", tab="database", **export_args),
            edit_url=url_for("students", tab="database", **export_args),
            download_buttons=[
                {"label": "Unduh PDF", "url": url_for("export_students_pdf", download=1, **export_args), "icon": "fa-file-pdf", "class": "btn-primary"},
                {"label": "Unduh Excel", "url": url_for("export_students_excel", download=1, **export_args), "icon": "fa-file-excel", "class": "btn-outline"},
                {"label": "Unduh CSV", "url": url_for("export_students", download=1, **export_args), "icon": "fa-file-csv", "class": "btn-outline"},
            ],
        )
        attach_pdf_page_preview(
            context,
            build_students_database_pdf(class_filter, status_filter, q),
        )

    elif document_type == "students-template":
        if current_user.is_teacher:
            abort(403)
        context.update(
            title="Preview Template Import Santri",
            subtitle="Periksa nama dan urutan kolom sebelum mengisi file Excel.",
            preview_kind="table",
            back_url=url_for("students", tab="database"),
            preview_headers=["Nama Lengkap", "Nama Tampilan Publik", "NIS", "Kelas", "Nama Wali", "Nomor Kontak Wali", "Tanggal Masuk"],
            preview_rows=[["Contoh Santri", "Contoh", "Belum diisi", "Ar Rahman", "Nama Wali", "08123456789", date.today().isoformat()]],
            download_buttons=[{
                "label": "Unduh Template Excel", "url": url_for("student_import_template", download=1),
                "icon": "fa-file-excel", "class": "btn-primary",
            }],
        )

    elif document_type == "poster":
        filename = secure_filename(request.args.get("filename", ""))
        path = os.path.join(UPLOAD_DIR, "posters", filename)
        if not filename or not os.path.exists(path):
            abort(404)
        context.update(
            title="Preview Poster Santri Terbaik",
            subtitle=filename,
            preview_kind="image",
            preview_url=url_for("poster_file", filename=filename),
            print_url=url_for("poster_file", filename=filename),
            back_url=url_for("dashboard") + "#poster-santri-terbaik",
            download_buttons=[{
                "label": "Unduh Gambar", "url": url_for("poster_file", filename=filename, download=1),
                "icon": "fa-image", "class": "btn-primary",
            }],
        )

    elif document_type == "database-backup":
        if not current_user.is_superadmin:
            abort(403)
        db_path = os.path.join(app.instance_path, "tpq_hmarisa.db")
        if not os.path.exists(db_path):
            abort(404)
        context.update(
            title="Preview Cadangan Database",
            subtitle="Periksa ringkasan sebelum mengunduh berkas SQL.",
            preview_kind="summary",
            back_url=url_for("dashboard"),
            summary_items=[
                ("Nama file", f"tpq_hmarisa_{date.today().isoformat()}.sql"),
                ("Santri", str(Santri.query.count())),
                ("Silabus", str(WeeklyCurriculum.query.count())),
                ("Raport", str(Raport.query.count())),
                ("Data iuran", str(Iuran.query.count())),
                ("Kitab", str(Kitab.query.count())),
            ],
            notice="Berkas SQL berisi data internal. Simpan di tempat aman dan jangan dibagikan melalui grup publik.",
            download_buttons=[{
                "label": "Unduh Cadangan SQL", "url": url_for("backup_database", download=1),
                "icon": "fa-database", "class": "btn-primary",
            }],
        )

    else:
        abort(404)

    return render_template("document_preview.html", **context)


@app.route("/preview/pdf/<token>/page/<int:page_number>.png")
@admin_required
def document_preview_page(token, page_number):
    if not re.fullmatch(r"[a-f0-9]{32}", token or "") or page_number < 1:
        abort(404)
    page_path = os.path.join(
        app.instance_path,
        "document_preview_pages",
        token,
        f"page-{page_number}.png",
    )
    manifest_path = os.path.join(
        app.instance_path,
        "document_preview_pages",
        token,
        "manifest.json",
    )
    if not os.path.exists(page_path) or not os.path.exists(manifest_path):
        abort(404)
    try:
        manifest = json.loads(Path(manifest_path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        abort(404)
    if str(manifest.get("owner_id", "")) != str(getattr(current_user, "id", "")):
        abort(403)
    return send_file(
        page_path,
        mimetype="image/png",
        conditional=True,
        max_age=3600,
    )


@app.route("/daily-progress")
@admin_required
def daily_progress():
    classes = active_class_names()
    class_filter = normalize_class_name(request.args.get("class_name", ""))
    student_id = request.args.get("student_id", type=int)
    status_filter = request.args.get("status", "Aktif")

    rows = []
    total_matches = 0
    class_selected = class_filter in classes
    if class_selected:
        query = Santri.query.filter_by(class_name=class_filter)
        if status_filter == "Aktif":
            query = query.filter_by(is_active=True)
        elif status_filter == "Nonaktif":
            query = query.filter_by(is_active=False)
        if student_id:
            query = query.filter_by(id=student_id)
        total_matches = query.count()
        rows = query.order_by(func.lower(Santri.name)).limit(SEARCH_LIMIT).all()
    else:
        student_id = None

    students_by_class = {
        name: [
            {"id": st.id, "name": st.name, "nis": st.nis, "is_active": bool(st.is_active)}
            for st in Santri.query.filter_by(class_name=name)
            .order_by(func.lower(Santri.name)).all()
        ]
        for name in classes
    }
    return render_template(
        "daily_progress.html",
        students=rows,
        class_filter=class_filter if class_selected else "",
        class_selected=class_selected,
        student_id=student_id,
        status_filter=status_filter,
        students_by_class=students_by_class,
        total_matches=total_matches,
        search_limit=SEARCH_LIMIT,
    )


@app.route("/eraport")
@admin_required
def eraport():
    classes = active_class_names()
    class_filter = normalize_class_name(request.args.get("class_name", ""))
    student_id = request.args.get("student_id", type=int)
    academic_year = request.args.get("academic_year", "2026/2027")
    semester = request.args.get("semester", "Semester 1")
    status_filter = request.args.get("status", "Semua Status")
    query = Santri.query.filter_by(is_active=True)
    if class_filter in classes: query = query.filter_by(class_name=class_filter)
    else: query = query.filter(text("0=1"))
    if student_id: query = query.filter_by(id=student_id)
    students = query.order_by(func.lower(Santri.name)).all()
    report_rows = []
    for student in students:
        raport = get_or_create_raport(student)
        raport.academic_year = academic_year; raport.semester = semester
        raport.completeness = report_completeness(student, raport)
        if status_filter != "Semua Status" and raport.status != status_filter: continue
        report_rows.append((student, raport))
    db.session.commit()
    stats = {"total": len(students), "Belum Dibuat":0, "Draf":0, "Menunggu Pemeriksaan":0, "Diterbitkan":0, "Direvisi":0}
    for student in students:
        status = get_or_create_raport(student).status or "Draf"; stats[status] = stats.get(status,0)+1
    students_by_class = {name: [{"id": st.id, "name": st.name, "nis": st.nis}
                                for st in Santri.query.filter_by(class_name=name, is_active=True)
                                .order_by(func.lower(Santri.name)).all()] for name in classes}
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    return render_template("eraport.html", class_filter=class_filter, student_id=student_id,
                           academic_year=academic_year, semester=semester, status_filter=status_filter,
                           report_rows=report_rows, stats=stats, students_by_class=students_by_class,
                           academic_years=years)


@app.route("/eraport/bulk", methods=["GET", "POST"])
@admin_required
def eraport_bulk():
    class_name = normalize_class_name(request.values.get("class_name", ""))
    academic_year = request.values.get("academic_year", "2026/2027")
    semester = request.values.get("semester", "Semester 1")
    if class_name not in active_class_names():
        flash("Pilih kelas terlebih dahulu.", "warning")
        return redirect(url_for("eraport"))
    students = Santri.query.filter_by(class_name=class_name, is_active=True).order_by(func.lower(Santri.name)).all()
    subjects = class_subjects(class_name)
    if request.method == "POST":
        for student in students:
            raport = get_or_create_raport(student)
            scores = raport.scores()
            for subject in subjects:
                raw = request.form.get(f"score_{student.id}_{subject}", "").strip()
                if raw:
                    value = int(raw)
                    if 60 <= value <= 100:
                        scores[subject] = value
            raport.scores_json = json.dumps(scores, ensure_ascii=False)
            raport.academic_year = academic_year
            raport.semester = semester
            raport.status = "Draf"
            raport.completeness = report_completeness(student, raport)
        db.session.commit()
        flash("Nilai satu kelas berhasil disimpan sebagai draf.", "success")
        return redirect(url_for("eraport", class_name=class_name, academic_year=academic_year, semester=semester))
    rows = [(st, get_or_create_raport(st)) for st in students]
    db.session.commit()
    return render_template("eraport_bulk.html", class_name=class_name, academic_year=academic_year,
                           semester=semester, subjects=subjects, rows=rows)

@app.route("/ananda/<int:student_id>")
@login_required
def guardian_student_detail(student_id):
    """Halaman perkembangan ananda tanpa nilai dan tanpa E-Raport."""
    if current_user.is_admin:
        return redirect(url_for("student_detail", student_id=student_id))
    student = db.get_or_404(Santri, student_id)
    selected = selected_guardian_student()
    if not selected or selected.id != student.id:
        abort(403)
    raport = get_or_create_raport(student)
    db.session.commit()
    bills = (Iuran.query
             .filter_by(santri_id=student.id)
             .order_by(Iuran.year.desc(), Iuran.id.desc())
             .all())
    return render_template(
        "guardian_student.html",
        student=student,
        raport=raport,
        mutabaah=raport.mutabaah(),
        hafalan=raport.hafalan(),
        bills=bills,
    )


@app.route("/student/<int:student_id>")
@admin_required
def student_detail(student_id):
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    db.session.commit()
    bills = Iuran.query.filter_by(santri_id=student.id).order_by(Iuran.year.desc()).all()
    return render_template("student_detail.html", student=student, raport=raport, bills=bills)


@app.route("/student/<int:student_id>/toggle", methods=["POST"])
@admin_required
def toggle_student(student_id):
    student = db.get_or_404(Santri, student_id)
    student.is_active = not student.is_active
    db.session.commit()
    flash(f"{student.name} telah {'diaktifkan kembali' if student.is_active else 'dinonaktifkan/diarsipkan'}.", "success")
    return redirect(request.referrer or url_for("students"))


@app.route("/student/<int:student_id>/delete", methods=["POST"])
@admin_required
def delete_student(student_id):
    student = db.get_or_404(Santri, student_id)
    confirmation = request.form.get("confirmation", "")
    if confirmation != "HAPUS PERMANEN":
        flash("Konfirmasi penghapusan permanen tidak sesuai.", "danger")
        return redirect(request.referrer or url_for("students"))
    name = student.name
    db.session.delete(student); db.session.commit()
    flash(f"Data {name} dihapus permanen.", "info")
    return redirect(url_for("students"))


@app.route("/students/export.csv")
@admin_required
def export_students():
    class_filter = normalize_class_name(request.args.get("class_name", ""))
    status_filter = request.args.get("status", "Semua Status")
    query = Santri.query
    if class_filter in active_class_names(include_inactive=True): query = query.filter_by(class_name=class_filter)
    if status_filter == "Aktif": query = query.filter_by(is_active=True)
    elif status_filter == "Nonaktif": query = query.filter_by(is_active=False)
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(["NIS","Nama Santri","Nama Publik","Kelas","Status","Nama Wali","Nomor Kontak","Tanggal Masuk"])
    for st in query.order_by(Santri.class_name, Santri.name).all():
        writer.writerow([st.nis, st.name, st.public_name or st.name, st.class_name,
                         "Aktif" if st.is_active else "Nonaktif", st.guardian_name or st.guardian.full_name,
                         st.guardian_phone or "", st.joined_date or ""])
    slug = (class_filter or "semua_kelas").lower().replace(" ","_")
    return Response(output.getvalue().encode("utf-8-sig"), mimetype="text/csv",
                    headers={"Content-Disposition":f"attachment; filename=database_santri_{slug}_{date.today().isoformat()}.csv"})



@app.route("/student/<int:student_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_student(student_id):
    student = db.get_or_404(Santri, student_id)
    classes = active_class_names(include_inactive=True)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        nis = request.form.get("nis", "").strip()
        class_name = normalize_class_name(request.form.get("class_name", ""))
        if not name or not nis or class_name not in classes:
            flash("Nama, NIS, dan kelas wajib diisi.", "danger")
            return render_template("student_edit.html", student=student, classes=classes), 400
        duplicate_nis = Santri.query.filter(Santri.nis == nis, Santri.id != student.id).first()
        if duplicate_nis:
            flash("NIS tersebut sudah digunakan oleh santri lain.", "danger")
            return render_template("student_edit.html", student=student, classes=classes), 400
        student.name = name
        student.public_name = request.form.get("public_name", "").strip() or name
        student.nis = nis
        student.class_name = class_name
        student.guardian_name = request.form.get("guardian_name", "").strip()
        student.guardian_phone = request.form.get("guardian_phone", "").strip()
        student.joined_date = parse_date(request.form.get("joined_date"), student.joined_date or date.today())
        student.is_active = request.form.get("is_active") == "1"
        if student.guardian:
            student.guardian.full_name = student.guardian_name or student.guardian.full_name
        db.session.commit()
        flash(f"Data {student.name} berhasil diperbarui.", "success")
        return redirect(url_for("students", tab="database", class_name=student.class_name, student_id=student.id, status="Semua Status"))
    return render_template("student_edit.html", student=student, classes=classes)


@app.route("/students/template.xlsx")
@admin_required
def student_import_template():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Template Santri"
    headers = ["Nama Lengkap", "Nama Tampilan Publik", "NIS", "Kelas", "Nama Wali", "Nomor Kontak Wali", "Tanggal Masuk"]
    sheet.append(headers)
    sheet.append(["Contoh Santri", "Contoh", "", "Ar Rahman", "Nama Wali", "08123456789", date.today().isoformat()])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B6B50")
        cell.alignment = Alignment(horizontal="center")
    widths = [26, 24, 14, 18, 24, 22, 18]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name="template_import_santri_tpq_hmarisa.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/students/export.xlsx")
@admin_required
def export_students_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    class_filter = normalize_class_name(request.args.get("class_name", ""))
    status_filter = request.args.get("status", "Semua Status")
    q = request.args.get("q", "").strip()
    rows = student_export_query(class_filter, status_filter, q).order_by(Santri.class_name, func.lower(Santri.name)).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Database Santri"
    headers = ["NIS", "Nama Santri", "Nama Publik", "Kelas", "Status", "Nama Wali", "Nomor Kontak", "Tanggal Masuk"]
    sheet.append(headers)
    for student in rows:
        sheet.append([
            student.nis,
            student.name,
            student.public_name or student.nickname or "",
            student.class_name,
            "Aktif" if student.is_active else "Nonaktif",
            student.guardian_name or (student.guardian.full_name if student.guardian else ""),
            student.guardian_phone or "",
            student.joined_date.isoformat() if student.joined_date else "",
        ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B6B50")
        cell.alignment = Alignment(horizontal="center")
    widths = [14, 28, 22, 18, 14, 26, 22, 18]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width
    sheet.freeze_panes = "A2"
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    slug = (class_filter or "semua_kelas").lower().replace(" ", "_")
    filename = f"database_santri_{slug}_{date.today().isoformat()}.xlsx"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def build_students_database_pdf(class_filter="", status_filter="Semua Status", q=""):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    rows = student_export_query(class_filter, status_filter, q).order_by(
        Santri.class_name, func.lower(Santri.name)
    ).all()
    stream = io.BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("DATABASE SANTRI TPQ HMarisa", styles["Title"]),
        Paragraph(
            f"Kelas: {class_filter or 'Semua Kelas'} &nbsp;&nbsp; Status: {status_filter}",
            styles["Normal"],
        ),
        Spacer(1, 8),
    ]
    data = [["No", "NIS", "Nama Santri", "Kelas", "Status", "Nama Wali", "Nomor Kontak", "Tanggal Masuk"]]
    for number, student in enumerate(rows, start=1):
        data.append([
            number,
            student.nis,
            student.name,
            student.class_name,
            "Aktif" if student.is_active else "Nonaktif",
            student.guardian_name or (student.guardian.full_name if student.guardian else ""),
            student.guardian_phone or "-",
            format_date_id(student.joined_date),
        ])
    table = Table(
        data,
        repeatRows=1,
        colWidths=[12 * mm, 23 * mm, 47 * mm, 30 * mm, 22 * mm, 45 * mm, 35 * mm, 32 * mm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B6B50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), .4, colors.HexColor("#D7E3DD")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F8F6")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    doc.build(story)
    stream.seek(0)
    return stream


@app.route("/students/export.pdf")
@admin_required
def export_students_pdf():
    class_filter = normalize_class_name(request.args.get("class_name", ""))
    status_filter = request.args.get("status", "Semua Status")
    q = request.args.get("q", "").strip()
    stream = build_students_database_pdf(class_filter, status_filter, q)
    slug = (class_filter or "semua_kelas").lower().replace(" ", "_")
    filename = f"database_santri_{slug}_{date.today().isoformat()}.pdf"
    return send_file(
        stream,
        as_attachment=download_requested(),
        download_name=filename,
        mimetype="application/pdf",
    )


JOURNAL_NOTE_TEMPLATES = {
    "Lanjut": "Alhamdulillah, Ananda bisa mengikuti pelajaran hari ini. Semangat terus belajarnya.",
    "Cukup": "Alhamdulillah, Ananda bisa mengikuti pelajaran hari ini, tetapi masih membutuhkan bimbingan lagi. Semangat terus belajarnya.",
    "Ulangi": "Alhamdulillah, Ananda bisa mengikuti pelajaran hari ini, tetapi masih membutuhkan bimbingan khusus dan pendampingan lagi di rumah. Semangat terus belajarnya.",
}


def journal_material_rows(student):
    """Materi Mutabaah hanya bersumber dari Silabus Bulanan hasil import terbaru.

    Data CurriculumBank lama sengaja tidak digunakan. Setelah database Silabus
    dikosongkan, dropdown akan kosong. Begitu template Silabus baru diimpor,
    data otomatis muncul berdasarkan kelas, bulan/tahun, dan bidang pelajaran.
    """
    rows = (
        WeeklyCurriculum.query
        .filter_by(class_name=student.class_name, status="Aktif")
        .filter(WeeklyCurriculum.source_type == "Import Excel")
        .order_by(WeeklyCurriculum.year, WeeklyCurriculum.month, WeeklyCurriculum.subject, WeeklyCurriculum.topic)
        .all()
    )
    result = []
    seen = set()
    for row in rows:
        key = (row.month, row.year, row.subject, row.topic)
        if not row.topic or key in seen:
            continue
        seen.add(key)
        result.append({
            "month": row.month,
            "year": row.year,
            "academic_year": row.academic_year,
            "subject": row.subject,
            "topic": row.topic,
        })
    return result


def journal_form_payload(form):
    subject = form.get("subject", "").strip()
    material = form.get("material", "").strip() or form.get("manual_material", "").strip()
    tahsin_level = form.get("tahsin_level", "").strip()
    tilawati = form.get("tilawati", "").strip()
    alquran = form.get("alquran", "").strip()
    if tahsin_level == "Tilawati/Iqro":
        alquran = ""
    elif tahsin_level == "Al-Qur'an":
        tilawati = ""
    else:
        tahsin_level = ""
        tilawati = ""
        alquran = ""
    keterangan = form.get("keterangan", "").strip()
    notes = form.get("notes", "").strip() or JOURNAL_NOTE_TEMPLATES.get(keterangan, "")
    return {
        "date": form.get("date") or date.today().isoformat(),
        "subject": subject,
        "material": material,
        "tahsin_level": tahsin_level,
        "tilawati": tilawati,
        "alquran": alquran,
        "keterangan": keterangan,
        "notes": notes,
        "attendance": form.get("attendance", "Hadir"),
        "on_time": bool(form.get("on_time")),
        "adab_score": max(60, min(100, form.get("adab_score", type=int) or 80)),
    }


def validate_journal_payload(payload):
    if not payload["subject"]:
        return "Bidang Pelajaran wajib dipilih."
    if not payload["material"]:
        return "Materi Pelajaran wajib dipilih atau diisi manual."
    if payload["tahsin_level"] not in {"Tilawati/Iqro", "Al-Qur'an"}:
        return "Jenjang Tahsin wajib dipilih."
    if payload["tahsin_level"] == "Tilawati/Iqro" and not payload["tilawati"]:
        return "Kolom Tilawati/Iqro wajib diisi untuk jenjang yang dipilih."
    if payload["tahsin_level"] == "Al-Qur'an" and not payload["alquran"]:
        return "Kolom Al-Qur'an wajib diisi untuk jenjang yang dipilih."
    if payload["keterangan"] not in JOURNAL_NOTE_TEMPLATES:
        return "Keterangan wajib dipilih: Lanjut, Cukup, atau Ulangi."
    return ""


def journal_page_context(student):
    raport = get_or_create_raport(student)
    active_entries = [entry for entry in mutabaah_entries(student) if not entry.get("deleted")]
    return {
        "student": student,
        "raport": raport,
        "subjects": class_subjects(student.class_name),
        "material_rows": journal_material_rows(student),
        "total_journals": len(active_entries),
        "hafalan_done": sum(1 for row in build_hafalan_tracker_rows(student) if row["status"] == "Selesai"),
    }


def tracker_status_label(raw_status):
    if raw_status in {"Sudah Hafal", "Selesai"}:
        return "Selesai"
    if raw_status in {"Sedang Proses", "Perlu Murajaah", "Proses"}:
        return "Proses"
    return "Belum Mulai"


def build_hafalan_tracker_rows(student):
    all_records = (
        HafalanRecord.query
        .filter_by(santri_id=student.id, is_deleted=False)
        .order_by(HafalanRecord.entry_date.asc(), HafalanRecord.id.asc())
        .all()
    )
    grouped = {surah: [] for surah in SURAH_JUZ30}
    for row in all_records:
        if row.surah in grouped:
            grouped[row.surah].append(row)
    legacy = get_or_create_raport(student).hafalan()
    tracker = []
    for number, surah in enumerate(SURAH_JUZ30, 1):
        rows = grouped[surah]
        latest = rows[-1] if rows else None
        status = tracker_status_label(latest.status if latest else ("Sudah Hafal" if legacy.get(surah) else "Belum Hafal"))
        started = next((r.entry_date for r in rows if tracker_status_label(r.status) in {"Proses", "Selesai"}), None)
        completed_rows = [r for r in rows if tracker_status_label(r.status) == "Selesai"]
        completed = completed_rows[-1].entry_date if completed_rows else None
        tracker.append({
            "number": number,
            "surah": surah,
            "status": status,
            "start_date": started,
            "completion_date": completed,
            "score": latest.fluency if latest else None,
            "notes": latest.notes if latest else "",
            "latest_record": latest,
            "history_count": len(rows),
        })
    return tracker


@app.route("/academics/<int:student_id>", methods=["GET", "POST"])
@admin_required
def academics(student_id):
    """Halaman 1: Tambah Jurnal Mutabaah."""
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    entries = mutabaah_entries(student)

    if request.method == "POST":
        payload = journal_form_payload(request.form)
        error = validate_journal_payload(payload)
        if error:
            flash(error, "danger")
        else:
            duplicate = next((
                entry for entry in entries
                if not entry.get("deleted")
                and entry.get("date") == payload["date"]
                and entry.get("subject") == payload["subject"]
                and entry.get("material") == payload["material"]
            ), None)
            if duplicate and request.form.get("allow_duplicate") != "1":
                flash("Jurnal serupa sudah tersedia pada tanggal tersebut. Edit jurnal sebelumnya atau centang simpan duplikat.", "warning")
            else:
                payload.update({
                    "id": uuid.uuid4().hex,
                    "created_by": current_user.full_name,
                    "created_at": datetime.utcnow().isoformat(),
                    "deleted": False,
                })
                entries.insert(0, payload)
                raport.mutabaah_json = json.dumps(entries, ensure_ascii=False)
                db.session.commit()
                flash("Jurnal Mutabaah berhasil ditambahkan.", "success")
                return redirect(url_for("mutabaah_history", student_id=student.id))

    context = journal_page_context(student)
    return render_template("academics.html", **context)


@app.route("/academics/<int:student_id>/history")
@admin_required
def mutabaah_history(student_id):
    """Halaman 2: Riwayat Jurnal Mutabaah dengan filter dan pagination."""
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    month_filter = request.args.get("month", "")
    year_filter = request.args.get("year", type=int)
    subject_filter = request.args.get("subject", "")
    attendance_filter = request.args.get("attendance", "")
    keyword = request.args.get("keyword", "").strip().lower()
    page = max(1, request.args.get("page", type=int) or 1)
    per_page = request.args.get("per_page", type=int) or 10
    if per_page not in {10, 25, 50, 100}:
        per_page = 10

    month_names = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    entries = [entry for entry in mutabaah_entries(student) if not entry.get("deleted")]
    entries.sort(key=lambda entry: (entry.get("date", ""), entry.get("created_at", "")), reverse=True)

    def matches(entry):
        entry_date = parse_date(entry.get("date"))
        if month_filter in month_names:
            if not entry_date or month_names[entry_date.month - 1] != month_filter:
                return False
        if year_filter and (not entry_date or entry_date.year != year_filter):
            return False
        if subject_filter and entry.get("subject") != subject_filter:
            return False
        if attendance_filter and entry.get("attendance") != attendance_filter:
            return False
        if keyword:
            haystack = " ".join([
                entry.get("material", ""), entry.get("tilawati", ""), entry.get("alquran", ""),
                entry.get("keterangan", ""), entry.get("notes", ""), entry.get("subject", ""),
            ]).lower()
            if keyword not in haystack:
                return False
        return True

    filtered = [entry for entry in entries if matches(entry)]
    total_entries = len(filtered)
    page_count = max(1, (total_entries + per_page - 1) // per_page)
    page = min(page, page_count)
    start = (page - 1) * per_page
    paged_entries = filtered[start:start + per_page]
    context = journal_page_context(student)
    context.update({
        "entries": paged_entries,
        "all_total": len(entries),
        "total_entries": total_entries,
        "page": page,
        "page_count": page_count,
        "per_page": per_page,
        "month_filter": month_filter,
        "year_filter": year_filter,
        "subject_filter": subject_filter,
        "attendance_filter": attendance_filter,
        "keyword": keyword,
    })
    return render_template("mutabaah_history.html", **context)


@app.route("/academics/<int:student_id>/mutabaah/<int:entry_index>/detail")
@admin_required
def mutabaah_detail(student_id, entry_index):
    student = db.get_or_404(Santri, student_id)
    entries = mutabaah_entries(student)
    if entry_index < 0 or entry_index >= len(entries) or entries[entry_index].get("deleted"):
        abort(404)
    context = journal_page_context(student)
    context.update({"entry": entries[entry_index], "entry_index": entry_index})
    return render_template("mutabaah_detail.html", **context)


@app.route("/academics/<int:student_id>/mutabaah/<int:entry_index>/edit", methods=["GET", "POST"])
@admin_required
def edit_mutabaah(student_id, entry_index):
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    entries = mutabaah_entries(student)
    if entry_index < 0 or entry_index >= len(entries) or entries[entry_index].get("deleted"):
        abort(404)
    entry = entries[entry_index]
    if request.method == "POST":
        payload = journal_form_payload(request.form)
        error = validate_journal_payload(payload)
        if error:
            flash(error, "danger")
        else:
            entry.update(payload)
            entry.update({"updated_by": current_user.full_name, "updated_at": datetime.utcnow().isoformat()})
            raport.mutabaah_json = json.dumps(entries, ensure_ascii=False)
            db.session.commit()
            flash("Jurnal Mutabaah berhasil diperbarui.", "success")
            return redirect(url_for("mutabaah_detail", student_id=student.id, entry_index=entry_index))
    context = journal_page_context(student)
    context.update({"entry": entry, "entry_index": entry_index})
    return render_template("mutabaah_edit.html", **context)


@app.route("/academics/<int:student_id>/mutabaah/<int:entry_index>/delete", methods=["POST"])
@admin_required
def delete_mutabaah(student_id, entry_index):
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    entries = mutabaah_entries(student)
    if entry_index < 0 or entry_index >= len(entries):
        abort(404)
    entries[entry_index]["deleted"] = True
    entries[entry_index]["deleted_by"] = current_user.full_name
    entries[entry_index]["deleted_at"] = datetime.utcnow().isoformat()
    raport.mutabaah_json = json.dumps(entries, ensure_ascii=False)
    db.session.commit()
    flash("Jurnal diarsipkan. Riwayat tetap tersimpan di database.", "info")
    return redirect(url_for("mutabaah_history", student_id=student.id))


@app.route("/academics/<int:student_id>/hafalan")
@admin_required
def hafalan_tracker(student_id):
    """Halaman 3: Tracker Hafalan Juz 30 yang sederhana dan terpisah."""
    student = db.get_or_404(Santri, student_id)
    status_filter = request.args.get("status", "")
    keyword = request.args.get("keyword", "").strip().lower()
    selected_surah = request.args.get("surah", "")
    tracker_rows = build_hafalan_tracker_rows(student)
    counts = {
        "Selesai": sum(1 for row in tracker_rows if row["status"] == "Selesai"),
        "Proses": sum(1 for row in tracker_rows if row["status"] == "Proses"),
        "Belum Mulai": sum(1 for row in tracker_rows if row["status"] == "Belum Mulai"),
    }
    filtered_rows = [
        row for row in tracker_rows
        if (not status_filter or row["status"] == status_filter)
        and (not keyword or keyword in row["surah"].lower() or keyword in (row["notes"] or "").lower())
    ]
    context = journal_page_context(student)
    context.update({
        "tracker_rows": filtered_rows,
        "tracker_total": len(tracker_rows),
        "counts": counts,
        "status_filter": status_filter,
        "keyword": keyword,
        "selected_surah": selected_surah,
    })
    return render_template("hafalan_tracker.html", **context)


@app.route("/academics/<int:student_id>/hafalan/save", methods=["POST"])
@admin_required
def save_hafalan_tracker(student_id):
    student = db.get_or_404(Santri, student_id)
    surah = request.form.get("surah", "")
    status = request.form.get("status", "Belum Mulai")
    if surah not in SURAH_JUZ30 or status not in {"Belum Mulai", "Proses", "Selesai"}:
        flash("Data tracker hafalan tidak valid.", "danger")
        return redirect(url_for("hafalan_tracker", student_id=student.id))

    start_date = parse_date(request.form.get("start_date"))
    completion_date = parse_date(request.form.get("completion_date"))
    score = request.form.get("score", type=int)
    if score is not None and not 60 <= score <= 100:
        flash("Nilai/kelancaran harus berada pada rentang 60–100.", "danger")
        return redirect(url_for("hafalan_tracker", student_id=student.id, surah=surah))
    notes = request.form.get("notes", "").strip()
    existing = HafalanRecord.query.filter_by(santri_id=student.id, surah=surah, is_deleted=False).first()

    if status == "Proses":
        entry_date = start_date or date.today()
        raw_status = "Sedang Proses"
        activity = "Mulai Hafalan" if not existing else "Update Progres"
    elif status == "Selesai":
        entry_date = completion_date or date.today()
        raw_status = "Sudah Hafal"
        activity = "Tandai Selesai"
        if start_date and not existing and start_date < entry_date:
            db.session.add(HafalanRecord(
                santri_id=student.id, surah=surah, activity_type="Mulai Hafalan",
                entry_date=start_date, status="Sedang Proses", notes="Mulai tracker hafalan.",
                created_by=current_user.full_name,
            ))
    else:
        entry_date = date.today()
        raw_status = "Belum Hafal"
        activity = "Reset Status"

    row = HafalanRecord(
        santri_id=student.id,
        surah=surah,
        activity_type=activity,
        entry_date=entry_date,
        status=raw_status,
        fluency=score,
        notes=notes,
        created_by=current_user.full_name,
    )
    db.session.add(row)
    db.session.commit()
    flash(f"Tracker {surah} berhasil diperbarui menjadi {status}.", "success")
    return redirect(url_for("hafalan_tracker", student_id=student.id, surah=surah))


# Route lama dipertahankan agar tautan lama tidak rusak.
@app.route("/hafalan/<int:record_id>/edit", methods=["POST"])
@admin_required
def edit_hafalan(record_id):
    row = db.get_or_404(HafalanRecord, record_id)
    row.activity_type = request.form.get("activity_type", row.activity_type)
    row.entry_date = parse_date(request.form.get("entry_date"), row.entry_date)
    row.status = request.form.get("hafalan_status", row.status)
    row.fluency = request.form.get("fluency", type=int)
    row.tajwid = request.form.get("tajwid", type=int)
    row.makhraj = request.form.get("makhraj", type=int)
    row.notes = request.form.get("hafalan_notes", "").strip()
    db.session.commit()
    flash("Setoran hafalan terakhir diperbarui.", "success")
    return redirect(url_for("hafalan_tracker", student_id=row.santri_id, surah=row.surah))


@app.route("/hafalan/<int:record_id>/delete", methods=["POST"])
@admin_required
def delete_hafalan(record_id):
    row = db.get_or_404(HafalanRecord, record_id)
    row.is_deleted = True
    db.session.commit()
    flash("Setoran hafalan diarsipkan.", "info")
    return redirect(url_for("hafalan_tracker", student_id=row.santri_id, surah=row.surah))


@app.route("/hafalan/<int:student_id>/<path:surah>/reset", methods=["POST"])
@admin_required
def reset_hafalan(student_id, surah):
    student = db.get_or_404(Santri, student_id)
    if surah not in SURAH_JUZ30:
        abort(404)
    row = HafalanRecord(
        santri_id=student.id, surah=surah, activity_type="Reset Status",
        entry_date=date.today(), status="Belum Hafal",
        notes="Status direset tanpa menghapus riwayat.", created_by=current_user.full_name,
    )
    db.session.add(row)
    db.session.commit()
    flash("Status surat dikembalikan menjadi Belum Mulai.", "info")
    return redirect(url_for("hafalan_tracker", student_id=student.id, surah=surah))


@app.route("/eraport/<int:student_id>/edit", methods=["GET","POST"])
@admin_required
def report_edit(student_id):
    student=db.get_or_404(Santri,student_id); raport=get_or_create_raport(student); subjects=class_subjects(student.class_name)
    if request.method=="POST":
        action=request.form.get("action","save_draft")
        scores={}
        for subject in subjects:
            raw=request.form.get(f"score_{subject}","").strip()
            if raw:
                value=int(raw)
                if value<60 or value>100:
                    flash("Nilai harus berada pada rentang 60–100.","danger"); return redirect(url_for("report_edit",student_id=student.id))
                scores[subject]=value
        raport.scores_json=json.dumps(scores,ensure_ascii=False)
        raport.semester=request.form.get("semester","Semester 1"); raport.academic_year=request.form.get("academic_year","2026/2027")
        attitude={key:request.form.get(f"attitude_{key}","") for key in ["Kehadiran","Kedisiplinan","Keterlibatan","Pergaulan/Perilaku"]}
        raport.attitude_json=json.dumps(attitude,ensure_ascii=False)
        absence={key:{"count":request.form.get(f"absence_{key}",type=int) or 0,"notes":request.form.get(f"absence_note_{key}","").strip()} for key in ["Sakit","Izin","Tanpa Keterangan","Keterangan Lain"]}
        raport.absence_json=json.dumps(absence,ensure_ascii=False)
        raport.development_notes=request.form.get("development_notes","").strip(); raport.publish_date=parse_date(request.form.get("publish_date"),raport.publish_date)
        raport.completeness=report_completeness(student,raport)
        if action=="send_review": raport.status="Menunggu Pemeriksaan"
        elif action=="publish":
            if raport.completeness<100:
                flash("Raport belum lengkap dan belum dapat diterbitkan.","danger"); db.session.commit(); return redirect(url_for("report_edit",student_id=student.id))
            raport.status="Diterbitkan"; raport.published_at=datetime.utcnow()
            latest=current_hafalan_status(student.id); done=sum(1 for s in SURAH_JUZ30 if latest.get(s) and latest[s].status=="Sudah Hafal")
            raport.snapshot_json=json.dumps({"student_name":student.name,"nis":student.nis,"class_name":student.class_name,"teacher":class_teacher(student.class_name),"principal":PRINCIPAL,"scores":scores,"attitude":attitude,"absence":absence,"development_notes":raport.development_notes,"hafalan_done":done,"academic_year":raport.academic_year,"semester":raport.semester,"publish_date":str(raport.publish_date)},ensure_ascii=False)
        elif action=="revise": raport.status="Direvisi"; raport.version=(raport.version or 1)+1
        else: raport.status="Draf"
        db.session.commit(); flash("Data E-Raport berhasil disimpan.","success")
        if action=="publish": return redirect(url_for("report_preview",student_id=student.id))
    latest=current_hafalan_status(student.id); legacy=raport.hafalan(); done=sum(1 for s in SURAH_JUZ30 if (latest.get(s) and latest[s].status=="Sudah Hafal") or (not latest.get(s) and legacy.get(s)))
    raport.completeness=report_completeness(student,raport); db.session.commit()
    return render_template("report_edit.html",student=student,raport=raport,subjects=subjects,scores=raport.scores(),attitude=raport.attitude(),absence=raport.absence(),hafalan_done=done)


@app.route("/eraport/<int:student_id>/publish", methods=["POST"])
@admin_required
def report_publish(student_id):
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    raport.completeness = report_completeness(student, raport)
    if raport.completeness < 100:
        flash("Raport belum lengkap dan belum dapat diterbitkan.", "danger")
        return redirect(url_for("report_preview", student_id=student.id))
    raport.status = "Diterbitkan"
    raport.published_at = datetime.utcnow()
    if not raport.publish_date:
        raport.publish_date = jakarta_now().date()
    latest = current_hafalan_status(student.id)
    legacy = raport.hafalan()
    done = sum(1 for surah in SURAH_JUZ30 if
               (latest.get(surah) and latest[surah].status == "Sudah Hafal") or
               (not latest.get(surah) and legacy.get(surah)))
    raport.snapshot_json = json.dumps({
        "student_name": student.name, "nis": student.nis, "class_name": student.class_name,
        "teacher": class_teacher(student.class_name), "principal": PRINCIPAL,
        "scores": raport.scores(), "attitude": raport.attitude(), "absence": raport.absence(),
        "development_notes": raport.development_notes, "hafalan_done": done,
        "academic_year": raport.academic_year, "semester": raport.semester,
        "publish_date": str(raport.publish_date), "version": raport.version,
    }, ensure_ascii=False)
    db.session.commit()
    flash("Raport berhasil diterbitkan dan kini dapat dilihat di Portal Wali.", "success")
    return redirect(url_for("report_preview", student_id=student.id))


@app.route("/eraport/<int:student_id>/revise", methods=["POST"])
@admin_required
def report_revise(student_id):
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    raport.version = (raport.version or 1) + 1
    raport.status = "Direvisi"
    db.session.commit()
    flash("Mode revisi dibuka. Versi sebelumnya tetap tersimpan sebagai snapshot.", "info")
    return redirect(url_for("report_edit", student_id=student.id))


@app.route("/eraport/<int:student_id>/preview")
@admin_required
def report_preview(student_id):
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    latest = current_hafalan_status(student.id)
    legacy = raport.hafalan()
    done = sum(1 for surah in SURAH_JUZ30 if
               (latest.get(surah) and latest[surah].status == "Sudah Hafal") or
               (not latest.get(surah) and legacy.get(surah)))
    return render_template(
        "report_preview.html",
        student=student,
        raport=raport,
        hafalan_done=done,
        teacher=class_teacher(student.class_name),
        subjects=class_subjects(student.class_name),
        scores=raport.scores(),
        attitude=raport.attitude(),
        absence=raport.absence(),
        current_date=jakarta_now().date(),
    )


def _report_visual_data(student, raport, include_draft=True):
    latest = current_hafalan_status(student.id)
    legacy = raport.hafalan()
    done = sum(1 for surah in SURAH_JUZ30 if
               (latest.get(surah) and latest[surah].status == "Sudah Hafal") or
               (not latest.get(surah) and legacy.get(surah)))
    pub_date = raport.publish_date or jakarta_now().date()
    return {
        "student_name": student.name,
        "nis": student.nis,
        "class_name": student.class_name,
        "semester": raport.semester or "Semester 1",
        "academic_year": raport.academic_year or "2026/2027",
        "teacher": class_teacher(student.class_name),
        "principal": PRINCIPAL,
        "subjects": class_subjects(student.class_name),
        "scores": raport.scores(),
        "attitude": raport.attitude(),
        "absence": raport.absence(),
        "development_notes": raport.development_notes,
        "hafalan_done": done,
        "hafalan_total": len(SURAH_JUZ30),
        "kkm": 70,
        "publish_date": format_date_id(pub_date),
        "address": "Jl. Kayu Gede 2, Paku Jaya, Kec. Serpong Utara, Kota Tangerang, Banten 15220",
        "is_draft": include_draft and raport.status != "Diterbitkan",
    }


def build_report_png(student, raport, include_draft=True):
    template = os.path.join(BASE_DIR, "static", "img", "eraport_reference.png")
    return render_report_image(template, _report_visual_data(student, raport, include_draft=include_draft))


@app.route("/eraport/<int:student_id>/image")
@admin_required
def report_image(student_id):
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    png = build_report_png(student, raport, include_draft=True)
    response = send_file(png, mimetype="image/png", download_name=f"preview_raport_{student.id}.png")
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return response


def build_report_pdf(student, raport):
    """Embed the exact approved preview PNG as one A4 portrait PDF page."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader

    png = build_report_png(student, raport, include_draft=False)
    buffer = io.BytesIO()
    page_width, page_height = A4
    pdf = canvas.Canvas(buffer, pagesize=A4, pageCompression=1)
    pdf.setTitle(f"Raport {student.name} - TPQ HMarisa")
    pdf.drawImage(ImageReader(png), 0, 0, width=page_width, height=page_height,
                  preserveAspectRatio=False, mask="auto")
    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer


@app.route("/report/<int:student_id>/pdf")
@admin_required
def report_pdf(student_id):
    student = db.get_or_404(Santri, student_id)
    raport = get_or_create_raport(student)
    db.session.commit()
    if raport.status != "Diterbitkan":
        flash("Raport harus diterbitkan terlebih dahulu sebelum diunduh atau dicetak.", "warning")
        return redirect(url_for("report_preview", student_id=student.id))
    pdf = build_report_pdf(student, raport)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", student.name)
    return send_file(pdf, mimetype="application/pdf", as_attachment=download_requested(),
                     download_name=f"E-Raport_{safe_name}.pdf")


@app.route("/data-master", methods=["GET", "POST"])
@admin_required
def data_master():
    allowed_tabs = {"classes", "teachers", "subjects", "years", "users"}
    tab = request.args.get("tab", request.form.get("tab", "classes"))
    if tab not in allowed_tabs:
        tab = "classes"
    if tab == "users" and not current_user.is_superadmin:
        abort(403)

    academic_year_options = [f"{year}/{year + 1}" for year in range(2026, 2050)]
    teacher_positions = ["Guru", "Wali Kelas", "Guru dan Wali Kelas", "Kepala TPQ", "Guru dan Kepala TPQ"]

    class_filter_options = active_class_names()
    subject_class_filter = normalize_class_name(
        request.args.get("subject_class", request.form.get("subject_class_filter", ""))
    )
    if subject_class_filter not in class_filter_options:
        subject_class_filter = class_filter_options[0] if class_filter_options else ""

    stored_year_names = [
        row[0] for row in db.session.query(AcademicYear.name)
        .distinct().order_by(AcademicYear.name.desc()).all()
    ]
    primary_year = current_academic_year()
    year_filter = request.args.get(
        "year_filter", request.form.get("year_filter_value", "")
    ).strip()
    if year_filter not in stored_year_names:
        if primary_year and primary_year.name in stored_year_names:
            year_filter = primary_year.name
        elif stored_year_names:
            year_filter = stored_year_names[0]
        else:
            year_filter = academic_year_options[0]
    year_filter_options = stored_year_names or academic_year_options

    def year_dates(year_name, semester):
        try:
            first_year, second_year = [int(part) for part in year_name.split("/")]
        except (TypeError, ValueError):
            raise ValueError("Format tahun ajaran tidak valid.")
        if semester == "Semester 1":
            return date(first_year, 7, 1), date(first_year, 12, 31)
        if semester == "Semester 2":
            return date(second_year, 1, 1), date(second_year, 6, 30)
        raise ValueError("Semester tidak valid.")

    def make_code(value, existing, prefix=""):
        words = re.findall(r"[A-Za-z0-9]+", value.upper())
        base = "".join(word[:1] for word in words) if len(words) > 1 else (words[0][:3] if words else "DATA")
        base = (prefix + base)[:24] or "DATA"
        candidate = base
        number = 2
        while candidate in existing:
            candidate = f"{base}{number}"
            number += 1
        return candidate

    def teacher_display(full_name, title):
        return f"{full_name}, {title}" if title else full_name

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        try:
            if action == "add_class":
                name = normalize_class_name(request.form.get("name", ""))
                if not name:
                    raise ValueError("Nama kelas wajib diisi.")
                if MasterClass.query.filter_by(name=name).first():
                    raise ValueError("Nama kelas sudah digunakan.")
                existing_codes = {row.code for row in MasterClass.query.all()}
                code = make_code(name, existing_codes)
                year_name = request.form.get("academic_year_start", "2026/2027")
                if year_name not in academic_year_options:
                    raise ValueError("Tahun ajaran mulai tidak valid.")
                max_order = db.session.query(func.max(MasterClass.sort_order)).scalar() or 0
                db.session.add(MasterClass(
                    name=name, code=code,
                    teacher_name=request.form.get("teacher_name", "").strip(),
                    sort_order=max_order + 1,
                    academic_year_start=year_name,
                    is_active=True,
                    notes=request.form.get("notes", "").strip(),
                ))
                flash("Kelas baru berhasil ditambahkan.", "success")

            elif action == "update_class":
                row = db.get_or_404(MasterClass, request.form.get("id", type=int))
                old_name = row.name
                new_name = normalize_class_name(request.form.get("name", ""))
                if not new_name:
                    raise ValueError("Nama kelas wajib diisi.")
                duplicate = MasterClass.query.filter(MasterClass.name == new_name, MasterClass.id != row.id).first()
                if duplicate:
                    raise ValueError("Nama kelas sudah digunakan.")
                year_name = request.form.get("academic_year_start", row.academic_year_start)
                if year_name not in academic_year_options:
                    raise ValueError("Tahun ajaran mulai tidak valid.")
                if new_name != old_name:
                    Santri.query.filter_by(class_name=old_name).update({Santri.class_name: new_name}, synchronize_session=False)
                    WeeklyCurriculum.query.filter_by(class_name=old_name).update({WeeklyCurriculum.class_name: new_name}, synchronize_session=False)
                    Subject.query.filter_by(class_name=old_name).update({Subject.class_name: new_name}, synchronize_session=False)
                    CurriculumBank.query.filter_by(class_name=old_name).update({CurriculumBank.class_name: new_name}, synchronize_session=False)
                    Teacher.query.filter_by(class_name=old_name).update({Teacher.class_name: new_name}, synchronize_session=False)
                    User.query.filter_by(assigned_class=old_name).update({User.assigned_class: new_name}, synchronize_session=False)
                    WeeklyWinner.query.filter_by(class_name=old_name).update({WeeklyWinner.class_name: new_name}, synchronize_session=False)
                row.name = new_name
                row.teacher_name = request.form.get("teacher_name", "").strip()
                row.academic_year_start = year_name
                row.notes = request.form.get("notes", "").strip()
                flash("Data kelas berhasil diperbarui.", "success")

            elif action == "toggle_class":
                row = db.get_or_404(MasterClass, request.form.get("id", type=int))
                row.is_active = not row.is_active
                flash("Status kelas berhasil diperbarui.", "success")

            elif action == "delete_class":
                row = db.get_or_404(MasterClass, request.form.get("id", type=int))
                dependencies = {
                    "santri": Santri.query.filter_by(class_name=row.name).count(),
                    "bidang pelajaran": Subject.query.filter_by(class_name=row.name).count(),
                    "silabus": WeeklyCurriculum.query.filter_by(class_name=row.name).count(),
                    "bank silabus": CurriculumBank.query.filter_by(class_name=row.name).count(),
                    "guru": Teacher.query.filter_by(class_name=row.name).count(),
                }
                used = [f"{label} ({count})" for label, count in dependencies.items() if count]
                if used:
                    raise ValueError("Kelas belum dapat dihapus karena masih terhubung dengan " + ", ".join(used) + ". Gunakan Nonaktifkan.")
                db.session.delete(row)
                flash("Kelas berhasil dihapus permanen.", "success")

            elif action == "add_teacher":
                full_name = request.form.get("full_name", "").strip()
                position = request.form.get("position", "")
                class_name = normalize_class_name(request.form.get("class_name", ""))
                if not full_name:
                    raise ValueError("Nama guru wajib diisi.")
                if position not in teacher_positions:
                    raise ValueError("Jabatan tidak valid.")
                title = request.form.get("title", "").strip()
                row = Teacher(
                    full_name=full_name, title=title,
                    phone=request.form.get("phone", "").strip(),
                    class_name=class_name, position=position,
                    is_active=True, notes=request.form.get("notes", "").strip(),
                )
                db.session.add(row)
                if class_name and "Wali Kelas" in position:
                    master = MasterClass.query.filter_by(name=class_name).first()
                    if master:
                        master.teacher_name = teacher_display(full_name, title)
                flash("Data guru berhasil ditambahkan.", "success")

            elif action == "update_teacher":
                row = db.get_or_404(Teacher, request.form.get("id", type=int))
                old_display = teacher_display(row.full_name, row.title)
                old_class = row.class_name
                full_name = request.form.get("full_name", "").strip()
                title = request.form.get("title", "").strip()
                position = request.form.get("position", "")
                class_name = normalize_class_name(request.form.get("class_name", ""))
                if not full_name:
                    raise ValueError("Nama guru wajib diisi.")
                if position not in teacher_positions:
                    raise ValueError("Jabatan tidak valid.")
                new_display = teacher_display(full_name, title)
                if old_class and old_class != class_name:
                    old_master = MasterClass.query.filter_by(name=old_class, teacher_name=old_display).first()
                    if old_master:
                        old_master.teacher_name = ""
                row.full_name = full_name
                row.title = title
                row.phone = request.form.get("phone", "").strip()
                row.class_name = class_name
                row.position = position
                row.notes = request.form.get("notes", "").strip()
                MasterClass.query.filter_by(teacher_name=old_display).update({MasterClass.teacher_name: new_display}, synchronize_session=False)
                CurriculumBank.query.filter_by(teacher_name=old_display).update({CurriculumBank.teacher_name: new_display}, synchronize_session=False)
                if class_name and "Wali Kelas" in position:
                    master = MasterClass.query.filter_by(name=class_name).first()
                    if master:
                        master.teacher_name = new_display
                flash("Data guru berhasil diperbarui.", "success")

            elif action == "toggle_teacher":
                row = db.get_or_404(Teacher, request.form.get("id", type=int))
                row.is_active = not row.is_active
                flash("Status guru berhasil diperbarui.", "success")

            elif action == "delete_teacher":
                row = db.get_or_404(Teacher, request.form.get("id", type=int))
                display_name = teacher_display(row.full_name, row.title)
                if MasterClass.query.filter_by(teacher_name=display_name, is_active=True).first():
                    raise ValueError("Guru masih tercatat sebagai wali kelas aktif. Ubah wali kelas terlebih dahulu.")
                db.session.delete(row)
                flash("Data guru berhasil dihapus permanen.", "success")

            elif action == "add_subject":
                name = request.form.get("name", "").strip()
                class_name = normalize_class_name(request.form.get("class_name", ""))
                if not name or not class_name:
                    raise ValueError("Bidang pelajaran dan kelas wajib diisi.")
                if Subject.query.filter(func.lower(Subject.name) == name.lower(), Subject.class_name == class_name).first():
                    raise ValueError("Bidang pelajaran tersebut sudah ada pada kelas ini.")
                class_row = MasterClass.query.filter_by(name=class_name).first()
                prefix = (class_row.code + "-") if class_row else ""
                existing_codes = {item.code for item in Subject.query.filter_by(class_name=class_name).all()}
                code = make_code(name, existing_codes, prefix)
                max_order = db.session.query(func.max(Subject.sort_order)).filter(Subject.class_name == class_name).scalar() or 0
                db.session.add(Subject(
                    name=name, code=code, class_name=class_name,
                    sort_order=max_order + 1, is_active=True,
                    notes=request.form.get("notes", "").strip(),
                ))
                flash("Bidang pelajaran berhasil ditambahkan.", "success")

            elif action == "update_subject":
                row = db.get_or_404(Subject, request.form.get("id", type=int))
                old_name, old_class = row.name, row.class_name
                new_name = request.form.get("name", "").strip()
                new_class = normalize_class_name(request.form.get("class_name", ""))
                if not new_name or not new_class:
                    raise ValueError("Bidang pelajaran dan kelas wajib diisi.")
                duplicate = Subject.query.filter(func.lower(Subject.name) == new_name.lower(), Subject.class_name == new_class, Subject.id != row.id).first()
                if duplicate:
                    raise ValueError("Bidang pelajaran tersebut sudah ada pada kelas ini.")
                WeeklyCurriculum.query.filter_by(subject=old_name, class_name=old_class).update({WeeklyCurriculum.subject: new_name, WeeklyCurriculum.class_name: new_class}, synchronize_session=False)
                CurriculumBank.query.filter_by(subject=old_name, class_name=old_class).update({CurriculumBank.subject: new_name, CurriculumBank.class_name: new_class}, synchronize_session=False)
                for raport in Raport.query.all():
                    scores = raport.scores()
                    if old_name in scores:
                        scores[new_name] = scores.pop(old_name)
                        raport.scores_json = json.dumps(scores, ensure_ascii=False)
                row.name = new_name
                row.class_name = new_class
                row.notes = request.form.get("notes", "").strip()
                flash("Bidang pelajaran berhasil diperbarui.", "success")

            elif action == "toggle_subject":
                row = db.get_or_404(Subject, request.form.get("id", type=int))
                row.is_active = not row.is_active
                flash("Status bidang pelajaran berhasil diperbarui.", "success")

            elif action == "delete_subject":
                row = db.get_or_404(Subject, request.form.get("id", type=int))
                used = WeeklyCurriculum.query.filter_by(subject=row.name, class_name=row.class_name).count() + CurriculumBank.query.filter_by(subject=row.name, class_name=row.class_name).count()
                used_in_reports = any(row.name in raport.scores() for raport in Raport.query.all())
                if used or used_in_reports:
                    raise ValueError("Bidang pelajaran sudah digunakan pada silabus atau raport. Gunakan Nonaktifkan.")
                db.session.delete(row)
                flash("Bidang pelajaran berhasil dihapus permanen.", "success")

            elif action == "add_year":
                year_name = request.form.get("name", "")
                semester = request.form.get("semester", "Semester 1")
                if year_name not in academic_year_options:
                    raise ValueError("Tahun ajaran tidak valid.")
                if AcademicYear.query.filter_by(name=year_name, semester=semester).first():
                    raise ValueError("Tahun ajaran dan semester tersebut sudah tersedia.")
                start_date, end_date = year_dates(year_name, semester)
                make_primary = bool(request.form.get("is_primary"))
                if make_primary:
                    AcademicYear.query.update({AcademicYear.is_primary: False})
                db.session.add(AcademicYear(
                    name=year_name, semester=semester,
                    start_date=start_date, end_date=end_date,
                    is_active=True, is_primary=make_primary,
                ))
                flash("Tahun ajaran berhasil ditambahkan.", "success")

            elif action == "update_year":
                row = db.get_or_404(AcademicYear, request.form.get("id", type=int))
                old_name, old_semester = row.name, row.semester
                year_name = request.form.get("name", "")
                semester = request.form.get("semester", "Semester 1")
                if year_name not in academic_year_options:
                    raise ValueError("Tahun ajaran tidak valid.")
                duplicate = AcademicYear.query.filter(AcademicYear.name == year_name, AcademicYear.semester == semester, AcademicYear.id != row.id).first()
                if duplicate:
                    raise ValueError("Tahun ajaran dan semester tersebut sudah tersedia.")
                start_date, end_date = year_dates(year_name, semester)
                if year_name != old_name:
                    WeeklyCurriculum.query.filter_by(academic_year=old_name).update({WeeklyCurriculum.academic_year: year_name}, synchronize_session=False)
                    Raport.query.filter_by(academic_year=old_name, semester=old_semester).update({Raport.academic_year: year_name, Raport.semester: semester}, synchronize_session=False)
                elif semester != old_semester:
                    Raport.query.filter_by(academic_year=old_name, semester=old_semester).update({Raport.semester: semester}, synchronize_session=False)
                row.name = year_name
                row.semester = semester
                row.start_date = start_date
                row.end_date = end_date
                if request.form.get("is_primary"):
                    AcademicYear.query.update({AcademicYear.is_primary: False})
                    row.is_primary = True
                flash("Tahun ajaran berhasil diperbarui.", "success")

            elif action == "set_primary_year":
                row = db.get_or_404(AcademicYear, request.form.get("id", type=int))
                AcademicYear.query.update({AcademicYear.is_primary: False})
                row.is_primary = True
                row.is_active = True
                flash("Tahun ajaran utama berhasil ditetapkan.", "success")

            elif action == "delete_year":
                row = db.get_or_404(AcademicYear, request.form.get("id", type=int))
                if row.is_primary:
                    raise ValueError("Tahun ajaran utama tidak dapat dihapus. Tetapkan tahun ajaran lain sebagai utama terlebih dahulu.")
                used = WeeklyCurriculum.query.filter_by(academic_year=row.name).count() + Raport.query.filter_by(academic_year=row.name, semester=row.semester).count()
                if used:
                    raise ValueError("Tahun ajaran sudah digunakan pada silabus atau E-Raport dan tidak dapat dihapus.")
                db.session.delete(row)
                flash("Tahun ajaran berhasil dihapus permanen.", "success")

            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        redirect_args = {"tab": tab}
        if tab == "subjects":
            redirect_args["subject_class"] = (
                request.form.get("subject_class_filter", "").strip()
                or request.form.get("class_name", "").strip()
                or subject_class_filter
            )
        elif tab == "years":
            redirect_args["year_filter"] = (
                request.form.get("year_filter_value", "").strip()
                or request.form.get("name", "").strip()
                or year_filter
            )
        return redirect(url_for("data_master", **redirect_args))

    users = []
    if current_user.is_superadmin:
        users = (User.query.filter(User.role.in_(["admin_utama", "admin", "guru"]))
                 .order_by(User.role, func.lower(User.full_name)).all())
    subject_query = Subject.query
    if subject_class_filter:
        subject_query = subject_query.filter_by(class_name=subject_class_filter)

    year_query = AcademicYear.query
    if year_filter:
        year_query = year_query.filter_by(name=year_filter)

    return render_template(
        "data_master.html", tab=tab,
        classes=MasterClass.query.order_by(MasterClass.sort_order, MasterClass.name).all(),
        teachers=Teacher.query.order_by(Teacher.full_name).all(),
        subjects=subject_query.order_by(Subject.sort_order, Subject.name).all(),
        years=year_query.order_by(AcademicYear.semester).all(),
        users=users,
        staff_roles={"admin": "Admin", "guru": "Guru"},
        academic_year_options=academic_year_options,
        teacher_positions=teacher_positions,
        class_filter_options=class_filter_options,
        subject_class_filter=subject_class_filter,
        year_filter_options=year_filter_options,
        year_filter=year_filter,
    )


@app.route("/users", methods=["GET", "POST"])
@superadmin_required
def user_management():
    if request.method == "GET":
        return redirect(url_for("data_master", tab="users"))
    staff_roles = {"admin": "Admin", "guru": "Guru"}
    action = request.form.get("action", "").strip()
    try:
        if action == "profile":
            full_name = request.form.get("full_name", "").strip()
            username = request.form.get("username", "").strip()
            old_password = request.form.get("old_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            if not full_name or not username:
                raise ValueError("Nama lengkap dan username wajib diisi.")
            duplicate = User.query.filter(User.username == username, User.id != current_user.id).first()
            if duplicate:
                raise ValueError("Username sudah digunakan akun lain.")
            current_user.full_name = full_name
            current_user.username = username
            if new_password:
                if not current_user.check_password(old_password):
                    raise ValueError("Kata sandi lama tidak sesuai.")
                if len(new_password) < 8:
                    raise ValueError("Kata sandi baru minimal 8 karakter.")
                if new_password != confirm_password:
                    raise ValueError("Konfirmasi kata sandi baru tidak sama.")
                current_user.set_password(new_password)
            db.session.commit()
            flash("Data akun Admin Utama berhasil diperbarui.", "success")

        elif action == "create":
            full_name = request.form.get("full_name", "").strip()
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            confirm_password = request.form.get("confirm_password", "")
            role = request.form.get("role", "")
            assigned_class = normalize_class_name(request.form.get("assigned_class", ""))
            if not full_name or not username or role not in staff_roles:
                raise ValueError("Nama, username, dan peran wajib diisi.")
            if User.query.filter_by(username=username).first():
                raise ValueError("Username sudah digunakan akun lain.")
            if len(password) < 8:
                raise ValueError("Password awal minimal 8 karakter.")
            if password != confirm_password:
                raise ValueError("Konfirmasi password tidak sama.")
            if role == "guru" and assigned_class not in active_class_names(include_inactive=True):
                raise ValueError("Kelas yang diampu wajib dipilih untuk akun Guru.")
            user = User(full_name=full_name, username=username, role=role,
                        assigned_class=assigned_class if role == "guru" else "", is_active=True)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash(f"Akun {staff_roles[role]} berhasil ditambahkan.", "success")

        elif action == "update":
            user = db.get_or_404(User, request.form.get("user_id", type=int))
            if user.is_superadmin:
                raise ValueError("Akun Admin Utama hanya dapat diubah melalui bagian Akun Saya.")
            if user.role not in staff_roles:
                raise ValueError("Akun tersebut bukan akun staf yang dapat dikelola di halaman ini.")
            full_name = request.form.get("full_name", "").strip()
            username = request.form.get("username", "").strip()
            role = request.form.get("role", "")
            assigned_class = normalize_class_name(request.form.get("assigned_class", ""))
            new_password = request.form.get("new_password", "")
            if not full_name or not username or role not in staff_roles:
                raise ValueError("Nama, username, dan peran wajib diisi.")
            duplicate = User.query.filter(User.username == username, User.id != user.id).first()
            if duplicate:
                raise ValueError("Username sudah digunakan akun lain.")
            if role == "guru" and assigned_class not in active_class_names(include_inactive=True):
                raise ValueError("Kelas yang diampu wajib dipilih untuk akun Guru.")
            user.full_name = full_name
            user.username = username
            user.role = role
            user.assigned_class = assigned_class if role == "guru" else ""
            if new_password:
                if len(new_password) < 8:
                    raise ValueError("Password baru minimal 8 karakter.")
                user.set_password(new_password)
            db.session.commit()
            flash("Akun pengguna berhasil diperbarui.", "success")

        elif action == "toggle":
            user = db.get_or_404(User, request.form.get("user_id", type=int))
            if user.id == current_user.id or user.is_superadmin:
                raise ValueError("Akun Admin Utama yang sedang digunakan tidak dapat dinonaktifkan.")
            if user.role not in staff_roles:
                raise ValueError("Akun tersebut bukan akun staf.")
            user.is_active = not user.is_active
            db.session.commit()
            flash("Status akun berhasil diperbarui.", "success")
        else:
            raise ValueError("Perintah tidak dikenali.")
    except Exception as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("data_master", tab="users"))


@app.route("/hadith", methods=["GET", "POST"])
@admin_required
def hadith_manager():
    if request.method == "POST":
        action = request.form.get("action", "add")
        try:
            if action == "add":
                arabic = request.form.get("arabic", "").strip(); translation = request.form.get("translation", "").strip()
                if not arabic or not translation:
                    raise ValueError("Teks Arab dan terjemahan wajib diisi.")
                db.session.add(DailyHadith(arabic=arabic, translation=translation,
                                           source=request.form.get("source", "").strip(),
                                           number=request.form.get("number", "").strip(),
                                           theme=request.form.get("theme", "Adab").strip(),
                                           priority=request.form.get("priority", type=int) or 0,
                                           special_date=parse_date(request.form.get("special_date")), is_active=True))
                flash("Hadis berhasil ditambahkan.", "success")
            elif action == "toggle":
                row = db.get_or_404(DailyHadith, request.form.get("id", type=int)); row.is_active = not row.is_active
                flash("Status hadis berhasil diperbarui.", "success")
            elif action == "delete":
                row = db.get_or_404(DailyHadith, request.form.get("id", type=int)); db.session.delete(row)
                flash("Hadis berhasil dihapus.", "info")
            db.session.commit()
        except Exception as exc:
            db.session.rollback(); flash(str(exc), "danger")
        return redirect(url_for("hadith_manager"))
    theme = request.args.get("theme", "")
    query = DailyHadith.query
    if theme: query = query.filter_by(theme=theme)
    rows = query.order_by(DailyHadith.is_active.desc(), DailyHadith.priority.desc(), DailyHadith.id).all()
    themes = [r[0] for r in db.session.query(DailyHadith.theme).distinct().order_by(DailyHadith.theme).all()]
    return render_template("hadith.html", rows=rows, themes=themes, theme_filter=theme,
                           today_hadith=current_daily_hadith())


@app.route("/curriculum/bank/reload", methods=["POST"])
@admin_required
def curriculum_bank_reload():
    CurriculumBank.query.delete(); db.session.commit()
    seed_master_data(); db.session.commit()
    flash(f"Bank Silabus berhasil dimuat ulang: {CurriculumBank.query.count()} materi.", "success")
    return redirect(url_for("curriculum"))


@app.route("/poster/<int:poster_id>/published", methods=["POST"])
@admin_required
def mark_poster_published(poster_id):
    poster = db.get_or_404(PosterRecord, poster_id)
    poster.status = "Sudah Dipublikasikan"
    db.session.commit()
    flash("Poster ditandai sudah dipublikasikan.", "success")
    return redirect(url_for("dashboard") + "#poster-santri-terbaik")


@app.route("/poster/<path:filename>")
@admin_required
def poster_file(filename):
    return send_from_directory(os.path.join(UPLOAD_DIR, "posters"), filename, as_attachment=request.args.get("download") == "1")


@app.route("/ananda/<int:student_id>/raport")
@login_required
def guardian_report(student_id):
    student = db.get_or_404(Santri, student_id)
    if current_user.is_admin:
        return redirect(url_for("report_preview", student_id=student.id))
    selected = selected_guardian_student()
    if not selected or selected.id != student.id:
        abort(403)
    raport = get_or_create_raport(student)
    if raport.status != "Diterbitkan":
        flash("Raport belum diterbitkan oleh wali kelas.", "info")
        return redirect(url_for("guardian_student_detail", student_id=student.id))
    return send_file(build_report_pdf(student, raport), mimetype="application/pdf", as_attachment=False,
                     download_name=f"raport_{secure_filename(student.name)}.pdf")


@app.route("/finance", methods=["GET", "POST"])
@admin_required
def finance():
    if request.method == "POST":
        try:
            student_id = int(request.form.get("student_id", "0"))
            month = request.form.get("month", "")
            year = int(request.form.get("year", "2026"))
            nominal = int(request.form.get("nominal", "50000"))
            status = request.form.get("status", "Belum Lunas")
            if month not in MONTHS or year < 2026 or year > 2040 or nominal < 0:
                raise ValueError("Data tagihan tidak valid")
            db.get_or_404(Santri, student_id)
            bill = Iuran(santri_id=student_id, month=month, year=year,
                         nominal=nominal, status=status)
            db.session.add(bill)
            db.session.commit()
            flash("Tagihan berhasil dibuat.", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Tagihan gagal dibuat: {exc}", "danger")
        return redirect(url_for("finance", q=request.form.get("student_name", ""), month=request.form.get("month", "")))

    q = request.args.get("q", "").strip()
    class_filter = request.args.get("class_name", "").strip()
    month_filter = request.args.get("month", "").strip()
    query = Iuran.query.join(Santri)
    if q:
        query = query.filter(Santri.name.ilike(f"%{q}%"))
    if class_filter in CLASSES:
        query = query.filter(Santri.class_name == class_filter)
    if month_filter in MONTHS:
        query = query.filter(Iuran.month == month_filter)
    total_matches = query.count()
    bills = query.order_by(Iuran.year.desc(), Iuran.id.desc()).limit(60).all()
    student_rows = Santri.query.order_by(Santri.name).all()
    return render_template(
        "finance.html", bills=bills, students=student_rows, years=range(2026, 2041),
        q=q, class_filter=class_filter, month_filter=month_filter,
        total_matches=total_matches
    )


@app.route("/finance/<int:bill_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_bill(bill_id):
    bill = db.get_or_404(Iuran, bill_id)
    if request.method == "POST":
        try:
            student_id = int(request.form.get("student_id", "0"))
            month = request.form.get("month", "").strip()
            year = int(request.form.get("year", "2026"))
            status = request.form.get("status", "Belum Lunas").strip()
            nominal = int(request.form.get("nominal", "50000"))
            if month not in MONTHS or not 2026 <= year <= 2040 or nominal < 0:
                raise ValueError("Data administrasi tidak valid")
            db.get_or_404(Santri, student_id)
            bill.santri_id = student_id
            bill.month = month
            bill.year = year
            bill.status = status
            bill.nominal = nominal
            if status == "Lunas" and not bill.verified_at:
                bill.verified_at = datetime.utcnow()
            elif status != "Lunas":
                bill.verified_at = None
            db.session.commit()
            flash("Data administrasi santri berhasil diperbarui.", "success")
            return redirect(url_for("finance", q=bill.santri.name, class_name=bill.santri.class_name, month=bill.month))
        except Exception as exc:
            db.session.rollback()
            flash(f"Perubahan gagal disimpan: {exc}", "danger")
    students_rows = Santri.query.order_by(Santri.name).all()
    return render_template("finance_edit.html", bill=bill, students=students_rows, years=range(2026, 2041))


@app.route("/bill/<int:bill_id>/upload", methods=["POST"])
@login_required
def upload_bill_proof(bill_id):
    bill = db.get_or_404(Iuran, bill_id)
    selected = selected_guardian_student()
    if current_user.is_admin or not selected or bill.santri_id != selected.id:
        abort(403)
    if bill.status == "Lunas":
        flash("Tagihan ini sudah lunas.", "info")
        return redirect(url_for("guardian_student_detail", student_id=bill.santri_id))
    try:
        filename = save_upload(request.files.get("proof"), PROOF_DIR, {"jpg", "jpeg", "png", "pdf"})
        if not filename:
            raise ValueError("Pilih file bukti transfer")
        if bill.proof_path:
            old = os.path.join(PROOF_DIR, bill.proof_path)
            if os.path.exists(old):
                os.remove(old)
        bill.proof_path = filename
        bill.status = "Menunggu Verifikasi"
        db.session.commit()
        flash("Bukti transfer terkirim dan menunggu verifikasi admin.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("guardian_student_detail", student_id=bill.santri_id))


@app.route("/bill/<int:bill_id>/verify", methods=["POST"])
@admin_required
def verify_bill(bill_id):
    bill = db.get_or_404(Iuran, bill_id)
    bill.status = "Lunas"
    bill.verified_at = datetime.utcnow()
    db.session.commit()
    flash("Pembayaran telah diverifikasi dan ditandai Lunas.", "success")
    return redirect(url_for("finance"))


@app.route("/bill/<int:bill_id>/receipt")
@login_required
def bill_receipt(bill_id):
    bill = db.get_or_404(Iuran, bill_id)
    if not current_user.is_admin:
        selected = selected_guardian_student()
        if not selected or bill.santri_id != selected.id:
            abort(403)
    return render_template("receipt.html", bill=bill)


@app.route("/bill/<int:bill_id>/whatsapp")
@login_required
def bill_whatsapp(bill_id):
    bill = db.get_or_404(Iuran, bill_id)
    if not current_user.is_admin:
        selected = selected_guardian_student()
        if not selected or bill.santri_id != selected.id:
            abort(403)
    text = (f"Assalamu'alaikum, berikut tagihan iuran TPQ ananda {bill.santri.name} "
            f"bulan {bill.month}. Nominal: {rupiah(bill.nominal)}. Terima kasih.")
    return redirect(f"https://wa.me/?text={quote(text)}")


@app.route("/proof/bill/<int:bill_id>")
@admin_required
def view_bill_proof(bill_id):
    bill = db.get_or_404(Iuran, bill_id)
    if not bill.proof_path:
        abort(404)
    return send_from_directory(PROOF_DIR, bill.proof_path)


@app.route("/library")
@login_required
def library():
    selected_category = request.args.get("category", "").strip()
    query = Kitab.query
    if selected_category in BOOK_CATEGORIES:
        query = query.filter_by(category=selected_category)
    books = query.order_by(Kitab.uploaded_at.desc()).all()
    selected_student = None
    access_map = {}
    if not current_user.is_admin:
        selected_student = selected_guardian_student()
        if not selected_student:
            flash("Silakan pilih data santri terlebih dahulu.", "info")
            return redirect(url_for("home"))
        access_map = {a.kitab_id: a for a in selected_student.book_accesses}
    pending = AksesKitab.query.filter_by(status="Menunggu Verifikasi").all() if current_user.is_admin else []
    return render_template(
        "library.html", books=books, student=selected_student, access_map=access_map,
        pending=pending, selected_category=selected_category
    )


@app.route("/library/upload", methods=["GET", "POST"])
@admin_required
def upload_book():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        category = request.form.get("category", "").strip()
        try:
            price = max(0, int(request.form.get("price", "0")))
            if category not in BOOK_CATEGORIES:
                raise ValueError("Pilih kategori kitab yang tersedia")
            pdf = request.files.get("pdf")
            if not title or not pdf or not pdf.filename:
                raise ValueError("Judul dan file PDF wajib diisi")
            original_filename = secure_filename(pdf.filename)
            if not original_filename.lower().endswith(".pdf"):
                raise ValueError("Buku harus berupa PDF")
            temp_name = f"tmp_{uuid.uuid4().hex}.pdf"
            temp_path = os.path.join(BOOK_DIR, temp_name)
            pdf.save(temp_path)
            original_size = os.path.getsize(temp_path)
            final_name = f"{uuid.uuid4().hex}.pdf"
            final_path = os.path.join(BOOK_DIR, final_name)
            try:
                import fitz
                doc = fitz.open(temp_path)
                doc.save(final_path, garbage=4, deflate=True, clean=True)
                doc.close()
                os.remove(temp_path)
            except Exception:
                shutil.move(temp_path, final_path)
            optimized_size = os.path.getsize(final_path)
            book = Kitab(title=title, description=description, category=category, price=price,
                         filename=final_name, original_filename=original_filename,
                         original_size=original_size, optimized_size=optimized_size)
            db.session.add(book)
            db.session.commit()
            flash("Buku berhasil diunggah dan dioptimalkan.", "success")
            return redirect(url_for("library"))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template("library_upload.html")


@app.route("/library/<int:book_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_book(book_id):
    book = db.get_or_404(Kitab, book_id)
    if request.method == "POST":
        try:
            title = request.form.get("title", "").strip()
            description = request.form.get("description", "").strip()
            category = request.form.get("category", "").strip()
            price = max(0, int(request.form.get("price", "0")))
            if not title:
                raise ValueError("Judul buku wajib diisi")
            if category not in BOOK_CATEGORIES:
                raise ValueError("Pilih kategori kitab yang tersedia")

            replacement_pdf = request.files.get("pdf")
            if replacement_pdf and replacement_pdf.filename:
                original_filename = secure_filename(replacement_pdf.filename)
                if not original_filename.lower().endswith(".pdf"):
                    raise ValueError("Buku harus berupa PDF")
                temp_name = f"tmp_{uuid.uuid4().hex}.pdf"
                temp_path = os.path.join(BOOK_DIR, temp_name)
                replacement_pdf.save(temp_path)
                original_size = os.path.getsize(temp_path)
                final_name = f"{uuid.uuid4().hex}.pdf"
                final_path = os.path.join(BOOK_DIR, final_name)
                try:
                    import fitz
                    doc = fitz.open(temp_path)
                    doc.save(final_path, garbage=4, deflate=True, clean=True)
                    doc.close()
                    os.remove(temp_path)
                except Exception:
                    shutil.move(temp_path, final_path)

                old_filename = book.filename
                old_path = os.path.join(BOOK_DIR, old_filename or "")
                if old_filename:
                    clear_book_preview_cache(old_filename)
                if old_filename and os.path.exists(old_path):
                    os.remove(old_path)
                book.filename = final_name
                book.original_filename = original_filename
                book.original_size = original_size
                book.optimized_size = os.path.getsize(final_path)

            book.title = title
            book.description = description
            book.category = category
            book.price = price
            db.session.commit()
            flash("Data buku berhasil diperbarui.", "success")
            return redirect(url_for("library"))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template("library_edit.html", book=book)


@app.route("/library/<int:book_id>/delete", methods=["POST"])
@admin_required
def delete_book(book_id):
    book = db.get_or_404(Kitab, book_id)
    title = book.title
    filename = book.filename
    path = os.path.join(BOOK_DIR, filename or "")
    proof_paths = [access.proof_path for access in book.accesses if access.proof_path]
    db.session.delete(book)
    db.session.commit()
    if filename:
        clear_book_preview_cache(filename)
    if filename and os.path.exists(path):
        os.remove(path)
    for proof_name in proof_paths:
        proof_path = os.path.join(PROOF_DIR, proof_name)
        if os.path.exists(proof_path):
            os.remove(proof_path)
    flash(f"Buku {title} berhasil dihapus.", "info")
    return redirect(url_for("library"))


@app.route("/library/access/<int:access_id>/revoke", methods=["POST"])
@admin_required
def revoke_book_access(access_id):
    access = db.get_or_404(AksesKitab, access_id)
    student_name = access.santri.name
    book_title = access.kitab.title
    access.status = "Terkunci"
    access.confirmed_at = None
    db.session.commit()
    flash(f"Akses {book_title} untuk {student_name} telah ditutup.", "info")
    return redirect(url_for("library"))


@app.route("/library/<int:book_id>/purchase")
@login_required
def purchase_book(book_id):
    if current_user.is_admin:
        return redirect(url_for("download_book", book_id=book_id))
    book = db.get_or_404(Kitab, book_id)
    student = selected_guardian_student()
    if not student:
        abort(403)
    access = AksesKitab.query.filter_by(kitab_id=book.id, santri_id=student.id).first()
    if has_book_access(book, student):
        return redirect(url_for("download_book", book_id=book.id))
    return render_template("library_purchase.html", book=book, student=student, access=access)


@app.route("/library/<int:book_id>/unlock", methods=["POST"])
@login_required
def request_book_unlock(book_id):
    if current_user.is_admin:
        abort(403)
    book = db.get_or_404(Kitab, book_id)
    if book.price == 0:
        flash("Buku ini gratis dan dapat langsung dibuka tanpa pembayaran.", "info")
        return redirect(url_for("library"))
    student_id = request.form.get("student_id", type=int)
    student = db.get_or_404(Santri, student_id)
    selected = selected_guardian_student()
    if not selected or student.id != selected.id:
        abort(403)
    try:
        filename = save_upload(request.files.get("proof"), PROOF_DIR, {"jpg", "jpeg", "png", "pdf"})
        if not filename:
            raise ValueError("Unggah bukti transfer terlebih dahulu")
        access = AksesKitab.query.filter_by(kitab_id=book.id, santri_id=student.id).first()
        if not access:
            access = AksesKitab(kitab_id=book.id, santri_id=student.id)
            db.session.add(access)
        if access.proof_path:
            old = os.path.join(PROOF_DIR, access.proof_path)
            if os.path.exists(old):
                os.remove(old)
        access.proof_path = filename
        access.status = "Menunggu Verifikasi"
        access.requested_at = datetime.utcnow()
        db.session.commit()
        flash("Permintaan unlock terkirim. Admin akan memverifikasi pembayaran.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("purchase_book", book_id=book.id))


@app.route("/library/access/<int:access_id>/confirm", methods=["POST"])
@admin_required
def confirm_book_access(access_id):
    access = db.get_or_404(AksesKitab, access_id)
    access.status = "Terbuka"
    access.confirmed_at = datetime.utcnow()
    db.session.commit()
    flash(f"Buku {access.kitab.title} dibuka untuk {access.santri.name}.", "success")
    return redirect(url_for("library"))


@app.route("/proof/book/<int:access_id>")
@admin_required
def view_book_proof(access_id):
    access = db.get_or_404(AksesKitab, access_id)
    if not access.proof_path:
        abort(404)
    return send_from_directory(PROOF_DIR, access.proof_path)


def has_book_access(book, student):
    # Admin selalu memiliki akses penuh. Buku harga Rp 0 otomatis gratis bagi wali.
    if current_user.is_admin or book.price == 0:
        return True
    access = AksesKitab.query.filter_by(kitab_id=book.id, santri_id=student.id).first()
    return bool(access and access.status == "Terbuka")


def preview_cache_directory(filename):
    safe_stem = os.path.splitext(os.path.basename(filename or "book"))[0]
    return os.path.join(PREVIEW_DIR, safe_stem)


def clear_book_preview_cache(filename):
    cache_dir = preview_cache_directory(filename)
    if os.path.isdir(cache_dir):
        shutil.rmtree(cache_dir, ignore_errors=True)


def build_book_preview(pdf_path, filename, page_number):
    """Render one real PDF page to a cached JPEG preview."""
    import fitz

    cache_dir = preview_cache_directory(filename)
    os.makedirs(cache_dir, exist_ok=True)
    output_path = os.path.join(cache_dir, f"page_{page_number}.jpg")
    if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
        return output_path

    doc = fitz.open(pdf_path)
    try:
        if doc.needs_pass:
            raise ValueError("PDF terkunci dengan kata sandi")
        if page_number < 1 or page_number > doc.page_count:
            return None
        page = doc.load_page(page_number - 1)
        pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5), alpha=False)
        image_bytes = pix.tobytes("jpeg", jpg_quality=84)
        temporary_path = f"{output_path}.tmp"
        with open(temporary_path, "wb") as preview_file:
            preview_file.write(image_bytes)
        os.replace(temporary_path, output_path)
        return output_path
    finally:
        doc.close()


@app.route("/library/<int:book_id>/preview/<int:page_number>")
@login_required
def book_preview(book_id, page_number):
    book = db.get_or_404(Kitab, book_id)
    if page_number < 1:
        abort(404)
    if not current_user.is_admin:
        student = selected_guardian_student()
        if not student:
            abort(403)
        unlocked = has_book_access(book, student)
        if not unlocked and page_number > PREVIEW_PAGE_LIMIT:
            abort(403)
    if not book.filename:
        abort(404)

    pdf_path = os.path.join(BOOK_DIR, book.filename)
    if not os.path.exists(pdf_path):
        abort(404)

    try:
        preview_path = build_book_preview(pdf_path, book.filename, page_number)
        if not preview_path:
            abort(404)
        return send_file(preview_path, mimetype="image/jpeg", conditional=True, max_age=86400)
    except (ValueError, RuntimeError):
        abort(422)
    except Exception:
        app.logger.exception("Gagal membuat pratinjau PDF untuk buku %s halaman %s", book.id, page_number)
        abort(500)


@app.route("/library/<int:book_id>/download")
@login_required
def download_book(book_id):
    book = db.get_or_404(Kitab, book_id)
    if not current_user.is_admin:
        student = selected_guardian_student()
        if not student:
            abort(403)
        if not has_book_access(book, student):
            flash("Buku penuh dapat diunduh setelah pembelian dan verifikasi admin.", "info")
            return redirect(url_for("purchase_book", book_id=book.id))
    path = os.path.join(BOOK_DIR, book.filename or "")
    if not book.filename or not os.path.exists(path):
        abort(404)
    return send_file(path, as_attachment=download_requested(), download_name=book.original_filename or f"{book.title}.pdf")


@app.route("/backup/academic.csv")
@admin_required
def backup_academic():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["NIS", "Nama", "Kelas", "Wali", "Semester", "Tahun Ajaran", "Nilai JSON", "Hafalan JSON", "Mutabaah JSON"])
    for student in Santri.query.order_by(Santri.nis).all():
        raport = get_or_create_raport(student)
        writer.writerow([student.nis, student.name, student.class_name, student.guardian.username,
                         raport.semester, raport.academic_year, raport.scores_json,
                         raport.hafalan_json, raport.mutabaah_json])
    db.session.commit()
    data = output.getvalue().encode("utf-8-sig")
    return Response(data, mimetype="text/csv", headers={
        "Content-Disposition": f"attachment; filename=backup_akademik_{date.today().isoformat()}.csv"
    })


@app.route("/backup/financial.csv")
@admin_required
def backup_financial():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["NIS", "Nama", "Bulan", "Tahun", "Status", "Nominal", "Tanggal Verifikasi"])
    for bill in Iuran.query.order_by(Iuran.year, Iuran.id).all():
        writer.writerow([bill.santri.nis, bill.santri.name, bill.month, bill.year,
                         bill.status, bill.nominal, bill.verified_at or ""])
    data = output.getvalue().encode("utf-8-sig")
    return Response(data, mimetype="text/csv", headers={
        "Content-Disposition": f"attachment; filename=backup_keuangan_{date.today().isoformat()}.csv"
    })


@app.route("/backup/database.sql")
@admin_required
def backup_database():
    db_path = os.path.join(app.instance_path, "tpq_hmarisa.db")
    if not os.path.exists(db_path):
        abort(404)
    connection = sqlite3.connect(db_path)
    dump = "\n".join(connection.iterdump())
    connection.close()
    return Response(dump, mimetype="application/sql", headers={
        "Content-Disposition": f"attachment; filename=tpq_hmarisa_{date.today().isoformat()}.sql"
    })


@app.errorhandler(403)
def forbidden(_error):
    return render_template("error.html", code=403, message="Anda tidak memiliki akses ke halaman ini."), 403


@app.errorhandler(404)
def not_found(_error):
    return render_template("error.html", code=404, message="Halaman atau data tidak ditemukan."), 404


@app.errorhandler(413)
def too_large(_error):
    flash("Ukuran file terlalu besar. Maksimal 25 MB.", "danger")
    return redirect(request.referrer or url_for("dashboard"))


def ensure_schema_updates():
    """Safely upgrade older SQLite databases without deleting existing data."""
    inspector = inspect(db.engine)
    table_names = set(inspector.get_table_names())

    additions = {
        "user": {
            "assigned_class": "VARCHAR(60) DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT 1 NOT NULL",
            "last_login_at": "DATETIME",
            "updated_at": "DATETIME",
        },
        "santri": {
            "nickname": "VARCHAR(80) DEFAULT ''",
            "public_name": "VARCHAR(100) DEFAULT ''",
            "guardian_name": "VARCHAR(140) DEFAULT ''",
            "guardian_phone": "VARCHAR(40) DEFAULT ''",
            "joined_date": "DATE",
            "is_active": "BOOLEAN DEFAULT 1 NOT NULL",
        },
        "raport": {
            "attitude_json": "TEXT DEFAULT '{}' NOT NULL",
            "absence_json": "TEXT DEFAULT '{}' NOT NULL",
            "development_notes": "TEXT DEFAULT ''",
            "status": "VARCHAR(30) DEFAULT 'Draf' NOT NULL",
            "completeness": "INTEGER DEFAULT 0",
            "publish_date": "DATE",
            "published_at": "DATETIME",
            "version": "INTEGER DEFAULT 1",
            "snapshot_json": "TEXT DEFAULT '{}' NOT NULL",
        },
        "weekly_curriculum": {
            "month": "VARCHAR(20) DEFAULT 'Juli' NOT NULL",
            "year": "INTEGER DEFAULT 2026 NOT NULL",
            "academic_year": "VARCHAR(20) DEFAULT '2026/2027' NOT NULL",
            "week1": "TEXT DEFAULT ''",
            "week2": "TEXT DEFAULT ''",
            "week3": "TEXT DEFAULT ''",
            "week4": "TEXT DEFAULT ''",
            "week5": "TEXT DEFAULT ''",
            "source_type": "VARCHAR(40) DEFAULT 'Data Silabus Lama'",
            "status": "VARCHAR(20) DEFAULT 'Aktif'",
            "version": "INTEGER DEFAULT 1",
        },
        "kitab": {"category": "VARCHAR(80) DEFAULT 'Kitab Fiqih'"},
    }
    with db.engine.begin() as connection:
        for table, columns_to_add in additions.items():
            if table not in table_names:
                continue
            current = {column["name"] for column in inspect(db.engine).get_columns(table)}
            for column_name, sql_type in columns_to_add.items():
                if column_name not in current:
                    connection.execute(text(f"ALTER TABLE {table} ADD COLUMN {column_name} {sql_type}"))
        if "user" in table_names:
            connection.execute(text("UPDATE user SET is_active=1 WHERE is_active IS NULL"))
            connection.execute(text("UPDATE user SET role='admin_utama' WHERE username='tpqhmarisa' AND role='admin'"))
        if "kitab" in table_names:
            connection.execute(text("UPDATE kitab SET category='Kitab Fiqih' WHERE category IS NULL OR category=''"))
        if "santri" in table_names:
            connection.execute(text("UPDATE santri SET class_name='Ar Rahman' WHERE class_name='Ar-Rahman'"))
            connection.execute(text("UPDATE santri SET class_name='Ar Rahim' WHERE class_name='Ar-Rahim'"))
            connection.execute(text("UPDATE santri SET class_name='Al-Bayyan' WHERE class_name IN ('Al Bayyan','Al-Bayyan')"))
            connection.execute(text("UPDATE santri SET is_active=1 WHERE is_active IS NULL"))
            connection.execute(text("UPDATE santri SET public_name=name WHERE public_name IS NULL OR TRIM(public_name)=''"))
        if "weekly_curriculum" in table_names:
            connection.execute(text("UPDATE weekly_curriculum SET class_name='Ar Rahman' WHERE class_name='Ar-Rahman'"))
            connection.execute(text("UPDATE weekly_curriculum SET class_name='Ar Rahim' WHERE class_name='Ar-Rahim'"))
            connection.execute(text("UPDATE weekly_curriculum SET source_type='Data Silabus Lama' WHERE source_type IS NULL OR source_type=''"))
            connection.execute(text("UPDATE weekly_curriculum SET week1=activities WHERE (week1 IS NULL OR week1='') AND activities IS NOT NULL"))


def create_sample_book():
    path = os.path.join(BOOK_DIR, "sample_panduan_wudhu.pdf")
    if os.path.exists(path):
        return path
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        c = canvas.Canvas(path, pagesize=A4)
        width, height = A4
        pages = [
            ("Panduan Wudhu Anak", "Belajar wudhu dengan tertib, bersih, dan menyenangkan."),
            ("Langkah 1", "Niat, membaca basmalah, lalu membasuh kedua telapak tangan."),
            ("Langkah 2", "Berkumur, membersihkan hidung, kemudian membasuh wajah."),
            ("Langkah 3", "Membasuh tangan, mengusap kepala dan telinga, lalu membasuh kaki."),
        ]
        for title, body in pages:
            c.setFillColorRGB(0.01, 0.29, 0.21)
            c.rect(0, height - 120, width, 120, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            c.setFont("Helvetica-Bold", 24)
            c.drawCentredString(width / 2, height - 72, title)
            c.setFillColorRGB(0.1, 0.1, 0.1)
            c.setFont("Helvetica", 15)
            text = c.beginText(70, height - 190)
            text.setLeading(24)
            for line in body.split(". "):
                text.textLine(line.strip())
            c.drawText(text)
            c.setFillColorRGB(0.9, 0.66, 0.23)
            c.circle(width / 2, height / 2 - 40, 70, fill=1, stroke=0)
            c.showPage()
        c.save()
    except Exception:
        return None
    return path



def seed_master_data():
    """Seed dynamic master data, curated hadith, and the semester curriculum bank."""
    class_codes = {"Ar Rahman": "ARR", "Ar Rahim": "ARH", "Al-Bayyan": "ALB"}
    for order, class_name in enumerate(CLASSES, 1):
        row = MasterClass.query.filter_by(name=class_name).first()
        if not row:
            row = MasterClass(name=class_name, code=class_codes[class_name],
                              teacher_name=TEACHERS[class_name], sort_order=order,
                              academic_year_start="2026/2027", is_active=True)
            db.session.add(row)
        else:
            row.teacher_name = TEACHERS[class_name]

    teacher_rows = [
        ("Yeni Susilawati", "", "Ar Rahman", "Guru/Wali Kelas"),
        ("Faisal Kazim Muyassar", "S.H", "Ar Rahim", "Guru/Wali Kelas"),
        ("Hj. Maryamah", "S.Ag", "Al-Bayyan", "Guru/Wali Kelas dan Kepala TPQ"),
    ]
    for full_name, title, class_name, position in teacher_rows:
        if not Teacher.query.filter_by(full_name=full_name, class_name=class_name).first():
            db.session.add(Teacher(full_name=full_name, title=title, class_name=class_name,
                                   position=position, is_active=True))

    for class_name, subjects in SCORE_FIELDS.items():
        for order, subject_name in enumerate(subjects, 1):
            if not Subject.query.filter_by(name=subject_name, class_name=class_name).first():
                code = re.sub(r"[^A-Z0-9]", "", "".join(word[:1].upper() for word in subject_name.split()))[:12] or f"BDG{order}"
                db.session.add(Subject(name=subject_name, code=f"{class_codes[class_name]}-{code}",
                                       class_name=class_name, sort_order=order, is_active=True))

    if not AcademicYear.query.filter_by(name="2026/2027", semester="Semester 1").first():
        db.session.add(AcademicYear(name="2026/2027", semester="Semester 1",
                                    start_date=date(2026,7,13), end_date=date(2026,12,21),
                                    is_active=True, is_primary=True))
    if not AcademicYear.query.filter_by(name="2026/2027", semester="Semester 2").first():
        db.session.add(AcademicYear(name="2026/2027", semester="Semester 2",
                                    is_active=True, is_primary=False))

    hadith_rows = [
        ("خَيْرُكُمْ مَنْ تَعَلَّمَ الْقُرْآنَ وَعَلَّمَهُ", "Sebaik-baik kalian adalah orang yang mempelajari Al-Qur'an dan mengajarkannya.", "HR. Al-Bukhari", "5027", "Al-Qur'an"),
        ("الطُّهُورُ شَطْرُ الإِيمَانِ", "Bersuci adalah separuh dari iman.", "HR. Muslim", "223", "Kebersihan"),
        ("إِنَّمَا بُعِثْتُ لِأُتَمِّمَ صَالِحَ الأَخْلَاقِ", "Sesungguhnya aku diutus untuk menyempurnakan akhlak yang mulia.", "HR. Ahmad", "8952", "Adab"),
        ("مَنْ لَا يَرْحَمْ لَا يُرْحَمْ", "Siapa yang tidak menyayangi, tidak akan disayangi.", "Muttafaq 'alaih", "", "Kasih Sayang"),
        ("الْمُسْلِمُ أَخُو الْمُسْلِمِ", "Seorang Muslim adalah saudara bagi Muslim lainnya.", "Muttafaq 'alaih", "", "Persaudaraan"),
    ]
    if DailyHadith.query.count() == 0:
        for priority, row in enumerate(hadith_rows, 1):
            db.session.add(DailyHadith(arabic=row[0], translation=row[1], source=row[2], number=row[3],
                                       theme=row[4], priority=priority, is_active=True))

    bank_path = os.path.join(BASE_DIR, "data", "curriculum_bank.json")
    if CurriculumBank.query.count() == 0 and os.path.exists(bank_path):
        try:
            with open(bank_path, "r", encoding="utf-8") as file:
                records = json.load(file)
            for item in records:
                db.session.add(CurriculumBank(
                    material_code=str(item.get("ID Materi") or uuid.uuid4().hex),
                    class_name=normalize_class_name(item.get("Kelas")),
                    teacher_name=item.get("Guru Kelas") or "",
                    meeting_number=int(item.get("Pertemuan") or 1),
                    meeting_date=parse_date(item.get("Tanggal Pertemuan")),
                    month=item.get("Bulan") or "Juli",
                    year=int(item.get("Tahun") or 2026),
                    week_in_month=int(item.get("Pekan dalam Bulan") or 1),
                    day_name=item.get("Hari") or "",
                    subject=item.get("Bidang Pelajaran") or "",
                    topic=item.get("Materi Pokok") or "",
                    learning_target=item.get("Target Pembelajaran (Draft)") or "",
                    notes=item.get("Catatan Guru") or "",
                    calendar_agenda=item.get("Agenda Kaldik") or "",
                    is_active=(item.get("Status") or "Aktif") == "Aktif",
                    source_file=item.get("Sumber File") or "",
                ))
        except Exception as exc:
            app.logger.exception("Gagal memuat bank silabus: %s", exc)

def seed_database():
    seed_master_data()
    admin = User.query.filter_by(username="tpqhmarisa").first()
    if not admin:
        admin = User(username="tpqhmarisa", full_name="Administrator TPQ HMarisa", role="admin_utama", is_active=True)
        admin.set_password("tpqhmarisa")
        db.session.add(admin)

    guardian = User.query.filter_by(username="walizaki").first()
    if not guardian:
        guardian = User(username="walizaki", full_name="Wali Ahmad Zaki", role="guardian")
        guardian.set_password("walizaki123")
        db.session.add(guardian)
        db.session.flush()

    student = Santri.query.filter_by(name="Ahmad Zaki Al-Fatih").first()
    if not student:
        student = Santri(nis="26001", name="Ahmad Zaki Al-Fatih", class_name="Ar Rahim",
                         guardian_id=guardian.id)
        db.session.add(student)
        db.session.flush()
        scores = {"BTQ": 88, "Hafalan Surat Juz 30 dan Doa Harian": 90, "Materi dan Praktik Wudhu dan Shalat": 86, "Fiqih Dasar": 92}
        mutabaah = [
            {"date": "2026-07-10", "tilawati": "Tilawati Jilid 2 halaman 14", "fiqih": "Rukun wudhu", "notes": "Makhraj semakin baik."},
            {"date": "2026-07-08", "tilawati": "Tilawati Jilid 2 halaman 12", "fiqih": "Adab ke masjid", "notes": "Perlu murojaah harakat kasrah."},
        ]
        hafalan = {surah: i >= len(SURAH_JUZ30) - 5 for i, surah in enumerate(SURAH_JUZ30)}
        db.session.add(Raport(santri_id=student.id, scores_json=json.dumps(scores, ensure_ascii=False),
                              mutabaah_json=json.dumps(mutabaah, ensure_ascii=False),
                              hafalan_json=json.dumps(hafalan, ensure_ascii=False)))
        db.session.add_all([
            Iuran(santri_id=student.id, month="Juli", year=2026, status="Lunas", nominal=50000, verified_at=datetime.utcnow()),
            Iuran(santri_id=student.id, month="Agustus", year=2026, status="Belum Lunas", nominal=50000),
        ])

    if Kitab.query.count() == 0:
        sample_path = create_sample_book()
        if sample_path:
            size = os.path.getsize(sample_path)
            book = Kitab(title="Panduan Wudhu Anak", description="Buku digital ringkas untuk mengenalkan urutan wudhu kepada santri.",
                         category="Kitab Fiqih", price=25000, filename=os.path.basename(sample_path),
                         original_filename="Panduan Wudhu Anak.pdf", original_size=size, optimized_size=size)
            db.session.add(book)

    db.session.commit()


with app.app_context():
    db.create_all()
    ensure_schema_updates()
    seed_database()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=os.environ.get("FLASK_DEBUG") == "1")
