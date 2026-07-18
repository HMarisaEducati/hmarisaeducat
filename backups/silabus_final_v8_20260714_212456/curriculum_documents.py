"""Dokumen resmi Silabus Bulanan TPQ HMarisa.

Modul ini sengaja dipisahkan dari app.py agar route tetap ringkas dan seluruh
format PDF/Excel memakai standar visual yang sama.
"""
from __future__ import annotations

import io
import os
import re
from collections import OrderedDict
from typing import Iterable, Mapping, Sequence

MONTH_ORDER = [
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
]
GREEN = "075F46"
GOLD = "C49A31"
LIGHT_GREEN = "EAF4F0"
LIGHT_GOLD = "FFF8E8"
BORDER = "AAB8B2"


def display_value(value) -> str:
    text = str(value or "").strip()
    return text if text and text != "-" else "Belum diisi"


def safe_filename_part(value) -> str:
    text = str(value or "dokumen").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or "dokumen"


def combined_week_plan(row) -> str:
    parts = []
    for number in range(1, 6):
        value = str(getattr(row, f"week{number}", "") or "").strip()
        if value:
            parts.append(f"Pekan {number}: {value}")
    return "\n".join(parts) if parts else "Belum diisi"


def _pdf_styles():
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "CurriculumTitle",
            parent=base["Title"],
            fontName="Helvetica-Bold",
            fontSize=15,
            leading=18,
            alignment=TA_CENTER,
            textColor="#143C30",
            spaceAfter=0,
        ),
        "subtitle": ParagraphStyle(
            "CurriculumSubtitle",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            alignment=TA_CENTER,
            textColor="#4B5D55",
        ),
        "cell": ParagraphStyle(
            "CurriculumCell",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=7.8,
            leading=9.8,
            alignment=TA_LEFT,
            wordWrap="CJK",
            splitLongWords=True,
            spaceAfter=0,
        ),
        "cell_bold": ParagraphStyle(
            "CurriculumCellBold",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=7.8,
            leading=9.8,
            alignment=TA_LEFT,
            wordWrap="CJK",
            splitLongWords=True,
            spaceAfter=0,
        ),
        "small": ParagraphStyle(
            "CurriculumSmall",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=7.2,
            leading=9,
            textColor="#51645B",
        ),
        "signature": ParagraphStyle(
            "CurriculumSignature",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=12,
            alignment=TA_CENTER,
        ),
    }


def _paragraph(value, style):
    from reportlab.platypus import Paragraph
    from xml.sax.saxutils import escape

    text = escape(display_value(value)).replace("\n", "<br/>")
    return Paragraph(text, style)


def _letterhead_flowable(letterhead_path: str, available_width: float):
    from reportlab.platypus import Image

    if not letterhead_path or not os.path.exists(letterhead_path):
        return None
    image = Image(letterhead_path)
    ratio = image.imageHeight / float(image.imageWidth)
    image.drawWidth = available_width
    image.drawHeight = available_width * ratio
    return image


def _pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColorRGB(0.35, 0.40, 0.38)
    canvas.drawString(doc.leftMargin, 18, "Portal TPQ HMarisa - Dokumen Silabus")
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 18, f"Halaman {doc.page}")
    canvas.restoreState()


def _signature_table(principal: str, second_role: str, second_name: str, date_text: str, width: float):
    from reportlab.platypus import Paragraph, Table, TableStyle
    from reportlab.lib import colors

    styles = _pdf_styles()
    left = Paragraph(
        f"Mengetahui,<br/>Kepala TPQ HMarisa<br/><br/><br/><b>{principal}</b>",
        styles["signature"],
    )
    right = Paragraph(
        f"Tangerang Selatan, {date_text}<br/>{second_role}<br/><br/><br/><b>{second_name}</b>",
        styles["signature"],
    )
    table = Table([[left, right]], colWidths=[width / 2, width / 2])
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LINEABOVE", (0, 0), (-1, 0), 0, colors.white),
    ]))
    return table


def build_single_curriculum_pdf(row, *, teacher_name: str, principal: str, date_text: str,
                                letterhead_path: str) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = _pdf_styles()
    output = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.55 * cm,
        bottomMargin=0.8 * cm,
        title=f"Silabus Bulanan {row.class_name} {row.month} {row.year}",
        author="TPQ HMarisa",
    )
    available = page_size[0] - doc.leftMargin - doc.rightMargin
    story = []
    kop = _letterhead_flowable(letterhead_path, available)
    if kop:
        story.extend([kop, Spacer(1, 0.33 * cm)])
    story.extend([
        Paragraph("SILABUS BULANAN TPQ HMarisa", styles["title"]),
        Spacer(1, 0.25 * cm),
    ])

    meta_data = [
        [_paragraph("Kelas", styles["cell_bold"]), _paragraph(row.class_name, styles["cell"]),
         _paragraph("Bulan/Tahun", styles["cell_bold"]), _paragraph(f"{row.month} {row.year}", styles["cell"])],
        [_paragraph("Bidang Pelajaran", styles["cell_bold"]), _paragraph(row.subject, styles["cell"]),
         _paragraph("Tahun Ajaran", styles["cell_bold"]), _paragraph(row.academic_year, styles["cell"])],
        [_paragraph("Guru/Wali Kelas", styles["cell_bold"]), _paragraph(teacher_name, styles["cell"]),
         _paragraph("Semester", styles["cell_bold"]), _paragraph(
             "Semester 1" if row.month in MONTH_ORDER[:6] else "Semester 2", styles["cell"])],
        [_paragraph("Status", styles["cell_bold"]), _paragraph(row.status, styles["cell"]),
         _paragraph("Versi", styles["cell_bold"]), _paragraph(f"V{row.version}", styles["cell"])],
    ]
    meta = Table(meta_data, colWidths=[3.4 * cm, 8.0 * cm, 3.3 * cm, available - 14.7 * cm])
    meta.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#AAB8B2")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF4F0")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#EAF4F0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([meta, Spacer(1, 0.32 * cm)])

    headers = ["Pekan", "Materi Pokok", "Target Pembelajaran", "Rencana Materi", "Catatan Guru"]
    data = [[_paragraph(h, styles["cell_bold"]) for h in headers]]
    for index in range(1, 6):
        data.append([
            _paragraph(f"Pekan {index}", styles["cell_bold"]),
            _paragraph(row.topic, styles["cell"]),
            _paragraph(row.learning_target, styles["cell"]),
            _paragraph(getattr(row, f"week{index}", ""), styles["cell"]),
            _paragraph(row.notes, styles["cell"]),
        ])
    widths = [1.8 * cm, 5.1 * cm, 6.2 * cm, 7.2 * cm, available - 20.3 * cm]
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#075F46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#AAB8B2")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBF9")]),
    ]))
    story.extend([
        table,
        Spacer(1, 0.35 * cm),
        KeepTogether(_signature_table(principal, "Guru/Wali Kelas", teacher_name, date_text, available)),
    ])
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    output.seek(0)
    return output


def build_database_pdf(rows: Sequence, *, filter_text: str, principal: str, second_role: str,
                       second_name: str, date_text: str, letterhead_path: str) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = _pdf_styles()
    output = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.55 * cm,
        bottomMargin=0.8 * cm,
        title="Database Silabus Bulanan TPQ HMarisa",
        author="TPQ HMarisa",
    )
    available = page_size[0] - doc.leftMargin - doc.rightMargin
    story = []
    kop = _letterhead_flowable(letterhead_path, available)
    if kop:
        story.extend([kop, Spacer(1, 0.32 * cm)])
    story.extend([
        Paragraph("DATABASE SILABUS BULANAN TPQ HMarisa", styles["title"]),
        Spacer(1, 0.15 * cm),
        Paragraph(filter_text or "Seluruh data silabus", styles["subtitle"]),
        Spacer(1, 0.35 * cm),
    ])
    headers = ["Bulan/Tahun", "Kelas", "Bidang Pelajaran", "Materi Pokok", "Target", "Status", "Versi"]
    data = [[_paragraph(h, styles["cell_bold"]) for h in headers]]
    for row in rows:
        data.append([
            _paragraph(f"{row.month} {row.year}", styles["cell"]),
            _paragraph(row.class_name, styles["cell"]),
            _paragraph(row.subject, styles["cell"]),
            _paragraph(row.topic, styles["cell"]),
            _paragraph(row.learning_target, styles["cell"]),
            _paragraph(row.status, styles["cell"]),
            _paragraph(f"V{row.version}", styles["cell"]),
        ])
    if not rows:
        data.append([_paragraph("Belum ada data pada filter ini", styles["cell"]), "", "", "", "", "", ""])
    widths = [3.1 * cm, 3.2 * cm, 4.7 * cm, 5.4 * cm, 7.5 * cm, 2.3 * cm, available - 26.2 * cm]
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#075F46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#AAB8B2")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBF9")]),
        ("SPAN", (0, 1), (-1, 1)) if not rows else ("LINEBELOW", (0, 0), (-1, 0), 0.7, colors.HexColor("#C49A31")),
    ]))
    story.extend([
        table,
        Spacer(1, 0.55 * cm),
        KeepTogether(_signature_table(principal, second_role, second_name, date_text, available)),
    ])
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    output.seek(0)
    return output


def build_semester_pdf(rows: Sequence, *, class_name: str, academic_year: str, semester: str,
                       teacher_name: str, principal: str, date_text: str,
                       letterhead_path: str) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = _pdf_styles()
    output = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=0.7 * cm,
        rightMargin=0.7 * cm,
        topMargin=0.5 * cm,
        bottomMargin=0.8 * cm,
        title=f"Rekap Silabus {semester} {class_name}",
        author="TPQ HMarisa",
    )
    available = page_size[0] - doc.leftMargin - doc.rightMargin
    story = []
    kop = _letterhead_flowable(letterhead_path, available)
    if kop:
        story.extend([kop, Spacer(1, 0.28 * cm)])
    story.extend([
        Paragraph(f"REKAP SILABUS {semester.upper()} TPQ HMarisa", styles["title"]),
        Spacer(1, 0.12 * cm),
        Paragraph(f"Tahun Ajaran {academic_year} - Kelas {class_name}", styles["subtitle"]),
        Spacer(1, 0.34 * cm),
    ])
    headers = ["No.", "Bulan", "Bidang Pelajaran", "Materi Pokok", "Target Pembelajaran",
               "Rencana Materi", "Status", "Versi"]
    data = [[_paragraph(h, styles["cell_bold"]) for h in headers]]
    for number, row in enumerate(rows, 1):
        data.append([
            _paragraph(number, styles["cell"]),
            _paragraph(f"{row.month} {row.year}", styles["cell"]),
            _paragraph(row.subject, styles["cell"]),
            _paragraph(row.topic, styles["cell"]),
            _paragraph(row.learning_target, styles["cell"]),
            _paragraph(combined_week_plan(row), styles["cell"]),
            _paragraph(row.status, styles["cell"]),
            _paragraph(f"V{row.version}", styles["cell"]),
        ])
    if not rows:
        data.append([_paragraph("Belum ada data pada semester ini", styles["cell"]), "", "", "", "", "", "", ""])
    widths = [1.0 * cm, 2.5 * cm, 3.7 * cm, 4.0 * cm, 5.0 * cm, 7.0 * cm, 2.0 * cm, available - 25.2 * cm]
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#075F46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (1, -1), "CENTER"),
        ("ALIGN", (6, 0), (7, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#AAB8B2")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBF9")]),
        ("SPAN", (0, 1), (-1, 1)) if not rows else ("LINEBELOW", (0, 0), (-1, 0), 0.7, colors.HexColor("#C49A31")),
    ]))
    story.extend([
        table,
        Spacer(1, 0.52 * cm),
        KeepTogether(_signature_table(principal, "Guru/Wali Kelas", teacher_name, date_text, available)),
    ])
    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    output.seek(0)
    return output


def _apply_excel_page_setup(ws, print_area: str | None = None):
    from openpyxl.worksheet.page import PageMargins

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.45, header=0.1, footer=0.2)
    ws.oddFooter.center.text = "Portal TPQ HMarisa - Halaman &P dari &N"
    if print_area:
        ws.print_area = print_area


def _insert_excel_letterhead(ws, letterhead_path: str, end_column: int):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    end_letter = get_column_letter(end_column)
    ws.merge_cells(f"A1:{end_letter}5")
    for row in range(1, 6):
        ws.row_dimensions[row].height = 27
    if letterhead_path and os.path.exists(letterhead_path):
        image = XLImage(letterhead_path)
        original_width = float(image.width or 1)
        original_height = float(image.height or 1)
        image.width = 1120
        image.height = 1120 * original_height / original_width
        ws.add_image(image, "A1")


def _style_excel_title(ws, row: int, start_col: int, end_col: int, text: str, size: int = 16):
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    start = get_column_letter(start_col)
    end = get_column_letter(end_col)
    ws.merge_cells(f"{start}{row}:{end}{row}")
    cell = ws.cell(row, start_col, text)
    cell.font = Font(name="Arial", size=size, bold=True, color="143C30")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = size + 12


def _style_excel_header(ws, row: int, start_col: int, end_col: int):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color=BORDER)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[row].height = 32


def _style_excel_body(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color=BORDER)
    for row in range(start_row, end_row + 1):
        max_chars = 0
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Arial", size=9, color="1F2E28")
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True, indent=1)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBF9")
            max_chars = max(max_chars, len(str(cell.value or "")))
        ws.row_dimensions[row].height = min(max(24, 15 + (max_chars // 45) * 12), 120)


def _excel_signatures(ws, start_row: int, end_col: int, principal: str, second_role: str,
                      second_name: str, date_text: str):
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    middle = max(2, end_col // 2)
    left_end = middle
    right_start = middle + 1
    end_letter = get_column_letter(end_col)
    left_end_letter = get_column_letter(left_end)
    right_start_letter = get_column_letter(right_start)
    ws.merge_cells(f"A{start_row}:{left_end_letter}{start_row + 5}")
    ws.merge_cells(f"{right_start_letter}{start_row}:{end_letter}{start_row + 5}")
    left = ws.cell(start_row, 1, f"Mengetahui,\nKepala TPQ HMarisa\n\n\n{principal}")
    right = ws.cell(start_row, right_start, f"Tangerang Selatan, {date_text}\n{second_role}\n\n\n{second_name}")
    for cell in (left, right):
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row in range(start_row, start_row + 6):
        ws.row_dimensions[row].height = 20


def build_database_excel(rows: Sequence, *, filter_text: str, principal: str, second_role: str,
                         second_name: str, date_text: str, letterhead_path: str) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Database Silabus"
    _insert_excel_letterhead(ws, letterhead_path, 7)
    _style_excel_title(ws, 7, 1, 7, "DATABASE SILABUS BULANAN TPQ HMarisa")
    ws.merge_cells("A8:G8")
    ws["A8"] = filter_text or "Seluruh data silabus"
    ws["A8"].font = Font(name="Arial", size=9, italic=True, color="596B63")
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[8].height = 25
    header_row = 10
    headers = ["Bulan/Tahun", "Kelas", "Bidang Pelajaran", "Materi Pokok", "Target", "Status", "Versi"]
    for col, header in enumerate(headers, 1):
        ws.cell(header_row, col, header)
    _style_excel_header(ws, header_row, 1, 7)
    row_number = header_row + 1
    for row in rows:
        values = [f"{row.month} {row.year}", row.class_name, row.subject, row.topic,
                  display_value(row.learning_target), row.status, f"V{row.version}"]
        for col, value in enumerate(values, 1):
            ws.cell(row_number, col, value)
        row_number += 1
    if not rows:
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=7)
        ws.cell(row_number, 1, "Belum ada data pada filter ini")
        ws.cell(row_number, 1).alignment = Alignment(horizontal="center")
        row_number += 1
    _style_excel_body(ws, header_row + 1, row_number - 1, 1, 7)
    widths = {"A": 18, "B": 18, "C": 28, "D": 34, "E": 55, "F": 14, "G": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A{header_row}:G{max(header_row, row_number - 1)}"
    ws.freeze_panes = f"A{header_row + 1}"
    signature_row = row_number + 2
    _excel_signatures(ws, signature_row, 7, principal, second_role, second_name, date_text)
    _apply_excel_page_setup(ws, f"A1:G{signature_row + 5}")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_semester_excel(rows: Sequence, *, class_name: str, academic_year: str, semester: str,
                         teacher_name: str, principal: str, date_text: str,
                         letterhead_path: str, summary: Sequence[Mapping],
                         completeness: Sequence[Mapping]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Ringkasan Semester"
    _insert_excel_letterhead(summary_ws, letterhead_path, 3)
    _style_excel_title(summary_ws, 7, 1, 3, f"REKAP SILABUS {semester.upper()} TPQ HMarisa")
    summary_ws.merge_cells("A8:C8")
    summary_ws["A8"] = f"Tahun Ajaran {academic_year} - Kelas {class_name}"
    summary_ws["A8"].alignment = Alignment(horizontal="center")
    summary_ws["A8"].font = Font(name="Arial", size=10, bold=True, color="596B63")
    summary_ws["A10"] = "Bidang Pelajaran"
    summary_ws["B10"] = "Jumlah Materi"
    summary_ws["C10"] = "Bulan Terisi"
    _style_excel_header(summary_ws, 10, 1, 3)
    row_no = 11
    for item in summary:
        summary_ws.cell(row_no, 1, item.get("subject"))
        summary_ws.cell(row_no, 2, item.get("count", 0))
        summary_ws.cell(row_no, 3, item.get("months", "Belum diisi"))
        row_no += 1
    _style_excel_body(summary_ws, 11, max(11, row_no - 1), 1, 3)
    row_no += 2
    summary_ws.cell(row_no, 1, "Pemeriksaan Kelengkapan")
    summary_ws.cell(row_no, 1).font = Font(name="Arial", size=12, bold=True, color="143C30")
    row_no += 1
    for col, header in enumerate(["Bulan", "Status", "Bidang Pelajaran Belum Diisi"], 1):
        summary_ws.cell(row_no, col, header)
    _style_excel_header(summary_ws, row_no, 1, 3)
    check_start = row_no + 1
    row_no += 1
    for item in completeness:
        summary_ws.cell(row_no, 1, item.get("month"))
        summary_ws.cell(row_no, 2, item.get("status"))
        summary_ws.cell(row_no, 3, item.get("missing_text"))
        if item.get("status") != "Lengkap":
            summary_ws.cell(row_no, 2).fill = PatternFill("solid", fgColor="FFF0ED")
        row_no += 1
    _style_excel_body(summary_ws, check_start, max(check_start, row_no - 1), 1, 3)
    for col, width in {"A": 42, "B": 18, "C": 65}.items():
        summary_ws.column_dimensions[col].width = width
    _excel_signatures(summary_ws, row_no + 2, 3, principal, "Guru/Wali Kelas", teacher_name, date_text)
    _apply_excel_page_setup(summary_ws, f"A1:C{row_no + 7}")

    detail = wb.create_sheet("Detail Materi")
    _insert_excel_letterhead(detail, letterhead_path, 13)
    _style_excel_title(detail, 7, 1, 13, f"DETAIL SILABUS {semester.upper()} - {class_name}")
    detail.merge_cells("A8:M8")
    detail["A8"] = f"Tahun Ajaran {academic_year}"
    detail["A8"].alignment = Alignment(horizontal="center")
    detail["A8"].font = Font(name="Arial", size=10, bold=True, color="596B63")
    header_row = 10
    headers = ["Bulan", "Kelas", "Bidang Pelajaran", "Materi Pokok", "Target Pembelajaran",
               "Rencana Materi Pekan 1", "Rencana Materi Pekan 2", "Rencana Materi Pekan 3",
               "Rencana Materi Pekan 4", "Rencana Materi Pekan 5", "Catatan Guru", "Status", "Versi"]
    for col, header in enumerate(headers, 1):
        detail.cell(header_row, col, header)
    _style_excel_header(detail, header_row, 1, 13)
    current = header_row + 1
    for row in rows:
        values = [f"{row.month} {row.year}", row.class_name, row.subject, row.topic,
                  display_value(row.learning_target), display_value(row.week1), display_value(row.week2),
                  display_value(row.week3), display_value(row.week4), display_value(row.week5),
                  display_value(row.notes), row.status, f"V{row.version}"]
        for col, value in enumerate(values, 1):
            detail.cell(current, col, value)
        current += 1
    if not rows:
        detail.merge_cells(start_row=current, start_column=1, end_row=current, end_column=13)
        detail.cell(current, 1, "Belum ada data pada semester ini")
        detail.cell(current, 1).alignment = Alignment(horizontal="center")
        current += 1
    _style_excel_body(detail, header_row + 1, current - 1, 1, 13)
    widths = [18, 18, 28, 32, 42, 32, 32, 32, 32, 32, 34, 14, 10]
    for index, width in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter
        detail.column_dimensions[get_column_letter(index)].width = width
    detail.freeze_panes = f"A{header_row + 1}"
    detail.auto_filter.ref = f"A{header_row}:M{max(header_row, current - 1)}"
    _excel_signatures(detail, current + 2, 13, principal, "Guru/Wali Kelas", teacher_name, date_text)
    _apply_excel_page_setup(detail, f"A1:M{current + 7}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_import_template(*, classes: Sequence[str], subjects_map: Mapping[str, Sequence[str]],
                          academic_years: Sequence[str], months: Sequence[str]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, quote_sheetname
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Silabus"
    headers = [
        "Kelas", "Bulan", "Tahun Ajaran", "Semester", "Bidang Pelajaran", "Materi Pokok",
        "Target Pembelajaran", "Rencana Materi Pekan 1", "Rencana Materi Pekan 2",
        "Rencana Materi Pekan 3", "Rencana Materi Pekan 4", "Rencana Materi Pekan 5",
        "Catatan Guru", "Status", "Versi",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    _style_excel_header(ws, 1, 1, len(headers))
    widths = [20, 16, 18, 16, 36, 38, 48, 36, 36, 36, 36, 36, 40, 14, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O501"
    thin = Side(style="thin", color="D7E1DC")
    for row in range(2, 502):
        ws.cell(row, 4, f'=IF(B{row}="","",IF(OR(B{row}="Juli",B{row}="Agustus",B{row}="September",B{row}="Oktober",B{row}="November",B{row}="Desember"),"Semester 1","Semester 2"))')
        ws.cell(row, 14, "Aktif")
        ws.cell(row, 15, 1)
        for col in range(1, 16):
            cell = ws.cell(row, col)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[row].height = 26
    ws.conditional_formatting.add(
        "A2:O501",
        FormulaRule(formula=['AND($A2<>"",OR($B2="",$C2="",$E2="",$F2=""))'],
                    fill=PatternFill("solid", fgColor="FFF0ED")),
    )

    refs = wb.create_sheet("Referensi")
    refs.sheet_state = "hidden"
    refs["A1"] = "Kelas"
    for idx, value in enumerate(classes, 2):
        refs.cell(idx, 1, value)
    refs["B1"] = "Bulan"
    for idx, value in enumerate(months, 2):
        refs.cell(idx, 2, value)
    refs["C1"] = "Tahun Ajaran"
    for idx, value in enumerate(academic_years, 2):
        refs.cell(idx, 3, value)
    refs["D1"] = "Status"
    refs["D2"] = "Aktif"
    refs["D3"] = "Nonaktif"

    # Daftar bidang pelajaran per kelas dalam named range.
    for offset, class_name in enumerate(classes, 5):
        refs.cell(1, offset, class_name)
        subjects = list(subjects_map.get(class_name, []))
        for row, value in enumerate(subjects, 2):
            refs.cell(row, offset, value)
        range_name = "SUBJECT_" + re.sub(r"[^A-Z0-9_]", "_", class_name.upper().replace(" ", "_"))
        col_letter = get_column_letter(offset)
        target = f"{quote_sheetname(refs.title)}!${col_letter}$2:${col_letter}${max(2, len(subjects)+1)}"
        wb.defined_names.add(DefinedName(range_name, attr_text=target))

    class_dv = DataValidation(type="list", formula1=f"={quote_sheetname(refs.title)}!$A$2:$A${len(classes)+1}", allow_blank=True)
    month_dv = DataValidation(type="list", formula1=f"={quote_sheetname(refs.title)}!$B$2:$B${len(months)+1}", allow_blank=True)
    year_dv = DataValidation(type="list", formula1=f"={quote_sheetname(refs.title)}!$C$2:$C${len(academic_years)+1}", allow_blank=True)
    status_dv = DataValidation(type="list", formula1=f"={quote_sheetname(refs.title)}!$D$2:$D$3", allow_blank=True)
    subject_formula = '=INDIRECT("SUBJECT_"&SUBSTITUTE(SUBSTITUTE(UPPER($A2)," ","_"),"-","_"))'
    subject_dv = DataValidation(type="list", formula1=subject_formula, allow_blank=True)
    for dv in [class_dv, month_dv, year_dv, status_dv, subject_dv]:
        ws.add_data_validation(dv)
    class_dv.add("A2:A501")
    month_dv.add("B2:B501")
    year_dv.add("C2:C501")
    subject_dv.add("E2:E501")
    status_dv.add("N2:N501")

    guide = wb.create_sheet("Petunjuk")
    guide.column_dimensions["A"].width = 6
    guide.column_dimensions["B"].width = 115
    guide["A1"] = "No."
    guide["B1"] = "Petunjuk Pengisian Template Import Silabus Bulanan"
    _style_excel_header(guide, 1, 1, 2)
    instructions = [
        "Satu baris mewakili satu bidang pelajaran untuk satu kelas dan satu bulan.",
        "Isi Kelas, Bulan, Tahun Ajaran, Bidang Pelajaran, Materi Pokok, dan minimal satu Rencana Materi Pekan.",
        "Semester terisi otomatis berdasarkan bulan: Juli-Desember = Semester 1; Januari-Juni = Semester 2.",
        "Bidang Pelajaran harus sesuai dengan Kelas pada Data Master.",
        "Kolom yang kosong akan ditampilkan sebagai 'Belum diisi', bukan tanda minus atau angka 0.",
        "Status dapat diisi Aktif atau Nonaktif. Versi harus berupa angka minimal 1.",
        "Jangan mengubah nama header pada baris pertama.",
        "Sebelum data disimpan, portal akan menampilkan pratinjau dan daftar kesalahan per baris.",
    ]
    for idx, text in enumerate(instructions, 1):
        guide.cell(idx + 1, 1, idx)
        guide.cell(idx + 1, 2, text)
        guide.cell(idx + 1, 2).alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[idx + 1].height = 32
    _style_excel_body(guide, 2, len(instructions) + 1, 1, 2)
    guide.sheet_view.showGridLines = False
    ws.sheet_view.showGridLines = False
    _apply_excel_page_setup(ws, "A1:O60")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
